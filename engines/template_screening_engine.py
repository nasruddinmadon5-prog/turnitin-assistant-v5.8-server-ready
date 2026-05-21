"""
Template Screening Engine v5.7.1
DOCX-first structure screening with AI as MULTI-SOURCE EVIDENCE EVALUATOR.

v5.7.1 changes (on top of v5.7):
- AI is now a MULTI-EVIDENCE CONSISTENCY EVALUATOR, not a single-source detector
- AI receives BOTH DOC snippet AND PDF snippet simultaneously
- New AI prompt compares DOC vs PDF evidence to identify extraction/layout failures
- New AI JSON response format: exists, confidence, preferred_source, conflict_reason,
  likely_issue, reason, evidence_summary
- AI reasoning rules:
    DOC YES + PDF NO => AI leans YES, likely_issue = pdf_extraction_failure
    DOC NO + PDF YES => AI leans YES, likely_issue = doc_extraction_failure
    DOC YES + PDF YES => AI usually skipped (no conflict)
    DOC NO + PDF NO  => AI checks weak semantic evidence
- build_ai_verifier_input() now always includes BOTH doc and pdf evidence snippets
- Caching and rate limiting preserved from v5.7
- Architecture unchanged from v5.7

v5.7 changes:
- AI now participates early: runs when DOC/PDF conflict, not only when both FAILED
- evidence_sources list replaces single final_source_used
- New manuscript-level status: PASS / NEEDS MANUAL REVIEW / REJECT
- Per-item voting logic: DOC + PDF + AI = explainable screening
- AI rate limiting + anti-spam: delay, cache, max calls per submission/run
- Model switching: text model for sections, vision model for graphical abstract
- Graphical abstract vision check with cropped image
- Privacy: full redaction before sending to AI
- Organized file storage: json/, pdf/, combined/, crops/, ai_cache/, pass/, needs_review/, reject/
- Excel columns: final_review_status, doc_detect_summary, pdf_detect_summary, etc.

Part of Turnitin Assistant v5.7.1
"""

import json
import re
import subprocess
import sys
import os
import shutil
import zipfile
import xml.etree.ElementTree as ET
import hashlib
import time
from pathlib import Path
from datetime import datetime

try:
    import fitz  # PyMuPDF
except ImportError:
    fitz = None

# ── v5.7: Import config for AI verifier flags, rate limits, model switching ──
try:
    from config import (
        USE_AI_VERIFIER, OPENAI_API_KEY,
        AI_TEXT_MODEL, AI_VISION_MODEL,
        AI_REQUEST_DELAY_SECONDS, AI_MAX_CALLS_PER_SUBMISSION,
        AI_MAX_CALLS_PER_RUN, AI_TIMEOUT_SECONDS, AI_CACHE_ENABLED,
        AI_VERIFIER_MAX_SNIPPET_CHARS,
        TEMPLATE_SCREENING_CONFIG,
    )
except ImportError:
    USE_AI_VERIFIER = False
    OPENAI_API_KEY = ""
    AI_TEXT_MODEL = "gpt-4.1-mini"
    AI_VISION_MODEL = "gpt-4.1-mini"
    AI_REQUEST_DELAY_SECONDS = 2.0
    AI_MAX_CALLS_PER_SUBMISSION = 8
    AI_MAX_CALLS_PER_RUN = 80
    AI_TIMEOUT_SECONDS = 30
    AI_CACHE_ENABLED = True
    AI_VERIFIER_MAX_SNIPPET_CHARS = 800
    TEMPLATE_SCREENING_CONFIG = {"reports_dir": ""}


# ── v5.7: Global AI call counters for rate limiting ──
_ai_call_count_per_submission = 0
_ai_call_count_per_run = 0
_ai_last_call_time = 0.0


# ── v5.7: AI Cache ──
def _get_ai_cache_dir():
    """Get the AI cache directory path."""
    reports_dir = TEMPLATE_SCREENING_CONFIG.get("reports_dir", "")
    if reports_dir:
        cache_dir = Path(reports_dir) / "ai_cache"
        cache_dir.mkdir(parents=True, exist_ok=True)
        return cache_dir
    # Fallback
    cache_dir = Path("storage/template_screening_reports/ai_cache")
    cache_dir.mkdir(parents=True, exist_ok=True)
    return cache_dir


def _compute_ai_cache_key(item_id, redacted_snippet):
    """Compute a deterministic cache key for an AI verification request."""
    raw = f"{item_id}:::{redacted_snippet}"
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _get_ai_cache(cache_key):
    """Retrieve cached AI result if available."""
    if not AI_CACHE_ENABLED:
        return None
    cache_dir = _get_ai_cache_dir()
    cache_file = cache_dir / f"{cache_key}.json"
    if cache_file.exists():
        try:
            with open(cache_file, "r", encoding="utf-8") as f:
                cached = json.load(f)
            # Check expiry (24 hours)
            cached_time = cached.get("cached_at", 0)
            if time.time() - cached_time < 86400:  # 24 hours
                return cached.get("result")
        except Exception:
            pass
    return None


def _set_ai_cache(cache_key, result):
    """Store AI result in cache."""
    if not AI_CACHE_ENABLED:
        return
    cache_dir = _get_ai_cache_dir()
    cache_file = cache_dir / f"{cache_key}.json"
    try:
        with open(cache_file, "w", encoding="utf-8") as f:
            json.dump({
                "cached_at": time.time(),
                "result": result,
            }, f, ensure_ascii=False)
    except Exception:
        pass


def _check_ai_rate_limit():
    """Check and enforce AI rate limits. Returns True if allowed, False if rate limited."""
    global _ai_call_count_per_submission, _ai_call_count_per_run, _ai_last_call_time

    if _ai_call_count_per_submission >= AI_MAX_CALLS_PER_SUBMISSION:
        return False
    if _ai_call_count_per_run >= AI_MAX_CALLS_PER_RUN:
        return False

    # Enforce delay between calls
    if _ai_last_call_time > 0:
        elapsed = time.time() - _ai_last_call_time
        if elapsed < AI_REQUEST_DELAY_SECONDS:
            time.sleep(AI_REQUEST_DELAY_SECONDS - elapsed)

    _ai_call_count_per_submission += 1
    _ai_call_count_per_run += 1
    _ai_last_call_time = time.time()
    return True


def _reset_ai_rate_limits():
    """Reset AI rate limit counters (called per submission)."""
    global _ai_call_count_per_submission
    _ai_call_count_per_submission = 0


def _reset_ai_run_limits():
    """Reset AI per-run counters (called at start of workflow)."""
    global _ai_call_count_per_run, _ai_last_call_time
    _ai_call_count_per_run = 0
    _ai_last_call_time = 0.0


# ── v5.7: Privacy redaction ──
def redact_sensitive(text):
    """
    Redact sensitive information from text before sending to AI.
    Never sends: emails, phone, DOI, URL, affiliations, funding, author names.
    """
    if not text:
        return ""
    redacted = text
    # Email addresses
    redacted = re.sub(r'[\w.+-]+@[\w-]+\.[\w.-]+', '[EMAIL]', redacted)
    # Phone numbers
    redacted = re.sub(r'\b(\+?\d{1,3}[-.\s]?)?\(?\d{2,4}\)?[-.\s]?\d{3,4}[-.\s]?\d{3,4}\b', '[PHONE]', redacted)
    # DOI
    redacted = re.sub(r'\b10\.\d{4,}/[^\s,;)]+', '[DOI]', redacted)
    # URLs
    redacted = re.sub(r'https?://[^\s,;)]+', '[URL]', redacted)
    # ORCID
    redacted = re.sub(r'\b0000-000[12]-?\d{4}-?\d{3}[0-9X]\b', '[ORCID]', redacted)
    # Affiliation-like lines (containing university, faculty, department, etc.)
    affiliation_patterns = [
        r'(?i)(?:^|\n)\s*(?:universit\w+|faculty|school\s+of|department\s+of|institute|centre|center|college)\s+[^\n]{3,100}',
    ]
    for pat in affiliation_patterns:
        redacted = re.sub(pat, '\n[AFFILIATION]', redacted)
    # Remove funding/acknowledgement grant numbers
    redacted = re.sub(r'(?i)(grant|funding|award)\s*(nos?|numbers?)?\s*[:\-]?\s*[A-Z0-9\-]{5,30}', '[GRANT_NUMBER]', redacted)
    # Remove acknowledgement lines
    redacted = re.sub(r'(?i)(?:^|\n)\s*acknowledg[^\n]{10,300}', '\n[ACKNOWLEDGEMENT_REMOVED]', redacted)
    # Truncate to max chars
    redacted = redacted[:AI_VERIFIER_MAX_SNIPPET_CHARS]
    return redacted


# ─── Global Exclusion Labels ──────────────────────────────────────────────────
# These are template/header labels that should never be accepted as manuscript
# metadata (title, authors, affiliations, etc.)
EXCLUSION_LABELS = [
    "jurnal teknologi",
    "full paper",
    "article history",
    "received",
    "accepted",
    "corresponding author",
    "graphical abstract",
    "abstract",
    "keywords",
    "abstrak",
    "kata kunci",
    "eissn",
    "penerbit utm press",
    "utm press",
    "www.jurnalteknologi.utm.my",
    "jurnalteknologi",
    "doi",
    "published",
    "references",
    "introduction",
    "methodology",
    "results and discussion",
    "conclusion",
    "acknowledgement",
]


def is_false_positive(text):
    """
    Legacy false positive check using global EXCLUSION_LABELS.
    Kept for backward compatibility but should NOT be used for section/component items.
    """
    if not text:
        return True

    text_clean = text.strip().lower()
    if not text_clean:
        return True

    # Direct match
    for label in EXCLUSION_LABELS:
        if text_clean == label:
            return True

    # Check if text is mostly (>=70%) comprised of exclusion label words
    text_words = set(text_clean.split())
    if len(text_words) <= 3:
        for label in EXCLUSION_LABELS:
            label_words = set(label.split())
            overlap = text_words & label_words
            if len(overlap) >= min(len(text_words), len(label_words)):
                return True

    return False


# ─── Item-Aware False Positive Detection ─────────────────────────────────────

def is_false_positive_for_item(text, item_id):
    """
    Item-aware false positive check.
    Only rejects template labels that are NOT the valid heading for the given item.

    - For metadata items (title, authors, affiliations):
      Reject known template labels like Jurnal Teknologi, Full Paper, etc.

    - For section/component items:
      DO NOT reject their own valid heading label.
      e.g. item_id="abstract" and text="Abstract" => NOT a false positive
    """
    if not text:
        return True

    text_clean = text.strip().lower()
    if not text_clean:
        return True

    # ── Metadata items: stricter filtering ──
    metadata_items = {"title", "authors", "affiliations"}
    if item_id in metadata_items:
        return is_false_positive(text)  # Use full global exclusion list

    # ── Section/component items: allow their own heading ──
    # Define which label patterns are VALID (NOT false positive) for each item
    own_label_patterns = {
        "corresponding_author": [
            re.compile(r'^\*?\s*corresponding\s+author', re.IGNORECASE),
            re.compile(r'^corresponding\s+author', re.IGNORECASE),
        ],
        "graphical_abstract": [
            re.compile(r'^graphical\s+abstract', re.IGNORECASE),
        ],
        "abstract": [
            re.compile(r'^abstract', re.IGNORECASE),
        ],
        "keywords": [
            re.compile(r'^keywords?', re.IGNORECASE),
        ],
        "abstrak": [
            re.compile(r'^abstrak', re.IGNORECASE),
        ],
        "kata_kunci": [
            re.compile(r'^kata\s+kunci', re.IGNORECASE),
        ],
        "introduction": [
            re.compile(r'^1\.?\s*introduction', re.IGNORECASE),
            re.compile(r'^introduction', re.IGNORECASE),
        ],
        "methodology": [
            re.compile(r'^2\.?\s*method', re.IGNORECASE),
            re.compile(r'^methodology', re.IGNORECASE),
            re.compile(r'^method', re.IGNORECASE),
        ],
        "results_discussion": [
            re.compile(r'^3\.?\s*result', re.IGNORECASE),
            re.compile(r'^results?\s+and?\s+discussion', re.IGNORECASE),
            re.compile(r'^results?', re.IGNORECASE),
        ],
        "conclusion": [
            re.compile(r'^4\.?\s*conclusion', re.IGNORECASE),
            re.compile(r'^conclusion', re.IGNORECASE),
        ],
        "acknowledgement": [
            re.compile(r'^acknowledgement', re.IGNORECASE),
            re.compile(r'^acknowledgments?', re.IGNORECASE),
        ],
        "references": [
            re.compile(r'^references?', re.IGNORECASE),
        ],
    }

    # If this item has own-label patterns, check if text matches them
    if item_id in own_label_patterns:
        for pattern in own_label_patterns[item_id]:
            if pattern.match(text_clean):
                return False  # This is the item's own valid heading — NOT false positive

    # For section items not matching own label, apply standard exclusion
    return is_false_positive(text)


# ─── Content Verification Helpers ────────────────────────────────────────────

def extract_nearby_text_after_heading(page_text, heading_match, max_chars=1200):
    """
    Extract text content found after a heading match within the same page.
    heading_match: a regex match object or the matched string.
    max_chars: maximum characters to extract after the heading.

    Returns the extracted text snippet.
    """
    if isinstance(heading_match, str):
        # Find the position of the heading string in page text
        idx = page_text.lower().find(heading_match.lower())
        if idx == -1:
            return ""
        start = idx + len(heading_match)
    else:
        start = heading_match.end()

    # Extract up to max_chars after the heading
    nearby = page_text[start:start + max_chars].strip()

    # Try to truncate at a natural boundary (next heading or double newline)
    # Look for patterns like "\n\n", "\n[A-Z][A-Za-z ]+\n", or numbered sections
    break_pattern = re.compile(r'\n\s*\n|\n(?=[A-Z][A-Za-z\s]{2,50}\n)|(?=\n\d+\.\s+[A-Z])')
    break_match = break_pattern.search(nearby)
    if break_match:
        nearby = nearby[:break_match.start()]

    return nearby.strip()


def looks_like_abstract_content(text):
    """
    Check if text looks like abstract content.
    Valid if text has approximately 50–400 words and contains research-style language.
    """
    if not text:
        return False

    words = text.split()
    word_count = len(words)

    if word_count < 30:
        return False
    if word_count > 600:
        return False

    # Research-style language indicators (English)
    en_indicators = [
        "study", "research", "objective", "method", "result", "finding",
        "analysis", "approach", "aim", "purpose", "investigat", "evaluat",
        "assess", "demonstrat", "show", "conclude", "significant",
        "propose", "develop", "framework", "data", "sample", "experiment",
    ]
    # Research-style language indicators (Bahasa Melayu)
    ms_indicators = [
        "kajian", "objektif", "kaedah", "dapatan", "hasil", "analisis",
        "pendekatan", "tujuan", "menunjuk", "kesimpulan", "data",
        "sampel", "eksperimen", "penyelidikan", "perbincangan",
    ]

    text_lower = text.lower()
    en_count = sum(1 for word in en_indicators if word in text_lower)
    ms_count = sum(1 for word in ms_indicators if word in text_lower)

    # At least 2 research indicators should be present
    return (en_count + ms_count) >= 2


def looks_like_keywords_content(text):
    """
    Check if text looks like keywords content.
    Valid if contains comma/semicolon-separated terms or 3–8 short keyword phrases.
    """
    if not text:
        return False

    # Remove the label itself if present
    text_clean = re.sub(r'^(Keywords|Kata kunci|Kata Kunci)\s*[:\-–]?\s*', '', text, flags=re.IGNORECASE).strip()
    if not text_clean:
        return False

    # Split by comma, semicolon, or bullet
    parts = re.split(r'[;,•·\-•\n]+', text_clean)
    parts = [p.strip() for p in parts if p.strip() and len(p.strip()) > 1]

    # Should have 2–12 keyword phrases
    if len(parts) < 2 or len(parts) > 15:
        return False

    # Each part should be relatively short (under 10 words typically)
    long_parts = sum(1 for p in parts if len(p.split()) > 10)
    if long_parts > len(parts) // 2:
        return False

    return True


def looks_like_references_content(text):
    """
    Check if text looks like references content.
    Valid if contains multiple citation-like entries, years, author names,
    journal/book indicators, DOI, volume/issue, etc.
    """
    if not text:
        return False

    lines = text.split('\n')
    lines = [l.strip() for l in lines if l.strip()]
    if len(lines) < 2:
        return False

    # Count citation indicators
    citation_count = 0
    for line in lines[:30]:  # Check first 30 lines
        # Has year (19xx or 20xx)
        if re.search(r'\b(19|20)\d{2}\b', line):
            citation_count += 1
        # Starts with [number] or number.
        elif re.match(r'^\[?\d+\]?[\s.]', line):
            citation_count += 1
        # Contains DOI
        elif re.search(r'\b10\.\d{4,}/', line, re.IGNORECASE):
            citation_count += 1
        # Contains author-like pattern with journal indicators
        elif re.search(r'\b(Journal|Proceedings|Conference|Review|Letters|Research|Science)\b', line, re.IGNORECASE):
            citation_count += 1

    return citation_count >= 2


def looks_like_corresponding_author_content(text):
    """
    Check if text contains corresponding author contact info.
    Valid if nearby text contains email or author contact pattern.
    """
    if not text:
        return False

    # Email pattern
    if re.search(r'[\w.+-]+@[\w-]+\.[\w.-]+', text):
        return True

    # Phone/fax pattern
    if re.search(r'\b(\+?\d{1,3}[-.\s]?)?\(?\d{2,4}\)?[-.\s]?\d{3,4}[-.\s]?\d{3,4}\b', text):
        return True

    # Postal address indicators
    address_indicators = ["address", "tel", "fax", "email", "correspondence",
                          "orcid", "universit", "institute", "department"]
    text_lower = text.lower()
    if sum(1 for ind in address_indicators if ind in text_lower) >= 2:
        return True

    return False


def looks_like_graphical_abstract_content(page_data, region_bbox):
    """
    Check if the region contains graphical abstract content.
    Uses image detection, drawing detection, or figure-like content.
    """
    detected, method, confidence = detect_graphical_region(page_data, region_bbox)
    return detected


def looks_like_section_content(text, item_id):
    """
    Generic section content checker.
    Delegates to specific content checkers based on item_id.
    """
    if not text:
        return False

    content_checkers = {
        "abstract": looks_like_abstract_content,
        "abstrak": looks_like_abstract_content,
        "keywords": looks_like_keywords_content,
        "kata_kunci": looks_like_keywords_content,
        "references": looks_like_references_content,
    }

    checker = content_checkers.get(item_id)
    if checker:
        return checker(text)
    return True  # For generic sections (introduction, methodology, etc.), assume content is valid


# ═══════════════════════════════════════════════════════════════════════════════
# v3.0 NEW: DOCX Source Text Extraction
# ═══════════════════════════════════════════════════════════════════════════════

def extract_source_text(source_file):
    """
    Extract raw text from the original source file.
    
    For .docx: reads word/document.xml from within the ZIP archive.
    For .doc: tries LibreOffice headless conversion (docx > txt > pdf fallback).
    For .pdf: returns None (PDF text extraction is handled separately).
    
    Returns:
        str or None: The extracted source text, or None if extraction fails.
    """
    if not source_file or not Path(source_file).exists():
        return None
    
    source_path = Path(source_file)
    ext = source_path.suffix.lower()
    
    try:
        if ext == '.docx':
            return _extract_docx_text(source_path)
        elif ext == '.doc':
            return _extract_doc_text(source_path)
        else:
            return None
    except Exception as e:
        print(f"[SOURCE_TEXT] Error extracting from {source_path.name}: {e}")
        return None


def _extract_doc_text(doc_path):
    """
    Extract text from a .doc file using LibreOffice headless conversion.
    Tries: docx > txt > pdf fallback. Returns None if all fail.
    """
    doc_path = Path(doc_path)
    output_dir = doc_path.parent

    # Method 1: DOC to DOCX conversion
    try:
        cmd = ["soffice", "--headless", "--convert-to", "docx",
               "--outdir", str(output_dir), str(doc_path)]
        subprocess.run(cmd, capture_output=True, text=True, timeout=120)
        possible = sorted(output_dir.glob(f"{doc_path.stem}.docx"))
        if possible:
            text = _extract_docx_text(possible[0])
            try:
                possible[0].unlink()
            except Exception:
                pass
            if text:
                return text
    except Exception:
        pass

    # Method 2: DOC to TXT conversion
    try:
        cmd = ["soffice", "--headless", "--convert-to", "txt:Text",
               "--outdir", str(output_dir), str(doc_path)]
        subprocess.run(cmd, capture_output=True, text=True, timeout=120)
        possible = sorted(output_dir.glob(f"{doc_path.stem}.txt"))
        if possible:
            with open(possible[0], 'r', encoding='utf-8', errors='replace') as f:
                text = f.read()
            try:
                possible[0].unlink()
            except Exception:
                pass
            if text and len(text.strip()) > 50:
                return text
    except Exception:
        pass

    # Method 3: DOC to PDF fallback
    try:
        cmd = ["soffice", "--headless", "--convert-to", "pdf",
               "--outdir", str(output_dir), str(doc_path)]
        subprocess.run(cmd, capture_output=True, text=True, timeout=120)
        possible = sorted(output_dir.glob(f"{doc_path.stem}.pdf"))
        if possible and fitz:
            doc = fitz.open(str(possible[0]))
            text = "\n".join(page.get_text() for page in doc)
            doc.close()
            try:
                possible[0].unlink()
            except Exception:
                pass
            if text and len(text.strip()) > 50:
                return text
    except Exception:
        pass

    return None


def _extract_docx_text(docx_path):
    """
    Extract text from a .docx file by parsing word/document.xml.
    
    DOCX is a ZIP archive containing XML files. The main document
    content is in word/document.xml. Text is stored in <w:t> elements
    within <w:p> (paragraph) elements.
    
    Returns:
        str: Full text content with paragraphs separated by newlines.
    """
    try:
        with zipfile.ZipFile(docx_path, 'r') as z:
            # Try main document first
            xml_paths = ['word/document.xml', 'word/document2.xml']
            xml_content = None
            
            for xp in xml_paths:
                try:
                    xml_content = z.read(xp)
                    break
                except KeyError:
                    continue
            
            if xml_content is None:
                # Try to find any document XML
                xml_files = [f for f in z.namelist() if f.startswith('word/') and f.endswith('.xml')]
                if xml_files:
                    # Use the first document-like XML
                    doc_candidates = [f for f in xml_files if 'document' in f.lower()]
                    if doc_candidates:
                        xml_content = z.read(doc_candidates[0])
                
                if xml_content is None:
                    return None
            
            # Parse XML and extract text from <w:t> elements
            # Register the Word namespace
            namespaces = {
                'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main',
            }
            
            root = ET.fromstring(xml_content)
            
            # Find all <w:p> elements (paragraphs)
            paragraphs = root.findall('.//w:p', namespaces)
            
            text_parts = []
            for para in paragraphs:
                # Find all <w:t> elements within this paragraph
                text_elements = para.findall('.//w:t', namespaces)
                para_text = ''.join(elem.text or '' for elem in text_elements).strip()
                if para_text:
                    text_parts.append(para_text)
            
            if not text_parts:
                # Fallback: extract all text from all w:t elements
                all_texts = root.findall('.//w:t', namespaces)
                text_parts = [elem.text or '' for elem in all_texts if elem.text and elem.text.strip()]
            
            return '\n'.join(text_parts)
            
    except zipfile.BadZipFile:
        print(f"[SOURCE_TEXT] {docx_path} is not a valid ZIP/DOCX file")
        return None
    except Exception as e:
        print(f"[SOURCE_TEXT] Error parsing DOCX {docx_path}: {e}")
        return None


# ═══════════════════════════════════════════════════════════════════════════════
# v3.0 NEW: DOCX Source Text Section Detector
# ═══════════════════════════════════════════════════════════════════════════════

def source_text_section_detector(source_text, rules):
    """
    Detect manuscript sections in source text using regex patterns.
    
    Uses the same rule definitions as the PDF detector but operates
    on raw DOCX source text. This is more reliable because DOCX text
    extraction preserves structure that PDF conversion may distort.
    
    Each rule's text_pattern is applied against the source text.
    For section items (abstract, keywords, etc.), we extract the
    heading match and nearby content for verification.
    
    Returns:
        dict: {
            "items": {
                "abstract": {"heading_found": True, "evidence": "Abstract...", "nearby": "..."},
                ...
            },
            "full_text": str  # the source text
        }
    """
    if not source_text:
        return {"items": {}, "full_text": ""}
    
    detected = {}
    
    # Section items that can be detected in source text
    section_item_ids = {
        "abstract", "keywords", "abstrak", "kata_kunci",
        "introduction", "methodology", "results_discussion",
        "conclusion", "acknowledgement", "references",
        "corresponding_author", "graphical_abstract",
    }
    
    for rule in rules:
        item_id = rule.get("id")
        if item_id not in section_item_ids:
            continue
        
        text_pattern = rule.get("text_pattern", "")
        if not text_pattern or text_pattern == ".+":
            continue
        
        try:
            pattern = re.compile(text_pattern, re.IGNORECASE | re.MULTILINE | re.DOTALL)
        except re.error:
            continue
        
        match = pattern.search(source_text)
        
        if match:
            heading_text = match.group().strip()
            # Get raw heading position for nearby content extraction
            match_start = match.start()
            match_end = match.end()
            
            # Extract nearby content (up to 1500 chars after heading)
            nearby_start = match_end
            nearby_end = min(nearby_start + 1500, len(source_text))
            raw_nearby = source_text[nearby_start:nearby_end].strip()
            
            # Try to truncate at next section-like boundary
            next_section = re.search(
                r'\n\s*\n(?=[A-Z][A-Za-z\s]{2,60}\n)|'
                r'\n(?=\d+\.\s+[A-Z])|'
                r'\n(?=Abstract|Abstrak|Keywords|Kata kunci|Introduction|'
                r'Methodology|Results?|Discussion|Conclusion|References?'
                r'|Acknowledgement|Graphical abstract)',
                raw_nearby, re.IGNORECASE
            )
            if next_section:
                raw_nearby = raw_nearby[:next_section.start()]
            
            detected[item_id] = {
                "heading_found": True,
                "heading_text": heading_text[:100],
                "evidence": heading_text[:200],
                "nearby": raw_nearby[:800],
                "confidence": 0.85,  # DOCX direct match is high confidence
            }
        else:
            # Check if content exists semantically even without exact heading
            if item_id in ("abstract", "abstrak"):
                if looks_like_abstract_content(source_text):
                    detected[item_id] = {
                        "heading_found": False,
                        "heading_text": "",
                        "evidence": "Semantic abstract content detected in source text",
                        "nearby": source_text[:800],
                        "confidence": 0.70,
                    }
                    continue
    
    return {
        "items": detected,
        "full_text": source_text,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# v5.7.1: AI Verifier - MULTI-SOURCE EVIDENCE EVALUATOR
# rate limiting, caching, model switching preserved from v5.7
# ═══════════════════════════════════════════════════════════════════════════════

def build_ai_verifier_input(item_id, doc_found, doc_confidence, doc_snippet,
                            pdf_found, pdf_confidence, pdf_snippet):
    """
    v5.7.1: Build the AI verifier input package.

    Required behavior:
    - include item/component id
    - include DOC detector result and DOC snippet
    - include PDF detector result and PDF snippet
    - include detector scores
    - include current detector confidence
    - include conflict status/summary

    This function intentionally returns only small snippets. Redaction is handled
    before this function is called by run_ai_verifier_multi_source().
    """
    doc_confidence = float(doc_confidence or 0.0)
    pdf_confidence = float(pdf_confidence or 0.0)
    doc_status = "FOUND" if doc_found else "NOT FOUND"
    pdf_status = "FOUND" if pdf_found else "NOT FOUND"

    if doc_found and pdf_found:
        conflict_status = "DOC_PDF_AGREE_FOUND"
        conflict_summary = "DOC and PDF detectors both found the component."
    elif doc_found and not pdf_found:
        conflict_status = "DOC_FOUND_PDF_NOT_FOUND"
        conflict_summary = "DOC detector found the component, but PDF detector did not. Possible PDF extraction/layout issue."
    elif not doc_found and pdf_found:
        conflict_status = "DOC_NOT_FOUND_PDF_FOUND"
        conflict_summary = "PDF detector found the component, but DOC detector did not. Possible DOC extraction issue."
    else:
        conflict_status = "DOC_PDF_AGREE_NOT_FOUND"
        conflict_summary = "Both DOC and PDF detectors did not find the component. Evidence may be weak/noisy."

    current_detector_confidence = max(doc_confidence, pdf_confidence)

    return {
        "item_id": item_id,
        "component": item_id.replace("_", " ").title(),
        "doc_detector_result": doc_status,
        "doc_detector_found": bool(doc_found),
        "doc_detector_confidence": doc_confidence,
        "doc_snippet": (doc_snippet or "")[:AI_VERIFIER_MAX_SNIPPET_CHARS],
        "pdf_detector_result": pdf_status,
        "pdf_detector_found": bool(pdf_found),
        "pdf_detector_confidence": pdf_confidence,
        "pdf_snippet": (pdf_snippet or "")[:AI_VERIFIER_MAX_SNIPPET_CHARS],
        "current_detector_confidence": current_detector_confidence,
        "conflict_status": conflict_status,
        "conflict_summary": conflict_summary,
    }


def _build_ai_multi_source_prompt(item_id, doc_found, doc_confidence, redacted_doc_snippet,
                                   pdf_found, pdf_confidence, redacted_pdf_snippet,
                                   is_vision=False, verifier_input=None):
    """
    v5.7.1: Build a multi-source AI prompt that compares BOTH DOC and PDF evidence.

    AI acts as MULTI-EVIDENCE CONSISTENCY EVALUATOR.
    AI compares DOC snippet vs PDF snippet to determine:
    - which extraction is trustworthy
    - whether extraction/layout failure caused one side to miss the component
    - final existence judgment based on combined evidence
    """
    item_desc = {
        "abstract": "English Abstract section (academic abstract with objectives, methods, results, conclusions)",
        "abstrak": "Malay Abstract (Abstrak) section (Malay language academic abstract)",
        "keywords": "Keywords section (comma or semicolon-separated list of 2-12 keyword terms)",
        "kata_kunci": "Kata Kunci section (Malay language keyword terms list)",
        "introduction": "Introduction section (research problem, context, objectives)",
        "methodology": "Methodology section (research methods, materials, procedures, experimental design)",
        "results_discussion": "Results and Discussion section (research findings and their implications)",
        "conclusion": "Conclusion section (summary of findings, conclusions, future work)",
        "references": "References section (list of academic citations/references)",
        "acknowledgement": "Acknowledgement section (funding, contributors, institutions)",
        "corresponding_author": "Corresponding Author section (author contact info: email, address, phone)",
        "graphical_abstract": "Graphical Abstract (visual summary figure of research)",
    }
    desc = item_desc.get(item_id, f"'{item_id}' section in an academic manuscript")

    if is_vision:
        return (
            "You are verifying if an academic manuscript contains a Graphical Abstract.\n"
            "A Graphical Abstract is a visual summary of research, typically a single figure/image\n"
            "that includes icons, diagrams, flowcharts, or illustrations summarizing the study.\n\n"
            "Analyze this image and determine if it is a Graphical Abstract.\n"
            "Return JSON only:\n"
            '{"is_graphical_abstract": true/false, "topic_relevance": "high/medium/low/unknown", '
            '"confidence": 0.0-1.0, "reason": "...", "visual_evidence": "..."}'
        )

    if verifier_input is None:
        verifier_input = build_ai_verifier_input(
            item_id,
            doc_found, doc_confidence, redacted_doc_snippet,
            pdf_found, pdf_confidence, redacted_pdf_snippet,
        )

    doc_status_str = verifier_input.get("doc_detector_result", "FOUND" if doc_found else "NOT FOUND")
    pdf_status_str = verifier_input.get("pdf_detector_result", "FOUND" if pdf_found else "NOT FOUND")
    current_detector_confidence = verifier_input.get("current_detector_confidence", max(doc_confidence, pdf_confidence))
    conflict_status = verifier_input.get("conflict_status", "UNKNOWN")
    conflict_summary = verifier_input.get("conflict_summary", "")

    prompt = (
        "You are a MULTI-EVIDENCE CONSISTENCY EVALUATOR for academic manuscript screening.\n\n"
        f"Component to verify: {desc}\n\n"
        "You receive TWO evidence sources: DOC (extracted from DOCX/DOC) and PDF (extracted from PDF).\n"
        f"Current detector confidence: {current_detector_confidence:.2f}\n"
        f"Current conflict status: {conflict_status}\n"
        f"Conflict summary: {conflict_summary}\n\n"
        "Your task:\n"
        "1. Compare both DOC and PDF snippets.\n"
        "2. Identify if an extraction or layout failure caused one source to miss the component.\n"
        "3. Determine if the component likely EXISTS in the manuscript.\n"
        "4. Explain the conflict if evidence disagrees.\n\n"
        "Reasoning Rules:\n"
        "- DOC FOUND + PDF NOT FOUND: If DOC snippet is strong, lean YES. Likely pdf_extraction_failure.\n"
        "- DOC NOT FOUND + PDF FOUND: If PDF snippet is strong, lean YES. Likely doc_extraction_failure.\n"
        "- DOC FOUND + PDF FOUND: Usually skipped, but resolve if confidence conflict exists.\n"
        "- DOC NOT FOUND + PDF NOT FOUND: Check both snippets for weak semantic evidence.\n"
        "  If both are garbled/empty/noisy, say weak_evidence. Do NOT hallucinate structure.\n\n"
        "Important Rules:\n"
        "- Return JSON only. Do NOT echo snippets back.\n"
        "- preferred_source must be one of: DOC, PDF, BOTH, NONE\n"
        "- likely_issue must be one of: pdf_extraction_failure, doc_extraction_failure, "
        "layout_issue, weak_evidence, no_structure, none\n\n"
        f"--- DOC Detector ---\n"
        f"Result: {doc_status_str}\n"
        f"Confidence: {doc_confidence:.2f}\n\n"
        f"DOC Snippet (redacted):\n{redacted_doc_snippet or '[No DOC snippet available]'}\n\n"
        f"--- PDF Detector ---\n"
        f"Result: {pdf_status_str}\n"
        f"Confidence: {pdf_confidence:.2f}\n\n"
        f"PDF Snippet (redacted):\n{redacted_pdf_snippet or '[No PDF snippet available]'}\n\n"
        "Based on BOTH evidence sources, does this manuscript likely contain a valid "
        f"{item_id.replace('_', ' ').title()} section?\n\n"
        'Return JSON only:\n'
        '{\n'
        '  "exists": true/false,\n'
        '  "confidence": 0.0-1.0,\n'
        '  "preferred_source": "DOC/PDF/BOTH/NONE",\n'
        '  "conflict_reason": "...",\n'
        '  "likely_issue": "pdf_extraction_failure/doc_extraction_failure/layout_issue/weak_evidence/no_structure/none",\n'
        '  "reason": "...",\n'
        '  "evidence_summary": "..."\n'
        '}'
    )
    return prompt


# Keep legacy single-source prompt for backward compat (vision only now uses this)
def _build_ai_prompt(item_id, redacted_snippet, is_vision=False):
    """Legacy single-source prompt builder. Only used for vision (graphical abstract)."""
    if is_vision:
        return _build_ai_multi_source_prompt(item_id, False, 0.0, "", False, 0.0, "", is_vision=True)
    # Fall back to multi-source with only one snippet
    return _build_ai_multi_source_prompt(item_id, True, 0.5, redacted_snippet, False, 0.0, "")


def run_ai_verifier_multi_source(
    item_id,
    doc_found, doc_confidence, doc_snippet,
    pdf_found, pdf_confidence, pdf_snippet,
    is_vision=False, vision_image_path=None, vision_context_text=""
):
    """
    v5.7.1: Multi-source AI verifier.
    Sends BOTH DOC and PDF evidence to AI.
    AI acts as MULTI-EVIDENCE CONSISTENCY EVALUATOR.

    Returns:
        dict: {
            "exists": True/False/None,
            "confidence": 0.0-1.0,
            "preferred_source": "DOC/PDF/BOTH/NONE",
            "conflict_reason": "...",
            "likely_issue": "pdf_extraction_failure/doc_extraction_failure/layout_issue/weak_evidence/no_structure/none",
            "reason": "...",
            "evidence_summary": "...",
            "evidence": "...",          # alias for evidence_summary (backward compat)
            "ai_status": "YES"/"NO"/"SKIPPED"/"DISABLED"/"ERROR"/"CACHED"
        }
    """
    global USE_AI_VERIFIER

    if not USE_AI_VERIFIER:
        return {
            "exists": None, "confidence": 0.0, "preferred_source": "NONE",
            "conflict_reason": "", "likely_issue": "none",
            "reason": "AI verifier is disabled in configuration",
            "evidence_summary": "", "evidence": "",
            "ai_status": "DISABLED",
        }

    if not _check_ai_rate_limit():
        return {
            "exists": None, "confidence": 0.0, "preferred_source": "NONE",
            "conflict_reason": "", "likely_issue": "none",
            "reason": "AI rate limit reached",
            "evidence_summary": "", "evidence": "",
            "ai_status": "SKIPPED",
        }

    if not OPENAI_API_KEY:
        return {
            "exists": None, "confidence": 0.0, "preferred_source": "NONE",
            "conflict_reason": "", "likely_issue": "none",
            "reason": "OpenAI API key not configured",
            "evidence_summary": "", "evidence": "",
            "ai_status": "DISABLED",
        }

    # Redact both snippets
    redacted_doc = redact_sensitive(doc_snippet or "") if not is_vision else ""
    redacted_pdf = redact_sensitive(pdf_snippet or "") if not is_vision else ""

    verifier_input = build_ai_verifier_input(
        item_id,
        doc_found, doc_confidence, redacted_doc,
        pdf_found, pdf_confidence, redacted_pdf,
    )

    # Cache key covers both sources
    cache_key = _compute_ai_cache_key(
        item_id,
        json.dumps(verifier_input, sort_keys=True, ensure_ascii=False)
    )
    cached = _get_ai_cache(cache_key)
    if cached is not None:
        cached["ai_status"] = "CACHED"
        return cached

    model = AI_VISION_MODEL if is_vision else AI_TEXT_MODEL

    try:
        import urllib.request
        import json as json_lib
        import base64

        prompt = _build_ai_multi_source_prompt(
            item_id,
            doc_found, doc_confidence, redacted_doc,
            pdf_found, pdf_confidence, redacted_pdf,
            is_vision=is_vision,
            verifier_input=verifier_input,
        )

        messages = [
            {"role": "system", "content": "You are an academic manuscript verification assistant. Return JSON only."},
        ]

        if is_vision and vision_image_path:
            with open(vision_image_path, "rb") as img_file:
                img_base64 = base64.b64encode(img_file.read()).decode("utf-8")
            content_parts = [{"type": "text", "text": prompt}]
            if vision_context_text:
                content_parts.append({"type": "text", "text": f"Context (redacted): {vision_context_text[:500]}"})
            content_parts.append({
                "type": "image_url",
                "image_url": {"url": f"data:image/png;base64,{img_base64}", "detail": "low"}
            })
            user_message = {"role": "user", "content": content_parts}
            max_tokens = 400
        else:
            user_message = {"role": "user", "content": prompt}
            max_tokens = 400

        messages.append(user_message)
        payload = {"model": model, "messages": messages, "temperature": 0.1, "max_tokens": max_tokens}
        data = json_lib.dumps(payload).encode("utf-8")

        req = urllib.request.Request(
            "https://api.openai.com/v1/chat/completions",
            data=data,
            headers={"Content-Type": "application/json", "Authorization": f"Bearer {OPENAI_API_KEY}"},
            method="POST",
        )
        response = urllib.request.urlopen(req, timeout=AI_TIMEOUT_SECONDS)
        response_data = json_lib.loads(response.read().decode("utf-8"))

        if "choices" in response_data and len(response_data["choices"]) > 0:
            content = response_data["choices"][0]["message"]["content"]
            result = json_lib.loads(content)

            if is_vision:
                exists = result.get("is_graphical_abstract", False)
                evidence = result.get("visual_evidence", result.get("reason", ""))
            else:
                exists = result.get("exists", False)
                evidence = result.get("evidence_summary", result.get("reason", ""))

            ai_status = "YES" if exists else "NO"
            out = {
                "exists": exists,
                "confidence": result.get("confidence", 0.0),
                "preferred_source": result.get("preferred_source", "NONE"),
                "conflict_reason": result.get("conflict_reason", ""),
                "likely_issue": result.get("likely_issue", "none"),
                "reason": result.get("reason", ""),
                "evidence_summary": result.get("evidence_summary", ""),
                "evidence": evidence[:200],
                "verifier_input": verifier_input,
                "ai_status": ai_status,
            }
            _set_ai_cache(cache_key, out)
            return out

    except Exception as e:
        print(f"[AI_VERIFIER_MULTI] Error: {e}")

    return {
        "exists": None, "confidence": 0.0, "preferred_source": "NONE",
        "conflict_reason": "", "likely_issue": "none",
        "reason": "AI verifier failed", "evidence_summary": "", "evidence": "",
        "ai_status": "ERROR",
    }


def run_ai_verifier(item_id, snippet, context="", is_vision=False, vision_image_path=None, vision_context_text=""):
    """
    v5.7 legacy single-source AI verifier — kept for vision/backward compat.
    In v5.7.1 text verification uses run_ai_verifier_multi_source() instead.

    Args:
        item_id: The section/item ID to verify.
        snippet: Small text snippet to verify (max ~800 chars).
        is_vision: If True, use vision model for graphical abstract.
        vision_image_path: Path to cropped image for vision analysis.
        vision_context_text: Redacted title/abstract context for vision.
    """
    global USE_AI_VERIFIER

    if not USE_AI_VERIFIER:
        return {
            "exists": None,
            "component": item_id,
            "confidence": 0.0,
            "evidence": "AI verifier is disabled in configuration",
            "ai_status": "DISABLED",
        }

    # Check rate limits
    if not _check_ai_rate_limit():
        return {
            "exists": None,
            "component": item_id,
            "confidence": 0.0,
            "evidence": "AI rate limit reached (max calls exceeded)",
            "ai_status": "SKIPPED",
        }

    if not OPENAI_API_KEY:
        return {
            "exists": None,
            "component": item_id,
            "confidence": 0.0,
            "evidence": "OpenAI API key not configured",
            "ai_status": "DISABLED",
        }

    # ── Build redacted snippet ──
    redacted = ""
    if not is_vision:
        if not snippet or len(snippet.strip()) < 20:
            return {
                "exists": None,
                "component": item_id,
                "confidence": 0.0,
                "evidence": "Snippet too short for AI verification",
                "ai_status": "SKIPPED",
            }
        redacted = redact_sensitive(snippet)
        if len(redacted.strip()) < 20:
            return {
                "exists": None,
                "component": item_id,
                "confidence": 0.0,
                "evidence": "After redaction, snippet too short for AI verification",
                "ai_status": "SKIPPED",
            }

    # ── Check cache ──
    cache_key = _compute_ai_cache_key(item_id, redacted if not is_vision else f"vision::{item_id}::{vision_image_path or 'no_image'}")
    cached = _get_ai_cache(cache_key)
    if cached is not None:
        cached["ai_status"] = "CACHED"
        return cached

    # ── Determine model ──
    model = AI_VISION_MODEL if is_vision else AI_TEXT_MODEL

    # ── Build request ──
    try:
        import urllib.request
        import json as json_lib
        import base64

        prompt = _build_ai_prompt(item_id, redacted, is_vision=is_vision)

        messages = [
            {"role": "system", "content": "You are an academic manuscript verification assistant. Return JSON only."},
        ]

        if is_vision and vision_image_path:
            # Send image + text context
            with open(vision_image_path, "rb") as img_file:
                img_base64 = base64.b64encode(img_file.read()).decode("utf-8")

            content_parts = [
                {"type": "text", "text": prompt},
            ]
            if vision_context_text:
                content_parts.append({
                    "type": "text",
                    "text": f"Context (redacted): {vision_context_text[:500]}"
                })
            content_parts.append({
                "type": "image_url",
                "image_url": {
                    "url": f"data:image/png;base64,{img_base64}",
                    "detail": "low",
                }
            })

            user_message = {"role": "user", "content": content_parts}
            max_tokens = 400
        else:
            # Text-only request
            user_message = {"role": "user", "content": prompt}
            max_tokens = 300

        messages.append(user_message)

        payload = {
            "model": model,
            "messages": messages,
            "temperature": 0.1,
            "max_tokens": max_tokens,
        }

        data = json_lib.dumps(payload).encode('utf-8')

        req = urllib.request.Request(
            "https://api.openai.com/v1/chat/completions",
            data=data,
            headers={
                "Content-Type": "application/json",
                "Authorization": f"Bearer {OPENAI_API_KEY}",
            },
            method="POST",
        )

        response = urllib.request.urlopen(req, timeout=AI_TIMEOUT_SECONDS)
        response_data = json_lib.loads(response.read().decode('utf-8'))

        if "choices" in response_data and len(response_data["choices"]) > 0:
            content = response_data["choices"][0]["message"]["content"]
            result = json_lib.loads(content)

            ai_status = "YES" if result.get("exists", False) or result.get("is_graphical_abstract", False) else "NO"
            confidence = result.get("confidence", 0.0)

            if is_vision:
                exists = result.get("is_graphical_abstract", False)
                evidence = result.get("visual_evidence", result.get("reason", ""))
            else:
                exists = result.get("exists", False)
                evidence = result.get("evidence", "")

            out = {
                "exists": exists,
                "component": result.get("component", item_id),
                "confidence": confidence,
                "evidence": evidence[:200],
                "ai_status": ai_status,
                "topic_relevance": result.get("topic_relevance", "unknown") if is_vision else None,
            }

            # Cache the result
            _set_ai_cache(cache_key, out)
            return out

    except json.JSONDecodeError as e:
        print(f"[AI_VERIFIER] JSON parse error: {e}")
    except urllib.error.HTTPError as e:
        print(f"[AI_VERIFIER] HTTP {e.code}: {e.reason}")
        if e.code == 429:
            return {
                "exists": None,
                "component": item_id,
                "confidence": 0.0,
                "evidence": "AI rate limited (HTTP 429)",
                "ai_status": "ERROR",
            }
    except urllib.error.URLError as e:
        print(f"[AI_VERIFIER] URL error: {e.reason}")
    except Exception as e:
        print(f"[AI_VERIFIER] Error: {e}")

    return {
        "exists": None,
        "component": item_id,
        "confidence": 0.0,
        "evidence": "AI verifier failed",
        "ai_status": "ERROR",
    }


# ═══════════════════════════════════════════════════════════════════════════════
# v3.0 NEW: DOCX-to-PDF text comparison for debug report
# ═══════════════════════════════════════════════════════════════════════════════

def _compare_docx_pdf_text(docx_text, pdf_text):
    """
    Compare DOCX source text vs PDF extracted text to generate debug info.
    
    Returns a dict with comparison metrics.
    """
    if not docx_text and not pdf_text:
        return {"docx_chars": 0, "pdf_chars": 0, "ratio": 0}
    
    docx_len = len(docx_text or "")
    pdf_len = len(pdf_text or "")
    
    # Calculate overlap approximation (simple word overlap)
    docx_words = set((docx_text or "").lower().split())
    pdf_words = set((pdf_text or "").lower().split())
    
    overlap = len(docx_words & pdf_words) if docx_words and pdf_words else 0
    union = len(docx_words | pdf_words) if docx_words or pdf_words else 1
    
    return {
        "docx_chars": docx_len,
        "pdf_chars": pdf_len,
        "overlap_ratio": round(overlap / union, 4) if union > 0 else 0,
    }


# ─── Helper: PDF Parsing ─────────────────────────────────────────────────────

def extract_pdf_layout(pdf_path):
    """
    Extract full layout information from a PDF using PyMuPDF.
    Returns a list of page dicts, each containing:
      - page_num (int)
      - page_width, page_height (float)
      - text (str): full page text
      - blocks (list): text blocks with bbox, lines, spans
      - images (list): image blocks with bbox
      - drawings (list): vector drawing regions
      - image_info (list): detailed image info
    """
    if fitz is None:
        raise ImportError("PyMuPDF (fitz) is required. Install with: pip install pymupdf")

    doc = fitz.open(pdf_path)
    pages = []
    for page_num in range(len(doc)):
        page = doc[page_num]
        pw = page.rect.width
        ph = page.rect.height
        blocks = page.get_text("dict", flags=fitz.TEXT_PRESERVE_WHITESPACE)["blocks"]

        text_blocks = []
        image_blocks = []
        full_text_parts = []

        for block in blocks:
            if block["type"] == 0:  # text block
                bbox = block["bbox"]
                lines = []
                for line in block["lines"]:
                    spans = []
                    for span in line["spans"]:
                        spans.append({
                            "text": span["text"],
                            "font": span.get("font", ""),
                            "size": span.get("size", 0),
                            "flags": span.get("flags", 0),
                            "color": span.get("color", 0),
                            "bbox": span.get("bbox", line["bbox"]),
                        })
                    lines.append({
                        "bbox": line["bbox"],
                        "spans": spans,
                        "text": "".join(s["text"] for s in spans),
                    })
                line_text = " ".join(l["text"] for l in lines).strip()
                if line_text:
                    full_text_parts.append(line_text)
                text_blocks.append({
                    "bbox": bbox,
                    "lines": lines,
                    "text": line_text,
                })
            elif block["type"] == 1:  # image block
                image_blocks.append({
                    "bbox": block["bbox"],
                })

        # Get drawings (vector graphics, rectangles, lines)
        drawings = page.get_drawings()
        drawing_regions = []
        for d in drawings:
            drawing_regions.append({
                "bbox": d.get("rect", d.get("bbox", (0, 0, 0, 0))),
            })

        # Get detailed image info
        image_info = page.get_images(full=True) if hasattr(page, 'get_images') else []
        image_info_blocks = []
        for img in image_info:
            try:
                img_rect = page.get_image_rect(img[0]) if hasattr(page, 'get_image_rect') else None
                if img_rect:
                    image_info_blocks.append({
                        "bbox": (img_rect.x0, img_rect.y0, img_rect.x1, img_rect.y1),
                        "width": img[2] if len(img) > 2 else 0,
                        "height": img[3] if len(img) > 3 else 0,
                    })
            except Exception:
                pass

        pages.append({
            "page_num": page_num + 1,
            "page_width": pw,
            "page_height": ph,
            "text": "\n".join(full_text_parts),
            "blocks": text_blocks,
            "images": image_blocks,
            "drawings": drawing_regions,
            "image_info": image_info_blocks,
        })

    doc.close()
    return pages


def normalize_bbox(bbox, page_width, page_height):
    """
    Convert pixel/point bbox to normalized coordinates 0.0 to 1.0.
    bbox: (x0, y0, x1, y1) in points/pixels.
    Returns (x_min, y_min, x_max, y_max) all 0.0–1.0.
    """
    if page_width <= 0 or page_height <= 0:
        return (0, 0, 0, 0)
    x0, y0, x1, y1 = bbox
    return (
        max(0.0, min(1.0, x0 / page_width)),
        max(0.0, min(1.0, y0 / page_height)),
        max(0.0, min(1.0, x1 / page_width)),
        max(0.0, min(1.0, y1 / page_height)),
    )


def bbox_in_region(norm_bbox, expected_region):
    """
    Check if a normalized bbox overlaps with the expected region at all.
    Returns True if there is any overlap.
    """
    if norm_bbox is None:
        return False
    fx0, fy0, fx1, fy1 = norm_bbox
    ex0 = expected_region.get("x_min", 0)
    ex1 = expected_region.get("x_max", 1)
    ey0 = expected_region.get("y_min", 0)
    ey1 = expected_region.get("y_max", 1)
    h_overlap = min(fx1, ex1) - max(fx0, ex0)
    v_overlap = min(fy1, ey1) - max(fy0, ey0)
    return h_overlap > 0 and v_overlap > 0


def find_text_matches(layout, pattern):
    """
    Search for regex pattern in the full text of all pages.
    Returns list of matches with page number and context info.
    """
    matches = []
    try:
        compiled = re.compile(pattern, re.IGNORECASE)
    except re.error:
        return matches

    for page_data in layout:
        page_text = page_data["text"]
        for match in compiled.finditer(page_text):
            # Find which block/span contains this match
            found_bbox = None
            found_span = None
            found_block = None
            matched_text = match.group()

            for block in page_data["blocks"]:
                if matched_text in block["text"]:
                    bbox = block["bbox"]
                    found_bbox = normalize_bbox(
                        bbox,
                        page_data["page_width"],
                        page_data["page_height"],
                    )
                    found_block = block
                    # Get first span for formatting check
                    if block["lines"] and block["lines"][0]["spans"]:
                        found_span = block["lines"][0]["spans"][0]
                    break

            matches.append({
                "page": page_data["page_num"],
                "page_width": page_data["page_width"],
                "page_height": page_data["page_height"],
                "matched_text": matched_text,
                "bbox_raw": found_bbox,
                "span": found_span,
                "block": found_block,
                "page_data": page_data,
            })

    return matches


def check_region(found_bbox, expected_region, tolerance):
    """
    Check if found bbox is within expected region ± tolerance.
    found_bbox: (x_min, y_min, x_max, y_max) normalized 0-1.
    expected_region: dict with x_min, x_max, y_min, y_max.
    tolerance: float 0.0–1.0 margin of error.

    Returns (passed, deviation) where deviation is how far outside.
    """
    if found_bbox is None:
        return False, 1.0

    fx0, fy0, fx1, fy1 = found_bbox
    ex0 = expected_region.get("x_min", 0)
    ex1 = expected_region.get("x_max", 1)
    ey0 = expected_region.get("y_min", 0)
    ey1 = expected_region.get("y_max", 1)

    # Check horizontal overlap
    h_overlap = min(fx1, ex1) - max(fx0, ex0)
    h_ok = h_overlap >= 0  # any overlap at all

    # Check vertical overlap
    v_overlap = min(fy1, ey1) - max(fy0, ey0)
    v_ok = v_overlap >= 0

    # Calculate deviation
    h_dev = 0.0
    if fx0 < ex0 - tolerance:
        h_dev = max(h_dev, ex0 - tolerance - fx0)
    if fx1 > ex1 + tolerance:
        h_dev = max(h_dev, fx1 - (ex1 + tolerance))

    v_dev = 0.0
    if fy0 < ey0 - tolerance:
        v_dev = max(v_dev, ey0 - tolerance - fy0)
    if fy1 > ey1 + tolerance:
        v_dev = max(v_dev, fy1 - (ey1 + tolerance))

    deviation = max(h_dev, v_dev)
    passed = h_ok and v_ok

    return passed, deviation


def check_formatting(found_span, format_rules):
    """
    Check formatting of a text span against expected rules.
    Returns (passed, details dict).
    """
    if found_span is None or format_rules is None:
        return True, {}

    issues = []
    all_ok = True

    if format_rules.get("uppercase_preferred"):
        text = found_span.get("text", "")
        # Check if first few words are uppercase
        words = text.split()[:5]
        upper_count = sum(1 for w in words if w.isupper() and len(w) > 1)
        if upper_count < len(words) // 2:
            issues.append("text is not predominantly uppercase")
            all_ok = False

    if format_rules.get("bold_preferred"):
        flags = found_span.get("flags", 0)
        is_bold = bool(flags & 2)  # font flag bit 1 = bold
        if not is_bold:
            issues.append("text is not bold")

    if format_rules.get("image_required"):
        # This is checked at page level (image bbox presence)
        pass

    return all_ok, {"issues": issues, "all_ok": all_ok}


def score_results(results):
    """
    Score the screening results based on item statuses.
    Returns (final_score, structure_score, layout_score, formatting_score).
    """
    scoring_config = {
        "structure": 40,
        "layout": 30,
        "formatting": 30,
    }

    total_items = len(results)
    if total_items == 0:
        return 0, 0, 0, 0

    passed = sum(1 for r in results if r.get("status") == "PASS")
    warnings = sum(1 for r in results if r.get("status") == "WARNING")
    failed = sum(1 for r in results if r.get("status") == "FAILED")

    # Structure items (required content presence)
    structure_items = [
        "title", "authors", "affiliations", "corresponding_author",
        "graphical_abstract", "abstract", "keywords", "abstrak", "kata_kunci",
        "introduction", "methodology", "results_discussion", "conclusion",
        "acknowledgement", "references",
    ]
    struct_results = [r for r in results if r.get("id") in structure_items]
    struct_score = 0
    if struct_results:
        struct_passed = sum(1 for r in struct_results if r.get("status") == "PASS")
        struct_score = round((struct_passed / len(struct_results)) * 100)

    # Layout items (coordinate checks)
    layout_items = [
        "title", "authors", "affiliations", "corresponding_author",
        "graphical_abstract", "abstract", "keywords", "abstrak", "kata_kunci",
    ]
    layout_results = [r for r in results if r.get("id") in layout_items]
    layout_score = 0
    if layout_results:
        layout_ok = sum(1 for r in layout_results if r.get("status") in ("PASS", "WARNING"))
        layout_score = round((layout_ok / len(layout_results)) * 100)

    # Formatting items
    fmt_results = [r for r in results if r.get("formatting_ok") is not None]
    fmt_score = 0
    if fmt_results:
        fmt_ok = sum(1 for r in fmt_results if r.get("formatting_ok"))
        fmt_score = round((fmt_ok / len(fmt_results)) * 100)

    # Weighted final score
    w_struct = scoring_config["structure"]
    w_layout = scoring_config["layout"]
    w_format = scoring_config["formatting"]

    final_score = round(
        (struct_score * w_struct + layout_score * w_layout + fmt_score * w_format) / 100
    )

    return final_score, struct_score, layout_score, fmt_score


# ─── DOCX to PDF Conversion ──────────────────────────────────────────────────

def convert_docx_to_pdf(docx_path, output_dir=None):
    """
    Convert a DOCX file to PDF using LibreOffice CLI.
    Returns path to generated PDF, or None on failure.
    """
    docx_path = Path(docx_path)
    if not docx_path.exists():
        return None

    if output_dir is None:
        output_dir = docx_path.parent

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    pdf_path = output_dir / f"{docx_path.stem}.pdf"

    # If PDF already exists, return it
    if pdf_path.exists():
        return str(pdf_path)

    try:
        # Try soffice command
        cmd = [
            "soffice",
            "--headless",
            "--convert-to", "pdf",
            "--outdir", str(output_dir),
            str(docx_path),
        ]
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=120,
        )

        if result.returncode != 0:
            print(f"[CONVERT] LibreOffice stderr: {result.stderr}")

        # Check if PDF was created
        if pdf_path.exists():
            return str(pdf_path)
        else:
            # Try alternate paths
            possible = sorted(output_dir.glob("*.pdf"))
            if possible:
                return str(possible[-1])
            return None

    except subprocess.TimeoutExpired:
        print(f"[CONVERT] LibreOffice timed out for {docx_path}")
        return None
    except FileNotFoundError:
        print(f"[CONVERT] LibreOffice (soffice) not found. Install LibreOffice.")
        return None
    except Exception as e:
        print(f"[CONVERT] Unexpected error: {e}")
        return None


# ─── Candidate Ranking System ────────────────────────────────────────────────

def score_text_block_as_title(block, page_data, expected_region):
    """
    Score a text block as a potential manuscript title candidate.
    Returns a score 0.0–1.0.
    """
    text = block.get("text", "").strip()
    if not text:
        return 0.0

    norm_bbox = normalize_bbox(block["bbox"], page_data["page_width"], page_data["page_height"])

    score = 0.0

    # Length check: title should be substantial (more than ~25 chars)
    if len(text) > 25:
        score += 0.25
    elif len(text) > 15:
        score += 0.15
    else:
        score += 0.05

    # Should be on page 1
    if page_data["page_num"] == 1:
        score += 0.15

    # In expected region (upper portion of page)
    if norm_bbox and bbox_in_region(norm_bbox, expected_region):
        score += 0.20

    # Font size check - title is typically larger
    max_font_size = 0
    for line in block.get("lines", []):
        for span in line.get("spans", []):
            max_font_size = max(max_font_size, span.get("size", 0))
    if max_font_size >= 12:
        score += 0.10
    if max_font_size >= 14:
        score += 0.05

    # Uppercase ratio - titles are often in title case or uppercase
    words = text.split()
    if len(words) > 2:
        upper_words = sum(1 for w in words if w[0].isupper() and len(w) > 1)
        upper_ratio = upper_words / len(words)
        if upper_ratio > 0.5:
            score += 0.10

    # Not starting with a number/section (bad for title)
    if text[0].isalpha():
        score += 0.05

    # Does not contain common section words
    section_indicators = ["introduction", "methodology", "result", "conclusion",
                          "acknowledgement", "reference", "abstract", "keyword"]
    text_lower = text.lower()
    if not any(indicator in text_lower for indicator in section_indicators):
        score += 0.10

    return min(score, 1.0)


def score_text_block_as_author(block, page_data, expected_region):
    """
    Score a text block as a potential authors candidate.
    Returns a score 0.0–1.0.
    """
    text = block.get("text", "").strip()
    if not text:
        return 0.0

    norm_bbox = normalize_bbox(block["bbox"], page_data["page_width"], page_data["page_height"])

    score = 0.0

    # Should be on page 1
    if page_data["page_num"] == 1:
        score += 0.10

    # In expected region (below title)
    if norm_bbox and bbox_in_region(norm_bbox, expected_region):
        score += 0.15

    # Authors typically contain commas separating names
    # Pattern: Name1, Name2, Name3 or Name1¹, Name2²
    comma_count = text.count(",")
    if 1 <= comma_count <= 10:
        score += 0.20

    # Contains superscript indicators (¹²³⁴ or [1][2] or 1,2,3)
    has_superscript = bool(re.search(r'[¹²³⁴⁵⁶⁷⁸⁹⁰]|\[\d+\]|\d+(?:,\d+)*$', text))
    if has_superscript:
        score += 0.10

    # Should have multiple words (names)
    words = text.split()
    if 2 <= len(words) <= 20:
        score += 0.10

    # Should contain person-name-like patterns
    # Pattern: "Ahmad Ali", "Siti Aminah", etc.
    name_pattern = bool(re.search(r'[A-Z][a-zà-ü]+(?:\s+[A-Z][a-zà-ü]+)+', text))
    if name_pattern:
        score += 0.15

    # Should not be too long (affiliations are longer)
    if len(text) < 100:
        score += 0.10

    # Should not contain institution words
    institution_words = ["university", "universiti", "faculty", "school of",
                         "department", "institute", "centre", "center",
                         "college", "malaysia", "campus", "utm"]
    text_lower = text.lower()
    if not any(word in text_lower for word in institution_words):
        score += 0.10

    return min(score, 1.0)


def score_text_block_as_affiliation(block, page_data, expected_region):
    """
    Score a text block as a potential affiliations candidate.
    Returns a score 0.0–1.0.
    """
    text = block.get("text", "").strip()
    if not text:
        return 0.0

    norm_bbox = normalize_bbox(block["bbox"], page_data["page_width"], page_data["page_height"])

    score = 0.0

    # Should be on page 1
    if page_data["page_num"] == 1:
        score += 0.10

    # In expected region (below authors)
    if norm_bbox and bbox_in_region(norm_bbox, expected_region):
        score += 0.15

    # Contains institution indicators (strong signal)
    institution_words = [
        "university", "universiti", "faculty", "school", "department",
        "institute", "centre", "center", "college", "malaysia",
        "campus", "utm", "utm", "penerbit", "press", "jalan",
        "skudai", "johor", "kuala lumpur", "pulau pinang",
    ]
    text_lower = text.lower()
    institution_matches = sum(1 for word in institution_words if word in text_lower)
    if institution_matches >= 2:
        score += 0.25
    elif institution_matches == 1:
        score += 0.15

    # Contains email (strong signal)
    if re.search(r'[\w.+-]+@[\w-]+\.[\w.-]+', text):
        score += 0.15

    # Contains postal code
    if re.search(r'\b\d{5}\b', text):
        score += 0.10

    # Contains comma-separated parts (address format)
    if text.count(",") >= 1:
        score += 0.10

    # Should be reasonably long (addresses are usually 50+ chars)
    if len(text) > 50:
        score += 0.10

    # Should not look like an author list (few commas, no superscript affiliation markers)
    # Affiliation text typically has more "content" per segment
    if text.count(",") <= 6:
        score += 0.05

    return min(score, 1.0)


def find_text_blocks_in_region(page_data, expected_region):
    """
    Find all text blocks on a page that overlap with the expected region.
    Returns list of blocks with normalized bbox.
    """
    matching_blocks = []
    for block in page_data["blocks"]:
        norm_bbox = normalize_bbox(
            block["bbox"],
            page_data["page_width"],
            page_data["page_height"],
        )
        if bbox_in_region(norm_bbox, expected_region):
            matching_blocks.append({
                "block": block,
                "norm_bbox": norm_bbox,
                "page": page_data["page_num"],
            })
    return matching_blocks


def find_best_candidate_for_rule(rule, layout):
    """
    Find the best candidate text block for a rule.
    Uses intelligent scoring to pick the most likely correct match.
    Uses item-aware false positive detection.

    For rules with text_pattern ".+", looks for text blocks in the expected region
    and scores them. For rules with specific patterns, uses regex matching.

    Returns dict with:
      - candidate_text: str
      - confidence: float 0.0–1.0
      - bbox: normalized bbox tuple or None
      - page: int
      - block: text block dict or None
      - span: span dict or None
      - rejected_false_positives: list of rejected texts
      - all_candidates: list of all candidate scores
      - detection_method: str
      - nearby_content: str (content after heading for section items)
    """
    item_id = rule["id"]
    text_pattern = rule.get("text_pattern", "")
    expected_region = rule.get("expected_region", {})
    is_generic_pattern = (text_pattern == ".+")
    rejected_false_positives = []
    all_candidates = []
    detection_method = "not_found"

    if is_generic_pattern:
        # For generic patterns, collect all text blocks in the expected region
        # across relevant pages and score each one
        target_page = expected_region.get("page", 1)
        page_min = expected_region.get("page_min", target_page)
        page_max = expected_region.get("page_max", page_min)

        for page_data in layout:
            if page_min <= page_data["page_num"] <= page_max:
                blocks_in_region = find_text_blocks_in_region(page_data, expected_region)
                for item in blocks_in_region:
                    block = item["block"]
                    text = block.get("text", "").strip()
                    if not text:
                        continue

                    # Check false positives using item-aware function
                    if is_false_positive_for_item(text, item_id):
                        rejected_false_positives.append(text)
                        continue

                    # Score based on item type
                    if item_id == "title":
                        score_val = score_text_block_as_title(block, page_data, expected_region)
                    elif item_id == "authors":
                        score_val = score_text_block_as_author(block, page_data, expected_region)
                    elif item_id == "affiliations":
                        score_val = score_text_block_as_affiliation(block, page_data, expected_region)
                    else:
                        # Generic scoring: prefer longer text in expected region
                        score_val = 0.5
                        if bbox_in_region(item["norm_bbox"], expected_region):
                            score_val += 0.2
                        if len(text) > 30:
                            score_val += 0.2
                        if not is_false_positive_for_item(text, item_id):
                            score_val += 0.1

                    span = None
                    if block["lines"] and block["lines"][0]["spans"]:
                        span = block["lines"][0]["spans"][0]

                    all_candidates.append({
                        "text": text,
                        "score": min(score_val, 1.0),
                        "bbox": item["norm_bbox"],
                        "page": page_data["page_num"],
                        "block": block,
                        "span": span,
                    })

        # Sort by score descending, pick best
        all_candidates.sort(key=lambda c: c["score"], reverse=True)

        if all_candidates:
            best = all_candidates[0]
            detection_method = "score_ranking"
            # Re-fetch span from block
            best_span = None
            if best["block"]["lines"] and best["block"]["lines"][0]["spans"]:
                best_span = best["block"]["lines"][0]["spans"][0]
            return {
                "candidate_text": best["text"],
                "confidence": best["score"],
                "bbox": best["bbox"],
                "page": best["page"],
                "block": best["block"],
                "span": best_span,
                "rejected_false_positives": rejected_false_positives,
                "all_candidates": all_candidates[:5],  # top 5
                "detection_method": detection_method,
                "nearby_content": "",
            }
        else:
            return {
                "candidate_text": "",
                "confidence": 0.0,
                "bbox": None,
                "page": 0,
                "block": None,
                "span": None,
                "rejected_false_positives": rejected_false_positives,
                "all_candidates": [],
                "detection_method": "not_found",
                "nearby_content": "",
            }
    else:
        # For specific patterns (e.g. "Abstract", "Keywords"), use regex matching
        matches = find_text_matches(layout, text_pattern)

        # Filter out false positives using item-aware function
        valid_matches = []
        for m in matches:
            if not is_false_positive_for_item(m["matched_text"], item_id):
                valid_matches.append(m)

        if not valid_matches:
            # If no valid matches, try semantic fallback — search for content that
            # strongly looks like the expected section even without the exact heading
            nearby_content = ""
            if item_id in ("abstract", "abstrak"):
                # Try to find abstract-like content anywhere in the document
                for page_data in layout:
                    if looks_like_abstract_content(page_data["text"]):
                        detection_method = "semantic_fallback"
                        nearby_content = page_data["text"][:800]
                        break
            return {
                "candidate_text": "",
                "confidence": 0.0,
                "bbox": None,
                "page": 0,
                "block": None,
                "span": None,
                "rejected_false_positives": [m["matched_text"] for m in matches if is_false_positive_for_item(m["matched_text"], item_id)],
                "all_candidates": [],
                "detection_method": detection_method if detection_method != "not_found" else "not_found",
                "nearby_content": nearby_content,
            }

        # Pick best match by region proximity
        best_match = None
        best_deviation = 999
        best_confidence = 0.5

        for match in valid_matches:
            found_bbox = match.get("bbox_raw")
            if found_bbox:
                passed, deviation = check_region(found_bbox, expected_region, rule.get("tolerance", 0.15))
                conf = 0.7 if passed else 0.5
                if deviation < best_deviation:
                    best_deviation = deviation
                    best_match = match
                    best_confidence = conf

        if best_match is None and valid_matches:
            best_match = valid_matches[0]
            best_confidence = 0.4

        if best_match:
            # Extract nearby content after the heading for section items
            nearby_content = ""
            heading_text = best_match.get("matched_text", "")
            if item_id in ("abstract", "abstrak", "keywords", "kata_kunci",
                           "references", "corresponding_author", "graphical_abstract",
                           "introduction", "methodology", "results_discussion",
                           "conclusion", "acknowledgement"):
                page_data = best_match.get("page_data")
                if page_data and heading_text:
                    nearby_content = extract_nearby_text_after_heading(
                        page_data["text"], heading_text, max_chars=1200
                    )

            detection_method = "exact_rule"
            best_confidence = max(best_confidence, 0.75)  # Exact heading match gets high confidence

            return {
                "candidate_text": heading_text,
                "confidence": best_confidence,
                "bbox": best_match.get("bbox_raw"),
                "page": best_match.get("page", 0),
                "block": best_match.get("block"),
                "span": best_match.get("span"),
                "rejected_false_positives": [m["matched_text"] for m in matches if is_false_positive_for_item(m["matched_text"], item_id)],
                "all_candidates": [
                    {"text": m["matched_text"], "score": 0.5, "page": m.get("page", 0)}
                    for m in valid_matches[:3]
                ],
                "detection_method": detection_method,
                "nearby_content": nearby_content,
            }

        return {
            "candidate_text": "",
            "confidence": 0.0,
            "bbox": None,
            "page": 0,
            "block": None,
            "span": None,
            "rejected_false_positives": [],
            "all_candidates": [],
            "detection_method": "not_found",
            "nearby_content": "",
        }


def extract_following_paragraph(blocks, heading_text, page_data):
    """
    Find the paragraph text that comes immediately after a heading.
    Returns the paragraph text or empty string.
    """
    heading_block = None
    for block in blocks:
        if heading_text.lower() in block.get("text", "").lower():
            heading_block = block
            break

    if not heading_block:
        return ""

    # Get all blocks sorted by vertical position
    sorted_blocks = sorted(
        [b for b in page_data["blocks"] if b.get("text", "").strip()],
        key=lambda b: b["bbox"][1]  # sort by y0
    )

    # Find heading block index
    heading_idx = -1
    for i, b in enumerate(sorted_blocks):
        if id(b) == id(heading_block):
            heading_idx = i
            break

    if heading_idx == -1 or heading_idx + 1 >= len(sorted_blocks):
        return ""

    # Collect paragraphs after heading until next heading
    paragraphs = []
    for i in range(heading_idx + 1, len(sorted_blocks)):
        block_text = sorted_blocks[i].get("text", "").strip()
        if not block_text:
            continue
        # Stop if we hit another heading-like block
        if re.match(r'^[A-Z][A-Za-z\s]+$', block_text) and len(block_text) < 50:
            break
        if re.match(r'^\d+\.\d+\s+[A-Z]', block_text):
            break
        paragraphs.append(block_text)

    return " ".join(paragraphs)


def detect_graphical_region(page_data, region_bbox):
    """
    Detect if there is a graphical/image region in the expected area.
    Returns (detected, method, confidence).
    Uses multi-level detection:
    1. Direct image blocks
    2. Detailed image info
    3. Drawing/vector regions
    4. Large blank/non-text regions
    """
    pw = page_data["page_width"]
    ph = page_data["page_height"]

    # Normalize region bbox
    rx0, ry0, rx1, ry1 = region_bbox

    # Method 1: Check image blocks
    for img in page_data.get("images", []):
        img_norm = normalize_bbox(img["bbox"], pw, ph)
        if bbox_in_region(img_norm, {"x_min": rx0, "x_max": rx1, "y_min": ry0, "y_max": ry1}):
            return True, "image_block", 0.95

    # Method 2: Check detailed image info
    for img_info in page_data.get("image_info", []):
        img_norm = normalize_bbox(
            (img_info["bbox"][0], img_info["bbox"][1],
             img_info["bbox"][2], img_info["bbox"][3]),
            pw, ph
        )
        if bbox_in_region(img_norm, {"x_min": rx0, "x_max": rx1, "y_min": ry0, "y_max": ry1}):
            # Check if image is large enough (>10% of expected region area)
            expected_area = (rx1 - rx0) * (ry1 - ry0)
            img_area = (img_norm[2] - img_norm[0]) * (img_norm[3] - img_norm[1])
            if img_area > expected_area * 0.1:
                return True, "image_info", 0.90

    # Method 3: Check drawing/vector regions
    for dwg in page_data.get("drawings", []):
        dwg_norm = normalize_bbox(dwg["bbox"], pw, ph)
        if bbox_in_region(dwg_norm, {"x_min": rx0, "x_max": rx1, "y_min": ry0, "y_max": ry1}):
            # Must be reasonably sized (not just a tiny line)
            dwg_width = dwg_norm[2] - dwg_norm[0]
            dwg_height = dwg_norm[3] - dwg_norm[1]
            region_width = rx1 - rx0
            region_height = ry1 - ry0
            if dwg_width * dwg_height > (region_width * region_height) * 0.05:
                return True, "drawing", 0.75

    # Method 4: Check for large blank/non-text region
    # Find text blocks in the graphical region area
    text_blocks_in_region = find_text_blocks_in_region(
        page_data,
        {"x_min": rx0, "x_max": rx1, "y_min": ry0, "y_max": ry1}
    )

    if not text_blocks_in_region:
        # No text in graphical region - likely contains an image
        return True, "no_text_region", 0.60

    # If there's very little text compared to region size, likely has image
    total_text_area = 0
    for item in text_blocks_in_region:
        b = item["norm_bbox"]
        total_text_area += (b[2] - b[0]) * (b[3] - b[1])

    region_area = (rx1 - rx0) * (ry1 - ry0)
    if region_area > 0 and (total_text_area / region_area) < 0.3:
        return True, "sparse_text_region", 0.55

    return False, "no_detection", 0.0


def get_pdf_full_text(layout):
    """
    Get the complete text from all pages of a PDF layout.
    Returns a single string.
    """
    return "\n".join(page["text"] for page in layout)


# ─── v5.7: AI Decision Logic Helpers ────────────────────────────────────────

def _get_ai_decision_string(item_id, doc_status, pdf_status, ai_result, is_disabled=False):
    """
    Determine the AI status string for display.
    Returns: "YES", "NO", "SKIPPED", "DISABLED", "ERROR"
    """
    if is_disabled:
        return "DISABLED"
    if ai_result is None:
        return "SKIPPED"
    status = ai_result.get("ai_status", "SKIPPED")
    return status


def _compute_item_vote(item_id, doc_status, pdf_status, ai_result, required, is_optional_item=False):
    """
    v5.7: Per-item voting logic.
    
    DOC detection => detected_from_docx (bool)
    PDF detection => detected_from_pdf (bool)
    AI result     => {"exists": bool/None, "confidence": float, "ai_status": str}
    
    Returns:
        dict: {
            "final_source": str (e.g. "DOC + PDF", "DOC + AI", "PDF only", ...)
            "evidence_sources": list,
            "final_status": str ("PASS", "NEEDS_REVIEW", "FAILED"),
            "reason": str,
            "ai_display": str,
        }
    """
    doc_yes = bool(doc_status)
    pdf_yes = bool(pdf_status)
    ai_exists = ai_result.get("exists") if ai_result else None
    ai_confidence = ai_result.get("confidence", 0.0) if ai_result else 0.0
    ai_status_str = ai_result.get("ai_status", "SKIPPED") if ai_result else "SKIPPED"

    evidence_sources = []
    if doc_yes:
        evidence_sources.append("DOC")
    if pdf_yes:
        evidence_sources.append("PDF")
    if ai_exists is True and ai_confidence >= 0.5:
        evidence_sources.append("AI")
    elif ai_exists is True and ai_confidence < 0.5:
        evidence_sources.append("AI")  # Still count as evidence, just weak

    # ── Determine ai_display string ──
    if ai_status_str == "DISABLED":
        ai_display = "DISABLED"
    elif ai_status_str == "CACHED":
        ai_display = "YES" if ai_exists is True else "NO" if ai_exists is False else "CACHED"
    elif ai_status_str == "ERROR":
        ai_display = "ERROR"
    elif ai_status_str == "SKIPPED":
        ai_display = "SKIPPED"
    elif ai_status_str == "YES":
        ai_display = "YES"
    elif ai_status_str == "NO":
        ai_display = "NO"
    else:
        ai_display = ai_status_str

    # ── Voting logic ──
    if doc_yes and pdf_yes:
        # Strongest case
        final_source = "DOC + PDF"
        final_status = "PASS"
        reason = f"{item_id.title()} confirmed by both DOC and PDF detection."
    elif doc_yes and ai_exists is True and ai_confidence >= 0.5:
        final_source = "DOC + AI"
        final_status = "NEEDS_REVIEW"
        reason = f"{item_id.title()} confirmed by DOC and AI, but PDF/layout detection failed."
    elif doc_yes and not pdf_yes and (ai_exists is None or ai_exists is False):
        final_source = "DOC only"
        final_status = "NEEDS_REVIEW" if required else "PASS"
        reason = f"{item_id.title()} detected in DOC, but PDF and AI did not confirm."
    elif pdf_yes and ai_exists is True:
        final_source = "PDF + AI"
        final_status = "NEEDS_REVIEW"
        reason = f"{item_id.title()} detected by PDF and AI, but DOC source text did not confirm."
    elif pdf_yes and not doc_yes and (ai_exists is None or ai_exists is False):
        final_source = "PDF only"
        final_status = "NEEDS_REVIEW" if required else "PASS"
        reason = f"{item_id.title()} detected in PDF layout, but DOC and AI did not confirm."
    elif not doc_yes and not pdf_yes and ai_exists is True:
        final_source = "AI only"
        final_status = "NEEDS_REVIEW" if required else "PASS"
        reason = f"{item_id.title()} confirmed by AI only. Deterministic detectors failed."
    elif not doc_yes and not pdf_yes and (ai_exists is False or ai_exists is None):
        final_source = "None"
        final_status = "FAILED" if required else "PASS"
        reason = f"{item_id.title()} not found by DOC, PDF, or AI."
    else:
        final_source = " + ".join(evidence_sources) if evidence_sources else "None"
        final_status = "NEEDS_REVIEW" if required else "PASS"
        reason = f"{item_id.title()} detection inconclusive."

    return {
        "final_source": final_source,
        "evidence_sources": evidence_sources,
        "final_status": final_status,
        "reason": reason,
        "ai_display": ai_display,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# v5.7.1 MAIN: DOCX-First Structure Screening with Multi-Source AI Evaluator
# ═══════════════════════════════════════════════════════════════════════════════

def run_docx_first_screening(source_file, pdf_path, rules, log_callback=None):
    """
    v5.7.1: Run DOCX-first structure screening with PDF layout verification
    and AI as MULTI-SOURCE EVIDENCE EVALUATOR.

    This is the v5.7.1 main screening function that:
    1. Extracts source text from original file (DOCX XML parsing)
    2. Detects sections from source text (more reliable)
    3. Runs PDF layout analysis for layout/formatting
    4. Merges: structure from DOCX, layout from PDF
    5. Runs AI as MULTI-SOURCE EVALUATOR when DOC/PDF conflict
       (AI receives BOTH DOC snippet AND PDF snippet simultaneously)
    6. Uses voting logic: DOC + PDF + AI = explainable final status
    7. Handles graphical abstract with vision model
    8. Tracks evidence_sources list + new AI fields (preferred_source, likely_issue)
    9. Enforces AI rate limiting + caching

    Args:
        source_file: Path to original source file (DOCX/DOC/PDF)
        pdf_path: Path to converted PDF for layout analysis
        rules: List of template rules
        log_callback: Optional log function

    Returns:
        List of item result dicts with v5.7.1 fields.
    """
    def log(msg):
        if log_callback:
            log_callback(msg)

    log(f"[SCREENING v5.7.1] DOCX-first multi-source screening for {Path(source_file).name}")

    # Reset per-submission AI rate limits
    _reset_ai_rate_limits()

    # ── Step 1: Extract source text from original file ──
    source_text = extract_source_text(source_file)
    if source_text:
        log(f"[SCREENING v5.7.1] Extracted {len(source_text)} chars from source file")
    else:
        log(f"[SCREENING v5.7.1] No source text extracted (falling back to PDF-only)")

    # ── Step 2: DOCX source text section detection ──
    source_detection = None
    if source_text:
        source_detection = source_text_section_detector(source_text, rules)
        detected_items = list(source_detection.get("items", {}).keys())
        if detected_items:
            log(f"[SCREENING v5.7.1] DOCX detected sections: {', '.join(detected_items)}")
        else:
            log(f"[SCREENING v5.7.1] No sections detected from DOCX source")
    
    # ── Step 3: Parse PDF for layout analysis ──
    layout = extract_pdf_layout(pdf_path)
    if not layout:
        return {"error": "Failed to parse PDF"}
    
    pdf_text = get_pdf_full_text(layout)
    
    # ── Step 4: Run existing PDF-based screening for comparison ──
    pdf_results = run_single_screening(pdf_path, rules, log_callback=log_callback)
    
    if isinstance(pdf_results, dict) and "error" in pdf_results:
        return pdf_results
    
    # ── Step 5: Merge DOCX structure detection + AI verification ──
    for item_result in pdf_results:
        item_id = item_result.get("id")
        rule = next((r for r in rules if r.get("id") == item_id), {})
        required = rule.get("required", False)
        
        # --- Default fields ---
        item_result["detected_from_docx"] = False
        item_result["detected_from_pdf"] = item_result.get("structure_status") == "PASS"
        item_result["detected_from_ai"] = False
        item_result["final_source_used"] = "pdf_text"  # Keep for backward compat
        item_result["evidence_sources"] = ["PDF"]
        item_result["ai_status"] = "SKIPPED"
        item_result["ai_confidence"] = 0.0
        item_result["ai_evidence"] = ""
        
        # --- Step 5a: DOCX override (same as v5.6.6) ---
        doc_override = False
        if source_detection and item_id in source_detection.get("items", {}):
            docx_info = source_detection["items"][item_id]
            
            if docx_info.get("heading_found"):
                doc_override = True
                item_result["detected_from_docx"] = True
                item_result["final_source_used"] = "source_docx_text"
                item_result["evidence_sources"] = ["DOC"]
                
                item_result["evidence_text"] = (
                    f"Matched heading: \"{docx_info.get('heading_text', '')[:100]}\" | "
                    f"Detection method: source_docx_text"
                )
                item_result["detection_method"] = "source_docx_text"
                item_result["confidence"] = max(item_result.get("confidence", 0), 0.85)
                item_result["candidate_confidence"] = max(item_result.get("candidate_confidence", 0), 0.85)
                
                # Structure: PASS (DOCX proves it exists)
                item_result["structure_status"] = "PASS"
                
                # Layout: keep PDF result or WARNING if PDF didn't find it
                if item_result.get("layout_status") == "FAILED":
                    item_result["layout_status"] = "WARNING"
                
                log(f"[SCREENING v5.7.1] {item_id}: DOCX source text found: "
                    f"\"{docx_info.get('heading_text', '')[:50]}\"")
        
        # --- Step 5b: Determine DOC status for voting ---
        doc_status = item_result.get("detected_from_docx", False)
        pdf_status = item_result.get("detected_from_pdf", False)
        
        # PDF layout status from PDF-only run
        pdf_structure_ok = item_result.get("structure_status") == "PASS"
        
        # --- Step 5c: AI participation logic ---
        # v5.7: AI runs when DOC/PDF conflict, not only when both FAILED
        should_run_ai = False
        ai_run_reason = ""
        
        if not USE_AI_VERIFIER:
            should_run_ai = False
            ai_run_reason = "AI disabled"
        elif item_id in ("acknowledgement",):
            # Skip AI for optional items that are low-value for AI verification
            should_run_ai = False
            ai_run_reason = "Optional low-priority item"
        elif doc_override and pdf_status:
            # DOC = YES and PDF = YES with strong confidence -> SKIP AI
            should_run_ai = False
            ai_run_reason = "DOC and PDF already agree"
        elif doc_status and pdf_status:
            should_run_ai = False
            ai_run_reason = "DOC and PDF already agree"
        elif doc_status and not pdf_status:
            # DOC = YES, PDF = NO -> RUN AI
            should_run_ai = True
            ai_run_reason = "DOC detected, PDF failed"
        elif not doc_status and pdf_status:
            # DOC = NO, PDF = YES -> RUN AI
            should_run_ai = True
            ai_run_reason = "PDF detected, DOC failed"
        elif not doc_status and not pdf_status:
            # Both failed, but only for required/important items
            if required or item_id in ("abstract", "abstrak", "title", "references", "graphical_abstract"):
                should_run_ai = True
                ai_run_reason = "Both DOC and PDF failed for required item"
            else:
                should_run_ai = False
                ai_run_reason = "Optional item, both DOC and PDF failed"
        
        # --- Step 5d: Execute AI if needed ---
        ai_result = None
        
        if should_run_ai:
            # Build snippet from source text or PDF text
            snippet = ""
            if source_text:
                snippet = source_text[:2000]
            elif pdf_text:
                snippet = pdf_text[:2000]
            
            # For section items with nearby content, use that
            if item_result.get("nearby_content"):
                snippet = item_result["nearby_content"]
            
            # Handle graphical abstract specially
            if item_id == "graphical_abstract":
                # Try to crop the image region
                vision_image_path = _crop_graphical_abstract_region(
                    pdf_path, layout, item_result.get("expected_region", {})
                )
                if vision_image_path and Path(vision_image_path).exists():
                    vision_context = ""
                    if source_text:
                        # Build redacted context: title + first 2 sentences of abstract + keywords
                        title_match = re.search(r'^(.+)$', source_text, re.MULTILINE)
                        if title_match:
                            vision_context += redact_sensitive(title_match.group(1))[:200] + "\n"
                        abstract_match = re.search(r'Abstract\b[^.]*\.[^.]+\.', source_text, re.IGNORECASE)
                        if abstract_match:
                            vision_context += redact_sensitive(abstract_match.group(0))[:300] + "\n"
                        kw_match = re.search(r'Keywords[^.]*', source_text, re.IGNORECASE)
                        if kw_match:
                            vision_context += redact_sensitive(kw_match.group(0))[:100]
                    vision_context = vision_context[:800]
                    
                    ai_result = run_ai_verifier(
                        item_id, "",
                        is_vision=True,
                        vision_image_path=vision_image_path,
                        vision_context_text=vision_context,
                    )
                else:
                    # No image found, skip vision AI
                    ai_result = None
                    log(f"[SCREENING v5.7.1] {item_id}: no cropped image available for vision AI")
            else:
                # v5.7.1: Normal text verification using MULTI-SOURCE AI
                # Build DOC snippet (from source_text or DOCX nearby content)
                doc_snippet_raw = ""
                if source_detection and item_id in source_detection.get("items", {}):
                    docx_info = source_detection["items"][item_id]
                    doc_snippet_raw = docx_info.get("nearby", "") or docx_info.get("evidence", "")
                if not doc_snippet_raw and source_text:
                    doc_snippet_raw = source_text[:1500]

                # Build PDF snippet (from nearby_content or pdf_text)
                pdf_snippet_raw = item_result.get("nearby_content", "") or ""
                if not pdf_snippet_raw and pdf_text:
                    pdf_snippet_raw = pdf_text[:1500]

                # Truncate snippets
                doc_snippet_raw = doc_snippet_raw[:AI_VERIFIER_MAX_SNIPPET_CHARS]
                pdf_snippet_raw = pdf_snippet_raw[:AI_VERIFIER_MAX_SNIPPET_CHARS]

                doc_conf = item_result.get("candidate_confidence", 0.0) if doc_status else 0.0
                pdf_conf = item_result.get("candidate_confidence", 0.0) if pdf_status else 0.0

                has_content = (
                    (doc_snippet_raw and len(doc_snippet_raw.strip()) > 30) or
                    (pdf_snippet_raw and len(pdf_snippet_raw.strip()) > 30)
                )
                if has_content:
                    ai_result = run_ai_verifier_multi_source(
                        item_id,
                        doc_found=doc_status,
                        doc_confidence=doc_conf,
                        doc_snippet=doc_snippet_raw,
                        pdf_found=pdf_status,
                        pdf_confidence=pdf_conf,
                        pdf_snippet=pdf_snippet_raw,
                    )
                    log(f"[SCREENING v5.7.1] {item_id}: multi-source AI => "
                        f"exists={ai_result.get('exists')}, "
                        f"preferred={ai_result.get('preferred_source','?')}, "
                        f"issue={ai_result.get('likely_issue','?')}")
                else:
                    ai_result = None
                    log(f"[SCREENING v5.7.1] {item_id}: no safe snippet for multi-source AI verification")
        
        # --- Step 5e: Apply voting logic ---
        doc_found = item_result.get("detected_from_docx", False)
        pdf_found = item_result.get("detected_from_pdf", False)
        
        # Normalize PDF structure status
        if item_result.get("structure_status") == "PASS" and not doc_found:
            pdf_found = True
        
        vote = _compute_item_vote(
            item_id,
            doc_found,
            pdf_found,
            ai_result,
            required,
            is_optional_item=not required
        )
        
        # Update item result with v5.7.1 fields
        item_result["evidence_sources"] = vote["evidence_sources"]
        item_result["final_source_used"] = vote["final_source"]
        item_result["ai_status"] = vote["ai_display"]
        item_result["ai_confidence"] = round(ai_result.get("confidence", 0.0), 2) if ai_result else 0.0
        item_result["ai_evidence"] = ai_result.get("evidence", "") if ai_result else ""
        # v5.7.1: New multi-source AI fields
        item_result["ai_preferred_source"] = ai_result.get("preferred_source", "NONE") if ai_result else "NONE"
        item_result["ai_likely_issue"] = ai_result.get("likely_issue", "none") if ai_result else "none"
        item_result["ai_conflict_reason"] = ai_result.get("conflict_reason", "") if ai_result else ""
        item_result["ai_evidence_summary"] = ai_result.get("evidence_summary", "") if ai_result else ""

        if ai_result and ai_result.get("exists") is True:
            item_result["detected_from_ai"] = True

        # Override status based on voting
        if vote["final_status"] == "PASS":
            item_result["structure_status"] = "PASS"
            item_result["status"] = "PASS"
        elif vote["final_status"] == "NEEDS_REVIEW":
            item_result["structure_status"] = "WARNING"
            item_result["status"] = "WARNING"
        elif vote["final_status"] == "FAILED":
            item_result["structure_status"] = "FAILED"
            item_result["status"] = "FAILED"

        item_result["notes"] = vote["reason"]

        # v5.7.1: Enrich notes with AI conflict explanation when present
        if ai_result and ai_result.get("likely_issue") not in (None, "none", ""):
            issue = ai_result.get("likely_issue", "")
            preferred = ai_result.get("preferred_source", "")
            if issue and issue != "none":
                item_result["notes"] += f" [AI: {issue}, preferred={preferred}]"

        log_msg = (
            f"[SCREENING v5.7.1] {item_id}: DOC={doc_found}, PDF={pdf_found}, "
            f"AI={vote['ai_display']} preferred={item_result.get('ai_preferred_source','?')} "
            f"issue={item_result.get('ai_likely_issue','?')} "
            f"=> {vote['final_source']} => {vote['final_status']}"
        )
        log(log_msg)
    
    return pdf_results


def _crop_graphical_abstract_region(pdf_path, layout, expected_region):
    """
    Crop the graphical abstract region from the PDF and save as PNG.
    Returns path to cropped image or None.
    """
    if not fitz:
        return None
    if not expected_region:
        return None
    
    from pathlib import Path
    
    reports_dir = TEMPLATE_SCREENING_CONFIG.get("reports_dir", "")
    if not reports_dir:
        return None
    
    crops_dir = Path(reports_dir) / "crops"
    crops_dir.mkdir(parents=True, exist_ok=True)
    
    # Find the page where graphical abstract likely is
    target_page = expected_region.get("page", 1)
    
    try:
        doc = fitz.open(pdf_path)
        if len(doc) < target_page:
            doc.close()
            return None
        
        page = doc[target_page - 1]
        
        # Crop region (normalized -> absolute)
        pw = page.rect.width
        ph = page.rect.height
        
        x0 = expected_region.get("x_min", 0) * pw
        y0 = expected_region.get("y_min", 0) * ph
        x1 = expected_region.get("x_max", 1) * pw
        y1 = expected_region.get("y_max", 1) * ph
        
        # Ensure minimum size
        if (x1 - x0) < 50 or (y1 - y0) < 50:
            doc.close()
            return None
        
        # Render the page and crop
        mat = fitz.Matrix(2, 2)  # 2x zoom for better quality
        pix = page.get_pixmap(matrix=mat)
        
        # Crop pixmap to region
        crop_rect = fitz.Rect(x0, y0, x1, y1) * mat  # Scale by matrix
        crop_rect = crop_rect.irect  # Convert to integer rect
        
        # Ensure crop rect is within page bounds
        crop_rect.x0 = max(0, int(crop_rect.x0))
        crop_rect.y0 = max(0, int(crop_rect.y0))
        crop_rect.x1 = min(pix.width, int(crop_rect.x1))
        crop_rect.y1 = min(pix.height, int(crop_rect.y1))
        
        if crop_rect.x1 <= crop_rect.x0 or crop_rect.y1 <= crop_rect.y0:
            doc.close()
            return None
        
        # Create cropped pixmap
        clip = fitz.Rect(crop_rect)
        cropped_pix = page.get_pixmap(matrix=mat, clip=clip)
        
        # Save
        source_path = Path(pdf_path)
        safe_stem = re.sub(r'[^\w\-_]', '_', source_path.stem)[:50]
        output_path = crops_dir / f"{safe_stem}_graphical_abstract.png"
        
        cropped_pix.save(str(output_path))
        doc.close()
        
        if output_path.exists():
            return str(output_path)
        
    except Exception as e:
        print(f"[CROP] Error cropping graphical abstract: {e}")
        try:
            doc.close()
        except Exception:
            pass
    
    return None


# ─── Screening Engine Core ───────────────────────────────────────────────────

def load_template_rules(rules_path):
    """Load template rules from JSON file."""
    with open(rules_path, "r", encoding="utf-8") as f:
        return json.load(f)


def detect_path_column(headers):
    """
    Detect which column header contains the file path.
    Returns the column name or None.
    """
    possible_names = [
        "download_path",
        "download location",
        "download_location",
        "file_path",
        "article_file_path",
        "manuscript_path",
        "downloaded_file",
    ]
    for h in headers:
        if h is not None:
            h_stripped = str(h).strip().lower()
            for name in possible_names:
                if h_stripped == name.lower():
                    return str(h).strip()
    return None


def detect_column_index(headers, possible_names):
    """Find the first matching column index for given names."""
    for name in possible_names:
        for i, h in enumerate(headers):
            if h is not None and str(h).strip().lower() == name.lower():
                return i
    return None


def count_references(text):
    """Count number of reference entries in text."""
    patterns = [
        r"^\[?\d+\]?\s",  # [1] or 1. at start of line
        r"^\d+\.\s",       # 1. at start
        r"^[A-Z][a-zà-ü]+.*\(?\d{4}\)?\.",  # Author (Year).
    ]
    count = 0
    lines = text.split("\n")
    for line in lines:
        line = line.strip()
        if not line:
            continue
        for pat in patterns:
            if re.match(pat, line):
                count += 1
                break
    return max(count, 0)


def count_keywords(text):
    """Count number of keywords (comma/semicolon separated)."""
    # Find the line containing Keywords/Kata kunci and count items after it
    kw_pattern = re.compile(
        r"(Keywords|Kata kunci|Kata Kunci)\s*[:\-–]?\s*(.+)", re.IGNORECASE
    )
    match = kw_pattern.search(text)
    if not match:
        return 0

    kw_text = match.group(2)
    # Split by comma, semicolon, or bullet
    parts = re.split(r"[;,•·\-•]+", kw_text)
    parts = [p.strip() for p in parts if p.strip() and len(p.strip()) > 1]
    return len(parts)


def count_keyword_terms(text, label):
    """
    Count keyword terms after a specific label (Keywords / Kata kunci).
    Returns the count.
    """
    patterns = [
        re.compile(rf"{re.escape(label)}\s*[:\-–]?\s*(.+)", re.IGNORECASE),
    ]
    for pat in patterns:
        match = pat.search(text)
        if match:
            kw_text = match.group(1)
            parts = re.split(r"[;,•·\-•]+", kw_text)
            parts = [p.strip() for p in parts if p.strip() and len(p.strip()) > 1]
            return len(parts)
    return 0


def run_single_screening(pdf_path, rules, log_callback=None):
    """
    Run template screening on a single PDF file.
    Returns list of item result dicts.
    
    LEGACY v2.1 function - kept for backward compatibility and
    used by run_docx_first_screening() for PDF layout analysis.
    
    v5.7 callers should use run_docx_first_screening() instead.
    """
    def log(msg):
        if log_callback:
            log_callback(msg)

    # Parse PDF
    layout = extract_pdf_layout(pdf_path)
    if not layout:
        return {"error": "Failed to parse PDF"}

    results = []
    total_pages = len(layout)

    for rule in rules:
        item_id = rule["id"]
        label = rule["label"]
        required = rule.get("required", False)
        text_pattern = rule.get("text_pattern", "")
        expected_region = rule.get("expected_region", {})
        tolerance = rule.get("tolerance", 0.15)
        format_rules = rule.get("format_rules")
        min_ref_count = rule.get("min_reference_count", 0)
        min_kw_count = rule.get("min_keyword_count", 0)

        is_generic_text = (text_pattern == ".+")

        item_result = {
            "id": item_id,
            "label": label,
            "status": "FAILED",
            "page": "",
            "expected_region": expected_region,
            "found_region": None,
            "evidence_text": "",
            "notes": "",
            "formatting_ok": None,
            "confidence": 0.0,
            "candidate_confidence": 0.0,
            "rejected_false_positives": [],
            "all_candidates": [],
            "structure_status": "FAILED",
            "layout_status": "FAILED",
            "formatting_status": "N/A",
            "detection_method": "not_found",
            "nearby_content": "",
            # v5.7 default fields
            "detected_from_docx": False,
            "detected_from_pdf": False,
            "detected_from_ai": False,
            "final_source_used": "pdf_text",
            "evidence_sources": [],
            "ai_status": "SKIPPED",
            "ai_confidence": 0.0,
            "ai_evidence": "",
        }

        # ── Find best candidate using item-aware candidate ranking ──
        candidate = find_best_candidate_for_rule(rule, layout)
        confidence = candidate.get("confidence", 0.0)
        item_result["candidate_confidence"] = round(confidence, 2)
        item_result["rejected_false_positives"] = candidate.get("rejected_false_positives", [])
        item_result["detection_method"] = candidate.get("detection_method", "not_found")
        item_result["nearby_content"] = candidate.get("nearby_content", "")
        all_cands = candidate.get("all_candidates", [])
        item_result["all_candidates"] = [
            {"text": c.get("text", "")[:100], "score": round(c.get("score", 0), 2)}
            for c in all_cands
        ]

        candidate_text = candidate.get("candidate_text", "")
        found_bbox = candidate.get("bbox")
        page = candidate.get("page", 0)
        found_block = candidate.get("block")
        found_span = candidate.get("span")
        nearby_content = candidate.get("nearby_content", "")

        # ── DETECTION METHOD assignment ──
        detection_method = candidate.get("detection_method", "not_found")

        # ── HEADING + CONTENT VERIFICATION FOR SECTION ITEMS ──
        section_items = {
            "corresponding_author", "graphical_abstract", "abstract", "keywords",
            "abstrak", "kata_kunci", "introduction", "methodology",
            "results_discussion", "conclusion", "acknowledgement", "references"
        }

        is_section_item = item_id in section_items

        # Determine if heading was found
        heading_found = bool(candidate_text) and confidence > 0.0

        # For section items with patterns (not generic), extract nearby content if not already done
        if is_section_item and not is_generic_text and heading_found and not nearby_content:
            # Try to extract using extract_following_paragraph for better accuracy
            for page_data in layout:
                if candidate_text.lower() in page_data["text"].lower():
                    nearby_content = extract_following_paragraph(
                        page_data["blocks"], candidate_text, page_data
                    )
                    if nearby_content:
                        break

        # ── CONTENT VERIFICATION ──
        content_valid = False
        content_verified = False

        if is_section_item and heading_found:
            if item_id == "graphical_abstract":
                # Graphical abstract: check for image in expected region
                if expected_region:
                    re_bbox = (
                        expected_region.get("x_min", 0),
                        expected_region.get("y_min", 0),
                        expected_region.get("x_max", 1),
                        expected_region.get("y_max", 1),
                    )
                    page_data = layout[0] if layout else None
                    if page_data:
                        content_valid, img_method, img_conf = detect_graphical_region(
                            page_data, re_bbox
                        )
                        content_verified = True
            elif item_id == "corresponding_author":
                content_valid = looks_like_corresponding_author_content(nearby_content)
                content_verified = bool(nearby_content)
            elif item_id in ("abstract", "abstrak"):
                content_valid = looks_like_abstract_content(nearby_content)
                content_verified = bool(nearby_content)
            elif item_id in ("keywords", "kata_kunci"):
                content_valid = looks_like_keywords_content(nearby_content)
                content_verified = bool(nearby_content)
            elif item_id == "references":
                content_valid = looks_like_references_content(nearby_content)
                content_verified = bool(nearby_content)
            else:
                # Generic sections: content is considered valid if any nearby text exists
                content_valid = bool(nearby_content) and len(nearby_content) > 50
                content_verified = bool(nearby_content)

        # ── STRUCTURE EVALUATION ──
        # Structure = does the component exist with valid content?
        structure_ok = False
        structure_reason = ""

        if is_section_item:
            if heading_found and content_valid:
                # Heading exists AND content is verified
                structure_ok = True
                item_result["structure_status"] = "PASS"
                structure_reason = f"Heading found and content verified ({item_id})."
                confidence = max(confidence, 0.80)
            elif heading_found and not content_verified:
                # Heading exists but no content area found to verify (e.g., text extraction issue)
                structure_ok = True
                item_result["structure_status"] = "WARNING"
                structure_reason = "Heading found but content could not be verified."
                confidence = max(confidence, 0.50)
            elif heading_found and not content_valid and content_verified:
                # Heading exists but content fails validation
                structure_ok = True
                item_result["structure_status"] = "WARNING"
                structure_reason = "Heading found but content could not be verified."
                confidence = max(confidence, 0.50)
            elif not heading_found and candidate.get("detection_method") == "semantic_fallback":
                # No heading but semantic content matches
                structure_ok = True
                item_result["structure_status"] = "WARNING"
                structure_reason = "Semantic content matched but heading not found."
                detection_method = "semantic_fallback"
                confidence = max(confidence, 0.70)
            else:
                # Neither heading nor content found
                structure_ok = False
                item_result["structure_status"] = "FAILED"
                structure_reason = f"Required item '{label}' heading and content not found."
        else:
            # Metadata items (title, authors, affiliations)
            if is_generic_text:
                structure_ok = confidence >= 0.45
            else:
                structure_ok = heading_found

            if structure_ok:
                item_result["structure_status"] = "PASS"
            else:
                item_result["structure_status"] = "FAILED"

        # ── LAYOUT / REGION EVALUATION ──
        # Layout mismatch becomes WARNING, not FAILED
        if found_bbox and expected_region:
            passed, deviation = check_region(found_bbox, expected_region, tolerance)
            if passed and deviation <= tolerance:
                item_result["layout_status"] = "PASS"
            else:
                item_result["layout_status"] = "WARNING"  # Layout mismatch is WARNING, not FAILED
            item_result["found_region"] = {
                "x_min": round(found_bbox[0], 4),
                "y_min": round(found_bbox[1], 4),
                "x_max": round(found_bbox[2], 4),
                "y_max": round(found_bbox[3], 4),
            }
        else:
            passed = False
            deviation = 1.0
            if structure_ok:
                item_result["layout_status"] = "WARNING"  # Structure exists but layout unknown
            else:
                item_result["layout_status"] = "FAILED"

        # ── FORMATTING EVALUATION ──
        fmt_ok, fmt_details = check_formatting(found_span, format_rules)
        item_result["formatting_ok"] = fmt_ok
        item_result["formatting_status"] = "PASS" if fmt_ok else ("WARNING" if fmt_ok is False else "N/A")

        # ── OVERALL ITEM STATUS ──
        if structure_ok and confidence >= 0.80:
            item_result["status"] = "PASS"
            if not item_result.get("notes"):
                item_result["notes"] = f"Found with high confidence ({confidence:.2f}). {structure_reason}"
        elif structure_ok and confidence >= 0.50:
            item_result["status"] = "WARNING"
            if not item_result.get("notes"):
                item_result["notes"] = f"Found but with issues ({confidence:.2f}). {structure_reason}"
        elif structure_ok and confidence > 0:
            item_result["status"] = "WARNING"
            if not item_result.get("notes"):
                item_result["notes"] = f"Found but confidence low ({confidence:.2f}). {structure_reason}"
        else:
            item_result["status"] = "FAILED"
            if not item_result.get("notes"):
                item_result["notes"] = f"Required item '{label}' not found in PDF."

        # ── EVIDENCE TEXT ──
        evidence_parts = []
        if candidate_text:
            evidence_parts.append(f"Matched heading: \"{candidate_text[:100]}\"")
        if nearby_content:
            evidence_parts.append(f"Nearby content: \"{nearby_content[:200]}\"")
        evidence_parts.append(f"Detection method: {detection_method}")
        item_result["evidence_text"] = " | ".join(evidence_parts)

        if candidate_text:
            item_result["page"] = str(page)

        # If rejected some false positives, mention that
        rejected = candidate.get("rejected_false_positives", [])
        if rejected:
            rejection_note = f" Rejected template labels: {', '.join(rejected[:3])}"
            if item_result["notes"]:
                item_result["notes"] += rejection_note
            else:
                item_result["notes"] = rejection_note.strip()

        # ── SPECIAL HANDLING: ABSTRACT ──
        if item_id == "abstract":
            # Check for "Graphical abstract" confusion
            full_text = " ".join(p["text"] for p in layout)
            has_graphical = "graphical abstract" in full_text.lower()
            has_standalone_abstract = bool(re.search(
                r'(?<![Gg]raphical )(?:^|\n)\s*Abstract\b',
                full_text,
                re.MULTILINE
            ))

            if has_graphical and not has_standalone_abstract and not heading_found:
                item_result["status"] = "FAILED"
                item_result["structure_status"] = "FAILED"
                item_result["notes"] = ("Only 'Graphical abstract' heading found, "
                                        "not standalone 'Abstract'. Note: 'Graphical abstract' "
                                        "is a separate component and does not fulfill the "
                                        "English Abstract requirement.")

        # ── SPECIAL HANDLING: GRAPHICAL ABSTRACT ──
        if item_id == "graphical_abstract":
            if heading_found and expected_region:
                page_data = layout[0]  # Page 1
                re_bbox = (
                    expected_region.get("x_min", 0),
                    expected_region.get("y_min", 0),
                    expected_region.get("x_max", 1),
                    expected_region.get("y_max", 1),
                )
                img_detected, img_method, img_confidence = detect_graphical_region(
                    page_data, re_bbox
                )

                if img_detected:
                    item_result["status"] = "PASS"
                    item_result["notes"] = (
                        f"Graphical abstract heading found and image/region "
                        f"detected via {img_method} (confidence: {img_confidence:.2f})."
                    )
                elif img_confidence >= 0.55:
                    item_result["status"] = "WARNING"
                    item_result["notes"] = (
                        "Graphical abstract heading found but no clear image "
                        "detected. Layout suggests graphical region may exist."
                    )
                else:
                    item_result["status"] = "WARNING"
                    item_result["notes"] = (
                        "Graphical abstract heading found but no image/object "
                        "detected in expected region. Treating as WARNING since "
                        "heading exists but image extraction may have failed."
                    )

        # ── SPECIAL HANDLING: KEYWORDS ──
        if item_id == "keywords" and structure_ok:
            full_text = " ".join(p["text"] for p in layout)
            kw_count = count_keyword_terms(full_text, "Keywords")
            item_result["evidence_text"] += f" | Keywords found: {kw_count}"
            if min_kw_count > 0 and kw_count < min_kw_count:
                item_result["status"] = "WARNING"  # Changed from FAILED to WARNING
                item_result["notes"] = (
                    f"Keywords heading found but only {kw_count} keyword terms "
                    f"detected (minimum {min_kw_count} required)."
                )
            else:
                item_result["notes"] = f"Keywords found: {kw_count} terms."

        # ── SPECIAL HANDLING: KATA KUNCI ──
        if item_id == "kata_kunci" and structure_ok:
            full_text = " ".join(p["text"] for p in layout)
            kk_count = count_keyword_terms(full_text, "Kata kunci")
            item_result["evidence_text"] += f" | Kata kunci found: {kk_count}"
            if min_kw_count > 0 and kk_count < min_kw_count:
                item_result["status"] = "WARNING"  # Changed from FAILED to WARNING
                item_result["notes"] = (
                    f"Kata kunci heading found but only {kk_count} terms "
                    f"detected (minimum {min_kw_count} required)."
                )
            else:
                item_result["notes"] = f"Kata kunci found: {kk_count} terms."

        # ── SPECIAL HANDLING: ABSTRAK ──
        if item_id == "abstrak":
            full_text = " ".join(p["text"] for p in layout)
            has_abstrak = bool(re.search(r'(?:^|\n)\s*Abstrak\b', full_text, re.MULTILINE | re.IGNORECASE))
            if not has_abstrak and not heading_found:
                # Don't fail if content exists semantically
                if not looks_like_abstract_content(full_text):
                    item_result["status"] = "FAILED"
                    item_result["structure_status"] = "FAILED"
                    item_result["notes"] = "Malay Abstract (Abstrak) heading not found in PDF."

        # ── SPECIAL HANDLING: REFERENCES ──
        if item_id == "references" and structure_ok and min_ref_count > 0:
            full_text = " ".join(p["text"] for p in layout)
            ref_count = count_references(full_text)
            item_result["evidence_text"] += f" | Reference entries found: {ref_count}"
            if ref_count < min_ref_count:
                item_result["status"] = "WARNING"  # Changed from FAILED to WARNING
                item_result["notes"] = (
                    f"Only {ref_count} references found (minimum {min_ref_count} required). "
                    f"Using WARNING status since heading exists."
                )

        # ── STRUCTURE STATUS OVERRIDE ──
        if item_result["structure_status"] == "PASS":
            pass  # Already set by heading+content verification
        elif confidence >= 0.45:
            item_result["structure_status"] = "PASS"
        elif candidate_text and confidence > 0:
            item_result["structure_status"] = "WARNING"

        # Track that PDF detected this
        if item_result["structure_status"] == "PASS":
            item_result["detected_from_pdf"] = True

        results.append(item_result)

    return results


def _describe_region(region):
    """Convert normalized region dict to human-readable description."""
    if not region:
        return "Not detected"
    x_min = region.get("x_min", 0)
    y_min = region.get("y_min", 0)
    x_max = region.get("x_max", 1)
    y_max = region.get("y_max", 1)

    # Horizontal position
    if x_min < 0.15:
        h_pos = "left column"
    elif x_min < 0.45:
        h_pos = "left-center area"
    elif x_min < 0.65:
        h_pos = "center area"
    elif x_min < 0.85:
        h_pos = "right-center area"
    else:
        h_pos = "right column"

    # Vertical position
    if y_min < 0.15:
        v_pos = "top area"
    elif y_min < 0.35:
        v_pos = "upper area"
    elif y_min < 0.55:
        v_pos = "upper-middle area"
    elif y_min < 0.75:
        v_pos = "lower-middle area"
    else:
        v_pos = "lower area"

    return f"Page 1, {h_pos}, {v_pos}"


# ═══════════════════════════════════════════════════════════════════════════════
# v5.7: PDF Report Generation
# ═══════════════════════════════════════════════════════════════════════════════

def generate_screening_pdf_report(report_data, excel_row_data, output_dir, log_callback=None):
    """
    v5.7: Generate screening PDF report with updated table format.
    
    Each item shows: Item | DOC Detection | PDF Detection | AI Detection |
                     Final Source | Final Status | Reason
    
    Args:
        report_data: dict with screening results (from JSON)
        excel_row_data: dict with metadata (submission_no, author_name, title, etc.)
        output_dir: directory to save PDF
        log_callback: optional log function

    Returns:
        Path to combined PDF, or None on failure
    """
    def log(msg):
        if log_callback:
            log_callback(msg)

    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm, cm
        from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            PageBreak, Frame, PageTemplate, BaseDocTemplate
        )
        from reportlab.platypus.flowables import HRFlowable
    except ImportError:
        log("[SCREENING] ❌ reportlab not installed. Cannot generate PDF report.")
        return None

    sub_no = report_data.get("submission_no", excel_row_data.get("submission_no", "N/A"))
    timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")

    # ── v5.7: Organized file paths ──
    reports_root = Path(output_dir)
    pdf_dir = reports_root / "pdf"
    combined_dir = reports_root / "combined"
    pdf_dir.mkdir(parents=True, exist_ok=True)
    combined_dir.mkdir(parents=True, exist_ok=True)

    report_filename = f"template_screening_report_{sub_no}_{timestamp_str}.pdf"
    combined_filename = f"template_screening_combined_{sub_no}_{timestamp_str}.pdf"

    report_file = pdf_dir / report_filename
    combined_file = combined_dir / combined_filename

    # ── Step 1: Generate report PDF ──
    temp_report_file = pdf_dir / f"_temp_report_{sub_no}_{timestamp_str}.pdf"

    doc = SimpleDocTemplate(
        str(temp_report_file),
        pagesize=A4,
        topMargin=2*cm,
        bottomMargin=2*cm,
        leftMargin=2*cm,
        rightMargin=2*cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle', parent=styles['Heading1'],
        fontSize=20, leading=24, spaceAfter=12,
        textColor=colors.HexColor("#003399"),
    )
    heading2 = ParagraphStyle(
        'H2', parent=styles['Heading2'],
        fontSize=14, leading=18, spaceBefore=16, spaceAfter=8,
        textColor=colors.HexColor("#003399"),
    )
    heading3 = ParagraphStyle(
        'H3', parent=styles['Heading3'],
        fontSize=11, leading=14, spaceBefore=10, spaceAfter=4,
        textColor=colors.HexColor("#0055aa"),
    )
    normal = ParagraphStyle(
        'CustomNormal', parent=styles['Normal'],
        fontSize=9, leading=13, spaceAfter=4,
    )
    small = ParagraphStyle(
        'Small', parent=styles['Normal'],
        fontSize=8, leading=10, spaceAfter=2,
    )
    bold_style = ParagraphStyle(
        'Bold', parent=normal, fontName='Helvetica-Bold',
    )

    elements = []

    # ───── 1. REPORT HEADER ─────
    elements.append(Paragraph("Template Screening Report (v5.7.1 - Multi-Source AI Evaluator)", title_style))
    elements.append(Paragraph(f"Journal Template: Jurnal Teknologi", normal))
    elements.append(Spacer(1, 4))
    elements.append(Paragraph(f"<b>Submission No:</b> {sub_no}", normal))
    elements.append(Paragraph(f"<b>Author Name:</b> {excel_row_data.get('author_name', 'N/A')}", normal))
    elements.append(Paragraph(f"<b>Article Title:</b> {excel_row_data.get('title', 'N/A')}", normal))
    elements.append(Paragraph(f"<b>Source File:</b> {Path(report_data.get('source_file', 'N/A')).name}", normal))
    elements.append(Paragraph(f"<b>Screened PDF:</b> {Path(report_data.get('pdf_file', 'N/A')).name}", normal))
    elements.append(Paragraph(f"<b>Screening Date:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", normal))
    elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#003399")))
    elements.append(Spacer(1, 8))

    # ───── v5.7: DEBUG REPORT - Detection Sources ─────
    elements.append(Paragraph("Detection Source Debug Report", heading2))
    items = report_data.get("items", [])
    
    # v5.7: Updated table headers
    debug_data = [["Item", "DOC Detection", "PDF Detection", "AI Detection", "Final Source"]]
    for item in items:
        detected_docx = "YES" if item.get("detected_from_docx") else "NO"
        detected_pdf = "YES" if item.get("detected_from_pdf") else "NO"
        detected_ai = item.get("ai_status", "SKIPPED")
        final_source = item.get("final_source_used", "N/A")
        
        debug_data.append([
            item.get("label", item.get("id", "")),
            detected_docx,
            detected_pdf,
            detected_ai,
            final_source,
        ])
    
    debug_table = Table(debug_data, colWidths=[80, 70, 70, 70, 100])
    debug_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#003399")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor("#f0f0f5"), colors.white]),
    ]))
    elements.append(debug_table)
    elements.append(Spacer(1, 4))
    
    # Summary
    docx_count = sum(1 for i in items if i.get("detected_from_docx"))
    pdf_count = sum(1 for i in items if i.get("detected_from_pdf"))
    ai_count = sum(1 for i in items if i.get("detected_from_ai"))
    elements.append(Paragraph(
        f"<b>Summary:</b> DOC detected {docx_count} items | PDF detected {pdf_count} items | "
        f"AI confirmed {ai_count} items",
        normal
    ))
    elements.append(Paragraph(
        "The system uses DOC + PDF + AI voting logic. AI serves as a third checker "
        "when DOC and PDF conflict. AI is skipped when DOC and PDF already agree.",
        small
    ))
    elements.append(Spacer(1, 8))

    # ───── 2. EXECUTIVE SUMMARY ─────
    final_status = report_data.get("final_status", "N/A")
    final_score = report_data.get("final_score", 0)
    struct_score = report_data.get("structure_score", 0)
    layout_score = report_data.get("layout_score", 0)
    fmt_score = report_data.get("formatting_score", 0)

    # Build critical/warning lists
    critical_failed_names = [
        r.get("label", r.get("id", "")) for r in items
        if r.get("status") == "FAILED" and r.get("id") in [
            "title", "abstract", "introduction", "methodology",
            "results_discussion", "conclusion", "references",
            "graphical_abstract"
        ]
    ]
    warning_names = [
        r.get("label", r.get("id", "")) for r in items
        if r.get("status") in ("WARNING", "NEEDS_REVIEW")
    ]

    status_colors_map = {
        "SCREENING PASS": colors.HexColor("#00aa00"),
        "SCREENING WARNING": colors.HexColor("#cc8800"),
        "SCREENING FAILED": colors.HexColor("#cc0000"),
        "SCREENING ERROR": colors.HexColor("#cc0000"),
        "PASS": colors.HexColor("#00aa00"),
        "NEEDS MANUAL REVIEW": colors.HexColor("#cc8800"),
        "REJECT": colors.HexColor("#cc0000"),
    }
    status_color = status_colors_map.get(final_status, colors.black)

    elements.append(Paragraph("Executive Summary", heading2))
    elements.append(Paragraph(f"<b>Overall Decision:</b> <font color='{status_color.hexval()}'>{final_status}</font>", normal))

    if "PASS" in final_status:
        main_reason = "The manuscript follows the main Jurnal Teknologi template structure, layout, and formatting."
        suggestion = "Proceed with the manuscript to the next workflow stage."
    elif "NEEDS" in final_status or "WARNING" in final_status:
        main_reason = "The manuscript generally follows the template but contains inconsistencies that require manual review."
        suggestion = "Editor should review the highlighted issues before proceeding."
    else:
        main_reason = "The manuscript is missing required template components or critically mismatched."
        suggestion = "Author should revise the manuscript using the official Jurnal Teknologi template before proceeding."

    elements.append(Paragraph(f"<b>Main Reason:</b> {main_reason}", normal))

    if critical_failed_names:
        elements.append(Paragraph("<b>Critical Failed Items:</b>", normal))
        for name in critical_failed_names:
            elements.append(Paragraph(f"&bull; {name}", normal))
    else:
        elements.append(Paragraph("<b>Critical Failed Items:</b> None", normal))

    if warning_names:
        elements.append(Paragraph("<b>Items Needing Review:</b>", normal))
        for name in warning_names[:5]:
            elements.append(Paragraph(f"&bull; {name}", normal))
    else:
        elements.append(Paragraph("<b>Items Needing Review:</b> None", normal))

    elements.append(Paragraph(f"<b>Suggested Action:</b> {suggestion}", normal))
    elements.append(Spacer(1, 8))

    # ───── 3. OVERALL RESULT ─────
    elements.append(Paragraph("Overall Result", heading2))

    result_data = [
        ["Metric", "Score / Value"],
        ["Final Status", final_status],
        ["Final Score", f"{final_score}/100"],
        ["Structure Score", f"{struct_score}%"],
        ["Layout Score", f"{layout_score}%"],
        ["Formatting Score", f"{fmt_score}%"],
        ["Detection Method", "DOCX-First + Multi-Source AI (v5.7.1)"],
        ["Decision", "ACCEPT" if "PASS" in final_status else "MANUAL REVIEW" if "NEEDS" in final_status else "REJECT"],
    ]
    result_table = Table(result_data, colWidths=[180, 200])
    result_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#003399")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#f0f0f5")),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor("#f0f0f5"), colors.white]),
    ]))
    elements.append(result_table)
    elements.append(Spacer(1, 8))

    # ───── v5.7: VOTING LOGIC EXPLANATION ─────
    elements.append(Paragraph("How Items Are Evaluated", heading2))
    vote_legend = [
        ["DOC + PDF + AI SKIPPED", "PASS", "Both deterministic detectors agree"],
        ["DOC + PDF + AI CONFIRMED", "PASS", "All three confirm"],
        ["DOC + AI", "NEEDS REVIEW", "DOC+AI confirm, but PDF layout conflicts"],
        ["PDF + AI", "NEEDS REVIEW", "PDF+AI confirm, but DOC source extraction fails"],
        ["DOC only", "NEEDS REVIEW", "DOC detects, but PDF+AI don't confirm"],
        ["PDF only", "NEEDS REVIEW", "PDF detects, but DOC+AI don't confirm"],
        ["AI only", "NEEDS REVIEW", "Only AI confirms (deterministic detectors failed)"],
        ["None", "FAILED", "All three fail to detect required item"],
    ]
    vote_table = Table(vote_legend, colWidths=[160, 80, 240])
    vote_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#003399")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor("#f0f0f5"), colors.white]),
    ]))
    elements.append(vote_table)
    elements.append(Spacer(1, 8))

    # ───── 5. STATUS EXPLANATION ─────
    elements.append(Paragraph("Status Explanation", heading2))
    elements.append(Paragraph("<b>PASS</b> &mdash; Item follows the template structure, position, and formatting.", normal))
    elements.append(Paragraph("<b>WARNING/NEEDS REVIEW</b> &mdash; Item exists but evidence is inconsistent or layout differs from template.", normal))
    elements.append(Paragraph("<b>FAILED</b> &mdash; Required item truly missing after DOC/PDF/AI checks.", normal))
    elements.append(Spacer(1, 8))

    # ───── v5.7: ITEM TABLE with full columns ─────
    elements.append(Paragraph("Item Screening Details", heading2))

    # Build the table with v5.7 columns
    item_table_data = [
        ["Item", "DOC", "PDF", "AI", "Final Source", "Status", "Reason"]
    ]
    for item in items:
        label = item.get("label", item.get("id", ""))
        doc_val = "YES" if item.get("detected_from_docx") else "NO"
        pdf_val = "YES" if item.get("detected_from_pdf") else "NO"
        ai_val = item.get("ai_status", "SKIPPED")
        final_src = item.get("final_source_used", "None")
        status = item.get("status", "FAILED")
        reason = item.get("notes", "")[:60]
        
        item_table_data.append([label, doc_val, pdf_val, ai_val, final_src, status, reason])

    item_table = Table(item_table_data, colWidths=[55, 30, 30, 40, 65, 55, 125])
    item_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#003399")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor("#f0f0f5"), colors.white]),
    ]))
    elements.append(item_table)
    elements.append(Spacer(1, 12))
    elements.append(PageBreak())

    # ───── 7. DETAILED FINDINGS ─────
    elements.append(Paragraph("Detailed Findings", heading2))

    from config import TEMPLATE_RULES_DIR
    rules_path = Path(TEMPLATE_RULES_DIR) / "jurnal_teknologi_rules.json"
    rules = []
    if rules_path.exists():
        try:
            with open(rules_path) as f:
                rule_config = json.load(f)
                rules = rule_config.get("rules", [])
        except Exception:
            pass

    rule_map = {r.get("id"): r for r in rules}

    for item in items:
        item_id = item.get("id", "")
        label = item.get("label", item_id)
        status = item.get("status", "")
        page_found = item.get("page", "")
        evidence = item.get("evidence_text", "")
        notes = item.get("notes", "")
        found_region = item.get("found_region", {})
        expected_region = item.get("expected_region", {})
        span_info = item.get("span", {})
        formatting_ok = item.get("formatting_ok")
        confidence = item.get("candidate_confidence", 0.0)
        rejected_fps = item.get("rejected_false_positives", [])
        all_cands = item.get("all_candidates", [])
        detection_method = item.get("detection_method", "not_found")
        # v5.7 fields
        detected_docx = item.get("detected_from_docx", False)
        detected_pdf = item.get("detected_from_pdf", False)
        detected_ai = item.get("detected_from_ai", False)
        ai_status = item.get("ai_status", "SKIPPED")
        ai_confidence = item.get("ai_confidence", 0.0)
        ai_evidence = item.get("ai_evidence", "")
        final_source = item.get("final_source_used", "N/A")
        evidence_sources = item.get("evidence_sources", [])

        # Color for status
        status_color_item = {
            "PASS": colors.HexColor("#00aa00"),
            "WARNING": colors.HexColor("#cc8800"),
            "FAILED": colors.HexColor("#cc0000"),
            "NEEDS_REVIEW": colors.HexColor("#cc8800"),
        }.get(status, colors.black)

        # Find rule for human-readable expected location
        rule = rule_map.get(item_id, {})
        expected_text = _describe_region(expected_region) if expected_region else "Not specified"
        found_text = _describe_region(found_region) if found_region else "Not detected"

        # Item header
        elements.append(Paragraph(f"<b>{label} [{item_id}]</b>", heading3))

        # A. Structure Evaluation (v5.7: DOC + PDF + AI)
        elements.append(Paragraph("<b>A. Structure Evaluation (DOC + PDF + AI)</b>", normal))
        if rule.get("required", True):
            expected_struct = f"The '{label}' {'section' if item_id != 'graphical_abstract' else 'component'} must exist."
        else:
            expected_struct = f"The '{label}' component is recommended."
        elements.append(Paragraph(f"<b>Expected:</b> {expected_struct}", small))
        if evidence:
            elements.append(Paragraph(f"<b>Found:</b> {evidence[:150]}", small))
        else:
            elements.append(Paragraph(f"<b>Found:</b> {'Detected on page ' + page_found if page_found else 'Not detected'}", small))
        elements.append(Paragraph(f"<b>Confidence:</b> {confidence:.2f}", small))
        if detection_method:
            elements.append(Paragraph(f"<b>Detection Method:</b> {detection_method}", small))

        # v5.7: DOC/PDF/AI detection display
        elements.append(Paragraph(f"<b>DOC Detection:</b> {'YES' if detected_docx else 'NO'}", small))
        elements.append(Paragraph(f"<b>PDF Detection:</b> {'YES' if detected_pdf else 'NO'}", small))
        elements.append(Paragraph(f"<b>AI Detection:</b> {ai_status}", small))
        if ai_confidence > 0:
            elements.append(Paragraph(f"<b>AI Confidence:</b> {ai_confidence:.2f}", small))
        if ai_evidence:
            elements.append(Paragraph(f"<b>AI Evidence:</b> {ai_evidence[:150]}", small))
        
        sources_str = " + ".join(evidence_sources) if evidence_sources else "None"
        elements.append(Paragraph(f"<b>Evidence Sources:</b> {sources_str}", small))
        elements.append(Paragraph(f"<b>Final Source Used:</b> {final_source}", small))

        if rejected_fps:
            elements.append(Paragraph(f"<b>Rejected false positives:</b> {', '.join(rejected_fps[:3])}", small))
        elements.append(Paragraph(f"<b>Status:</b> <font color='{status_color_item.hexval()}'>{status}</font>", small))

        if status == "PASS":
            reason = f"The required component '{label}' exists."
        elif status == "WARNING":
            reason = f"Component found but evidence is inconsistent."
        elif status == "FAILED":
            if "not found" in notes.lower():
                reason = f"Required component '{label}' was not found."
            else:
                reason = notes if notes else f"Component '{label}' failed validation."
        else:
            reason = notes
        elements.append(Paragraph(f"<b>Reason:</b> {reason}", small))

        # B. Layout / Coordinate Evaluation
        elements.append(Paragraph("<b>B. Layout / Coordinate Evaluation</b>", normal))
        e_page = expected_region.get("page", expected_region.get("page_min", "N/A"))
        tolerance = rule.get("tolerance", 0.10)

        elements.append(Paragraph(f"<b>Expected:</b> {expected_text}", small))
        elements.append(Paragraph(f"<b>Found:</b> {found_text}", small))
        if found_region and expected_region:
            fx0 = found_region.get("x_min", 0)
            fy0 = found_region.get("y_min", 0)
            ex0 = expected_region.get("x_min", 0)
            ey0 = expected_region.get("y_min", 0)
            if abs(fx0 - ex0) <= tolerance and abs(fy0 - ey0) <= tolerance:
                pos_status = "PASS"
                pos_reason = "Position matches expected template region."
            else:
                pos_status = "WARNING"
                pos_reason = f"Position deviates from expected region (tolerance: {tolerance:.2f})."
            elements.append(Paragraph(f"<b>Position Check:</b> <font color='{status_color_item.hexval()}'>{pos_status}</font>", small))
            elements.append(Paragraph(f"<b>Reason:</b> {pos_reason}", small))

        # C. Formatting Evaluation
        elements.append(Paragraph("<b>C. Formatting Evaluation</b>", normal))
        format_rules = rule.get("format_rules", {})
        if format_rules:
            if format_rules.get("uppercase_preferred"):
                elements.append(Paragraph("<b>Expected:</b> Heading should be in uppercase style.", small))
            if format_rules.get("bold_preferred"):
                elements.append(Paragraph("<b>Expected:</b> Heading should use bold font.", small))
            if format_rules.get("image_required"):
                elements.append(Paragraph("<b>Expected:</b> Image should be present.", small))
        else:
            elements.append(Paragraph("<b>Expected:</b> Standard paragraph formatting.", small))

        if formatting_ok is True:
            elements.append(Paragraph(f"<b>Formatting Check:</b> <font color='{colors.HexColor('#00aa00').hexval()}'>PASS</font>", small))
            elements.append(Paragraph("<b>Reason:</b> Formatting matches template expectations.", small))
        elif formatting_ok is False:
            elements.append(Paragraph(f"<b>Formatting Check:</b> <font color='{colors.HexColor('#cc8800').hexval()}'>WARNING</font>", small))
            elements.append(Paragraph("<b>Reason:</b> Formatting differs from template expectations.", small))
        else:
            elements.append(Paragraph("<b>Formatting Check:</b> N/A", small))

        elements.append(Paragraph(f"<b>Overall Item Status:</b> <font color='{status_color_item.hexval()}'>{status}</font>", normal))
        elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.lightgrey))
        elements.append(Spacer(1, 4))

    # ───── 8. EDITOR CONCLUSION ─────
    elements.append(PageBreak())
    elements.append(Paragraph("Editor Conclusion", heading2))

    if "PASS" in final_status:
        conclusion = (
            "The manuscript appears to follow the main Jurnal Teknologi template "
            "structure, layout, and formatting. No critical issues were detected. "
            "The manuscript is recommended to proceed."
        )
    elif "NEEDS" in final_status or "WARNING" in final_status:
        conclusion = (
            "The manuscript generally follows the template but has items that require "
            "manual review. These items had inconsistent evidence between DOC detection, "
            "PDF layout analysis, or AI verification. Please review the highlighted items "
            "before making a final decision."
        )
    else:
        conclusion = (
            "The manuscript does not fully comply with the template because one or "
            "more required items are missing or critically mismatched. The author "
            "should revise the manuscript using the official Jurnal Teknologi template."
        )

    elements.append(Paragraph(conclusion, normal))
    elements.append(Spacer(1, 8))

    elements.append(Paragraph(f"<b>Structure Score:</b> {struct_score}% (Weight: 40%)", normal))
    elements.append(Paragraph(f"<b>Layout Score:</b> {layout_score}% (Weight: 30%)", normal))
    elements.append(Paragraph(f"<b>Formatting Score:</b> {fmt_score}% (Weight: 30%)", normal))
    elements.append(Paragraph(f"<b>Final Score:</b> {final_score}/100", normal))
    elements.append(Spacer(1, 4))

    if critical_failed_names:
        elements.append(Paragraph("<b>Critical Items Failed:</b>", normal))
        for name in critical_failed_names:
            elements.append(Paragraph(f"&bull; {name} &mdash; <font color='#cc0000'>FAILED</font>", small))
    if warning_names:
        elements.append(Paragraph("<b>Items Needing Review:</b>", normal))
        for name in warning_names[:5]:
            elements.append(Paragraph(f"&bull; {name} &mdash; <font color='#cc8800'>WARNING</font>", small))

    elements.append(Spacer(1, 10))
    elements.append(Paragraph(
        "<i>This report was automatically generated by Turnitin Assistant v5.7.1 "
        "Template Screening Engine (DOCX-First + Multi-Source AI Evaluator).</i>", small
    ))

    # ── Step 2: Build the temporary report PDF ──
    try:
        doc.build(elements)
    except Exception as e:
        log(f"[SCREENING] Failed to generate report PDF: {e}")
        return None

    # ── Step 3: Save just the report PDF ──
    try:
        import shutil as shutil_mod
        shutil_mod.copy2(str(temp_report_file), str(report_file))
        log(f"[SCREENING] Report PDF saved: {report_file.name}")
    except Exception as e:
        log(f"[SCREENING] Failed to save report PDF: {e}")

    # ── Step 4: Merge report PDF + original article PDF into single file ──
    try:
        if fitz is None:
            log("[SCREENING] ⚠️ PyMuPDF not available. Outputting report-only PDF.")
            if not combined_file.exists():
                temp_report_file.rename(combined_file)
        else:
            # Open the temp report PDF
            report_doc = fitz.open(str(temp_report_file))

            # Open the original article PDF
            article_pdf_path = report_data.get("pdf_file", "")
            if not article_pdf_path or not Path(article_pdf_path).exists():
                log(f"[SCREENING] ⚠️ Article PDF not found: {article_pdf_path}. Outputting report-only PDF.")
                if not combined_file.exists():
                    temp_report_file.rename(combined_file)
            else:
                article_doc = fitz.open(article_pdf_path)

                # Append all article pages to the report document
                report_doc.insert_pdf(article_doc)

                # Save the combined PDF
                report_doc.save(str(combined_file))
                article_doc.close()

            report_doc.close()

        # Remove temp file
        if temp_report_file.exists():
            temp_report_file.unlink()

        log(f"[SCREENING] Combined screening report + article PDF saved: {combined_file.name}")
        return str(combined_file)

    except Exception as e:
        log(f"[SCREENING] Failed to merge article PDF: {e}")
        # Fall back: rename temp file
        if temp_report_file.exists() and not combined_file.exists():
            temp_report_file.rename(combined_file)
        return str(combined_file) if combined_file.exists() else str(report_file)


# ═══════════════════════════════════════════════════════════════════════════════
# v5.7.1: Main Workflow Entry Point
# ═══════════════════════════════════════════════════════════════════════════════

def run_template_screening_workflow(log_callback=None):
    """
    v5.7.1: Main callable entry point for Template Screening workflow.
    
    v5.7 changes:
    - Uses original source file for STRUCTURE detection (DOCX/DOC)
    - Uses converted PDF only for LAYOUT and FORMATTING detection
    - AI as third checker with rate limiting, caching, model switching
    - New manuscript-level status: PASS / NEEDS MANUAL REVIEW / REJECT
    - Organized file storage in subfolders
    - Updated Excel columns
    
    Steps:
    1. Read existing Excel data from 'download' sheet
    2. Find rows with downloaded files not yet screened
    3. Copy metadata (submission_no, author, title) to 'screening' sheet
    4. Convert DOCX to PDF if needed
    5. Run v5.7.1 DOCX-first template screening on source file + PDF
    6. Apply voting logic for final status
    7. Update 'screening' sheet with results
    8. Save screening report to organized storage
    """
    import openpyxl
    from config import TEMPLATE_SCREENING_CONFIG, OJS_DOWNLOADS_DIR

    def log(msg):
        if log_callback:
            log_callback(msg)
        else:
            print(msg)

    # Reset AI per-run counters
    _reset_ai_run_limits()

    log("[SCREENING v5.7.1] Starting DOCX-first template screening with Multi-Source AI Evaluator...")

    # Load rules
    rules_path = TEMPLATE_SCREENING_CONFIG["rules_file"]
    if not Path(rules_path).exists():
        log(f"[SCREENING] Rules file not found: {rules_path}")
        return {"status": "error", "message": f"Rules file not found: {rules_path}"}

    rules_config = load_template_rules(rules_path)
    rules = rules_config.get("rules", [])
    log(f"[SCREENING] Loaded {len(rules)} rules from {rules_config.get('template_name', 'unknown')}")

    # Read Excel
    excel_path = TEMPLATE_SCREENING_CONFIG["master_excel"]
    if not Path(excel_path).exists():
        log(f"[SCREENING] Excel not found: {excel_path}")
        return {"status": "error", "message": "Excel file not found"}

    log(f"[SCREENING] Reading Excel: {excel_path}")

    wb = openpyxl.load_workbook(excel_path)

    download_sheet_name = TEMPLATE_SCREENING_CONFIG.get("download_sheet", "download")
    screening_sheet_name = TEMPLATE_SCREENING_CONFIG.get("screening_sheet", "screening")

    # Ensure screening sheet exists
    if screening_sheet_name not in wb.sheetnames:
        wb.create_sheet(screening_sheet_name)
    ws_screening = wb[screening_sheet_name]

    # Ensure download sheet exists
    if download_sheet_name not in wb.sheetnames:
        log(f"[SCREENING] Download sheet '{download_sheet_name}' not found.")
        wb.close()
        return {"status": "error", "message": "Download sheet not found"}

    ws_download = wb[download_sheet_name]
    download_headers = [cell.value for cell in ws_download[1]]

    # Detect key columns in download sheet
    sub_col_idx = detect_column_index(download_headers, ["submission no", "submission_no", "submissionno"])
    author_col_idx = detect_column_index(download_headers, ["author name", "author_name", "author"])
    title_col_idx = detect_column_index(download_headers, ["title", "article title", "article_title"])
    ts_col_idx = detect_column_index(download_headers, ["timestamp", "date", "download timestamp"])

    # Detect path column
    path_col = detect_path_column(download_headers)
    if path_col is None:
        log("[SCREENING] No download path column detected in Excel.")
        possible_indices = [4, 5, 6, 7, 8]
        path_col_idx = None
        for idx in possible_indices:
            if idx < len(download_headers) and download_headers[idx]:
                h_val = str(download_headers[idx]).strip().lower()
                for name in ["download_path", "download location", "file_path", "path"]:
                    if h_val == name.lower():
                        path_col_idx = idx
                        break
                if path_col_idx is not None:
                    break
        if path_col_idx is None:
            log("[SCREENING] Cannot determine download path column.")
            wb.close()
            return {"status": "error", "message": "Cannot determine download path column"}
        path_col = download_headers[path_col_idx]
    else:
        path_col_idx = download_headers.index(path_col)

    log(f"[SCREENING] Path column detected: '{path_col}' (index {path_col_idx})")

    # === SET UP SCREENING SHEET HEADERS ===
    # v5.7: Added new columns
    screening_base_columns = [
        "timestamp",
        "submission_no",
        "author_name",
        "title",
        "download_location",
    ]

    screening_result_columns = [
        "template_screening_status",
        "template_screening_score",
        "template_screening_decision",
        "template_structure_status",
        "template_layout_status",
        "template_formatting_status",
        "template_screening_notes",
        "template_screening_date",
        "template_pdf_path",
        "template_screening_report_pdf",
        # v5.7: New columns
        "final_review_status",
        "doc_detect_summary",
        "pdf_detect_summary",
        "ai_detect_summary",
        "final_source_summary",
        "needs_manual_review_reason",
        "json_report_path",
        "pdf_report_path",
        "combined_report_path",
        "status_folder_path",
    ]

    all_screening_columns = screening_base_columns + screening_result_columns

    # Build header row on screening sheet
    screening_headers = [cell.value for cell in ws_screening[1]]
    if not any(screening_headers):
        # Empty sheet, write headers
        for col_idx, col_name in enumerate(all_screening_columns, 1):
            ws_screening.cell(row=1, column=col_idx, value=col_name)
        screening_headers = all_screening_columns
    else:
        # Ensure all result columns exist (don't break existing)
        existing = set(h for h in screening_headers if h)
        next_col = len(screening_headers) + 1
        for col_name in all_screening_columns:
            if col_name not in existing:
                ws_screening.cell(row=1, column=next_col, value=col_name)
                next_col += 1
                existing.add(col_name)
        screening_headers = [cell.value for cell in ws_screening[1]]

    # Build col index map for screening sheet
    scr_col_map = {}
    for idx, h in enumerate(screening_headers, 1):
        if h:
            scr_col_map[str(h)] = idx

    # Reports directory (v5.7: organized subfolders)
    reports_dir = Path(TEMPLATE_SCREENING_CONFIG["reports_dir"])
    reports_dir.mkdir(parents=True, exist_ok=True)
    
    # Ensure all v5.7 subdirectories exist
    for sub in ["json", "pdf", "combined", "converted_pdf", "crops", "ai_cache", "pass", "needs_review", "reject"]:
        (reports_dir / sub).mkdir(parents=True, exist_ok=True)
    
    convert_dir = reports_dir / "converted_pdf"
    convert_dir.mkdir(parents=True, exist_ok=True)

    # === SCAN DOWNLOAD SHEET FOR PENDING ROWS ===
    # Build a set of already-screened submission numbers from screening sheet
    screened_subs = set()
    if ws_screening.max_row >= 1:
        sub_col_index = scr_col_map.get("submission_no")
        if sub_col_index:
            for r in range(2, ws_screening.max_row + 1):
                sub_val = ws_screening.cell(row=r, column=sub_col_index).value
                if sub_val:
                    screened_subs.add(str(sub_val).strip())

    pending_rows = []
    for row_num in range(2, ws_download.max_row + 1):
        sub_val = None
        if sub_col_idx is not None:
            sub_val = ws_download.cell(row=row_num, column=sub_col_idx + 1).value
        sub_str = str(sub_val).strip() if sub_val else ""

        # Skip if already screened
        if sub_str in screened_subs:
            continue

        # Get file path
        file_path_cell = ws_download.cell(row=row_num, column=path_col_idx + 1)
        file_path = file_path_cell.value

        if not file_path:
            continue

        # Get metadata
        metadata = {
            "download_row": row_num,
            "submission_no": sub_str,
            "author_name": str(ws_download.cell(row=row_num, column=author_col_idx + 1).value or "").strip() if author_col_idx is not None else "",
            "title": str(ws_download.cell(row=row_num, column=title_col_idx + 1).value or "").strip() if title_col_idx is not None else "",
            "timestamp": str(ws_download.cell(row=row_num, column=ts_col_idx + 1).value or "").strip() if ts_col_idx is not None else "",
            "file_path": str(file_path),
        }
        pending_rows.append(metadata)

    log(f"[SCREENING] Found {len(pending_rows)} downloaded file(s) not yet screened.")

    if not pending_rows:
        log("[SCREENING] No pending files to screen.")
        wb.save(excel_path)
        wb.close()
        return {
            "status": "success",
            "message": "No pending files to screen.",
            "screened_count": 0,
        }

    # Track the next empty row in screening sheet
    next_screening_row = ws_screening.max_row + 1
    if next_screening_row < 2:
        next_screening_row = 2

    # === PROCESS EACH PENDING FILE ===
    screened_count = 0
    error_count = 0
    source_file_map = {}  # key: submission_no, value: resolved source file path

    for item in pending_rows:
        row_num = item["download_row"]
        sub_no = item["submission_no"]
        file_path_str = item["file_path"]
        file_path = Path(file_path_str)

        # Resolve file path (handle older v5.x paths)
        if not file_path.exists():
            # Try to find by submission_no folder in OJS_DOWNLOADS_DIR
            found = False
            if OJS_DOWNLOADS_DIR.exists():
                for sub_dir in sorted(OJS_DOWNLOADS_DIR.iterdir(), reverse=True):
                    if sub_dir.is_dir() and sub_no in sub_dir.name:
                        manuscripts = list(sub_dir.iterdir())
                        if manuscripts:
                            file_path = manuscripts[0]
                            found = True
                            break
                if not found:
                    # Search recursively by filename
                    for sub_dir in OJS_DOWNLOADS_DIR.iterdir():
                        if sub_dir.is_dir():
                            for f in sub_dir.iterdir():
                                if f.is_file() and f.suffix.lower() in (".docx", ".doc", ".pdf"):
                                    # Check if this file matches (by submission_no)
                                    if sub_no in f.parent.name or sub_no in f.stem:
                                        file_path = f
                                        found = True
                                        break
                        if found:
                            break

            if not found:
                log(f"[SCREENING] {sub_no}: file not found: {file_path_str}")
                # Write error to screening sheet
                screening_row = next_screening_row
                next_screening_row += 1
                ws_screening.cell(row=screening_row, column=scr_col_map.get("submission_no", 2), value=sub_no)
                ws_screening.cell(row=screening_row, column=scr_col_map.get("author_name", 3), value=item["author_name"])
                ws_screening.cell(row=screening_row, column=scr_col_map.get("title", 4), value=item["title"])
                ws_screening.cell(row=screening_row, column=scr_col_map.get("download_location", 5), value=file_path_str)
                ws_screening.cell(row=screening_row, column=scr_col_map.get("timestamp", 1), value=item["timestamp"])
                ws_screening.cell(row=screening_row, column=scr_col_map.get("template_screening_status", 6), value="SCREENING ERROR")
                ws_screening.cell(row=screening_row, column=scr_col_map.get("template_screening_notes", 12), value=f"File not found: {file_path_str}")
                ws_screening.cell(row=screening_row, column=scr_col_map.get("template_screening_date", 13), value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                error_count += 1
                continue

        source_file_map[sub_no] = file_path

        # Determine PDF path
        pdf_path = None
        ext = file_path.suffix.lower()

        if ext == ".pdf":
            pdf_path = str(file_path)
            log(f"[SCREENING] {sub_no}: source is PDF: {file_path.name}")
        elif ext in (".docx", ".doc"):
            log(f"[SCREENING] {sub_no}: converting DOCX to PDF...")
            pdf_path = convert_docx_to_pdf(file_path, output_dir=convert_dir)
            if pdf_path is None:
                log(f"[SCREENING] {sub_no}: DOCX to PDF conversion failed.")
                screening_row = next_screening_row
                next_screening_row += 1
                ws_screening.cell(row=screening_row, column=scr_col_map.get("submission_no", 2), value=sub_no)
                ws_screening.cell(row=screening_row, column=scr_col_map.get("author_name", 3), value=item["author_name"])
                ws_screening.cell(row=screening_row, column=scr_col_map.get("title", 4), value=item["title"])
                ws_screening.cell(row=screening_row, column=scr_col_map.get("download_location", 5), value=str(file_path))
                ws_screening.cell(row=screening_row, column=scr_col_map.get("timestamp", 1), value=item["timestamp"])
                ws_screening.cell(row=screening_row, column=scr_col_map.get("template_screening_status", 6), value="SCREENING ERROR")
                ws_screening.cell(row=screening_row, column=scr_col_map.get("template_screening_notes", 12), value="DOCX to PDF conversion failed.")
                ws_screening.cell(row=screening_row, column=scr_col_map.get("template_screening_date", 13), value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                error_count += 1
                continue
            log(f"[SCREENING] {sub_no}: converted to PDF: {Path(pdf_path).name}")
        else:
            log(f"[SCREENING] {sub_no}: unsupported file format: {ext}")
            screening_row = next_screening_row
            next_screening_row += 1
            ws_screening.cell(row=screening_row, column=scr_col_map.get("submission_no", 2), value=sub_no)
            ws_screening.cell(row=screening_row, column=scr_col_map.get("author_name", 3), value=item["author_name"])
            ws_screening.cell(row=screening_row, column=scr_col_map.get("title", 4), value=item["title"])
            ws_screening.cell(row=screening_row, column=scr_col_map.get("download_location", 5), value=str(file_path))
            ws_screening.cell(row=screening_row, column=scr_col_map.get("timestamp", 1), value=item["timestamp"])
            ws_screening.cell(row=screening_row, column=scr_col_map.get("template_screening_status", 6), value="SCREENING ERROR")
            ws_screening.cell(row=screening_row, column=scr_col_map.get("template_screening_notes", 12), value=f"Unsupported file format: {ext}")
            ws_screening.cell(row=screening_row, column=scr_col_map.get("template_screening_date", 13), value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            error_count += 1
            continue

        # Reset AI per-submission counters
        _reset_ai_rate_limits()

        # ── v5.7: Run DOCX-first screening ──
        log(f"[SCREENING v5.7.1] {sub_no}: running DOCX-first multi-source template screening...")
        try:
            # Use v5.7 DOCX-first screening function (with AI as 3rd checker)
            results = run_docx_first_screening(
                str(file_path),  # source file
                pdf_path,        # PDF for layout
                rules,
                log_callback=log_callback,
            )
        except Exception as e:
            log(f"[SCREENING] {sub_no}: screening error: {e}")
            screening_row = next_screening_row
            next_screening_row += 1
            ws_screening.cell(row=screening_row, column=scr_col_map.get("submission_no", 2), value=sub_no)
            ws_screening.cell(row=screening_row, column=scr_col_map.get("author_name", 3), value=item["author_name"])
            ws_screening.cell(row=screening_row, column=scr_col_map.get("title", 4), value=item["title"])
            ws_screening.cell(row=screening_row, column=scr_col_map.get("download_location", 5), value=str(file_path))
            ws_screening.cell(row=screening_row, column=scr_col_map.get("timestamp", 1), value=item["timestamp"])
            ws_screening.cell(row=screening_row, column=scr_col_map.get("template_screening_status", 6), value="SCREENING ERROR")
            ws_screening.cell(row=screening_row, column=scr_col_map.get("template_screening_notes", 12), value=f"Screening engine error: {str(e)}")
            ws_screening.cell(row=screening_row, column=scr_col_map.get("template_screening_date", 13), value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            error_count += 1
            continue

        if isinstance(results, dict) and "error" in results:
            log(f"[SCREENING] {sub_no}: screening failed: {results['error']}")
            screening_row = next_screening_row
            next_screening_row += 1
            ws_screening.cell(row=screening_row, column=scr_col_map.get("submission_no", 2), value=sub_no)
            ws_screening.cell(row=screening_row, column=scr_col_map.get("author_name", 3), value=item["author_name"])
            ws_screening.cell(row=screening_row, column=scr_col_map.get("title", 4), value=item["title"])
            ws_screening.cell(row=screening_row, column=scr_col_map.get("download_location", 5), value=str(file_path))
            ws_screening.cell(row=screening_row, column=scr_col_map.get("timestamp", 1), value=item["timestamp"])
            ws_screening.cell(row=screening_row, column=scr_col_map.get("template_screening_status", 6), value="SCREENING ERROR")
            ws_screening.cell(row=screening_row, column=scr_col_map.get("template_screening_notes", 12), value=results["error"])
            ws_screening.cell(row=screening_row, column=scr_col_map.get("template_screening_date", 13), value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            error_count += 1
            continue

        # Score results
        final_score, struct_score, layout_score, fmt_score = score_results(results)

        # ─── v5.7 OVERALL MANUSCRIPT DECISION ───
        # v5.7: PASS / NEEDS MANUAL REVIEW / REJECT
        # REJECT only if critical items are truly missing after DOC/PDF/AI
        # NEEDS MANUAL REVIEW if many inconsistent items
        
        critical_item_ids = {"title", "authors", "abstract", "abstrak", "references"}

        truly_critical_failed = []
        needs_review_count = 0
        passed_count = 0

        for r in results:
            rid = r.get("id")
            rstatus = r.get("status")
            r_evidence = r.get("evidence_sources", [])

            # Critical items
            if rid in critical_item_ids and rstatus == "FAILED":
                truly_critical_failed.append(rid)
            
            # Count needs_review items
            if rstatus == "WARNING" or rstatus == "NEEDS_REVIEW":
                needs_review_count += 1
            elif rstatus == "PASS":
                passed_count += 1

        # Check abstract/abstrak: need at least one
        abstract_status = next((r.get("status") for r in results if r.get("id") == "abstract"), "FAILED")
        abstrak_status = next((r.get("status") for r in results if r.get("id") == "abstrak"), "FAILED")
        if abstract_status == "FAILED" and abstrak_status == "FAILED":
            truly_critical_failed.append("abstract/abstrak")

        has_critical_fail = len(truly_critical_failed) > 0

        # v5.7: New decision logic
        if final_score >= 80 and not has_critical_fail and needs_review_count == 0:
            final_status = "PASS"
        elif final_score >= 80 and not has_critical_fail and needs_review_count > 0:
            final_status = "NEEDS MANUAL REVIEW"
        elif final_score >= 50 and not has_critical_fail:
            final_status = "NEEDS MANUAL REVIEW"
        else:
            final_status = "REJECT"

        # Build notes
        passed_items = [r["label"] for r in results if r.get("status") == "PASS"]
        failed_items = [r["label"] for r in results if r.get("status") == "FAILED"]
        review_items = [r["label"] for r in results if r.get("status") in ("WARNING", "NEEDS_REVIEW")]
        
        notes_parts = []
        if passed_items:
            notes_parts.append(f"Passed: {', '.join(passed_items[:5])}")
        if review_items:
            notes_parts.append(f"Review: {', '.join(review_items[:3])}")
        if failed_items:
            notes_parts.append(f"Failed: {', '.join(failed_items[:3])}")
        notes = " | ".join(notes_parts)

        log(f"[SCREENING] SUB-{sub_no}: {final_status} (Score: {final_score})")

        # Write to screening sheet
        screening_row = next_screening_row
        next_screening_row += 1

        # Write basic columns
        ws_screening.cell(row=screening_row, column=scr_col_map.get("timestamp", 1), value=item["timestamp"])
        ws_screening.cell(row=screening_row, column=scr_col_map.get("submission_no", 2), value=sub_no)
        ws_screening.cell(row=screening_row, column=scr_col_map.get("author_name", 3), value=item["author_name"])
        ws_screening.cell(row=screening_row, column=scr_col_map.get("title", 4), value=item["title"])
        ws_screening.cell(row=screening_row, column=scr_col_map.get("download_location", 5), value=str(file_path))
        
        # Determine ACCEPT/REVIEW/REJECT decision
        if final_status == "PASS":
            decision = "ACCEPT"
        elif final_status == "NEEDS MANUAL REVIEW":
            decision = "REVIEW"
        else:
            decision = "REJECT"

        ws_screening.cell(row=screening_row, column=scr_col_map.get("template_screening_status", 6), value=final_status)
        ws_screening.cell(row=screening_row, column=scr_col_map.get("template_screening_score", 7), value=final_score)
        if scr_col_map.get("template_screening_decision"):
            ws_screening.cell(row=screening_row, column=scr_col_map["template_screening_decision"], value=decision)
        ws_screening.cell(row=screening_row, column=scr_col_map.get("template_structure_status", 8), value="PASS" if struct_score >= 70 else "FAILED")
        ws_screening.cell(row=screening_row, column=scr_col_map.get("template_layout_status", 9), value="PASS" if layout_score >= 60 else "WARNING" if layout_score >= 40 else "FAILED")
        ws_screening.cell(row=screening_row, column=scr_col_map.get("template_formatting_status", 10), value="PASS" if fmt_score >= 60 else "WARNING" if fmt_score >= 40 else "FAILED")
        ws_screening.cell(row=screening_row, column=scr_col_map.get("template_screening_notes", 12), value=notes)
        ws_screening.cell(row=screening_row, column=scr_col_map.get("template_screening_date", 13), value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        if scr_col_map.get("template_pdf_path"):
            ws_screening.cell(row=screening_row, column=scr_col_map["template_pdf_path"], value=pdf_path)

        # ─── v5.7: New Excel columns ───
        if scr_col_map.get("final_review_status"):
            ws_screening.cell(row=screening_row, column=scr_col_map["final_review_status"], value=final_status)
        
        # Build summaries
        doc_yes = sum(1 for r in results if r.get("detected_from_docx"))
        pdf_yes = sum(1 for r in results if r.get("detected_from_pdf"))
        ai_yes = sum(1 for r in results if r.get("detected_from_ai"))
        
        if scr_col_map.get("doc_detect_summary"):
            ws_screening.cell(row=screening_row, column=scr_col_map["doc_detect_summary"],
                             value=f"DOC detected {doc_yes}/{len(results)} items")
        if scr_col_map.get("pdf_detect_summary"):
            ws_screening.cell(row=screening_row, column=scr_col_map["pdf_detect_summary"],
                             value=f"PDF detected {pdf_yes}/{len(results)} items")
        if scr_col_map.get("ai_detect_summary"):
            ai_statuses = [r.get("ai_status", "SKIPPED") for r in results]
            ai_yes_count = sum(1 for s in ai_statuses if s == "YES")
            ai_skipped = sum(1 for s in ai_statuses if s in ("SKIPPED", "DISABLED"))
            ai_error = sum(1 for s in ai_statuses if s == "ERROR")
            ws_screening.cell(row=screening_row, column=scr_col_map["ai_detect_summary"],
                             value=f"AI: {ai_yes_count} YES, {ai_skipped} skipped, {ai_error} errors")
        
        if scr_col_map.get("final_source_summary"):
            # Count unique final sources
            sources = [r.get("final_source_used", "None") for r in results]
            ws_screening.cell(row=screening_row, column=scr_col_map["final_source_summary"],
                             value=f"Sources: {', '.join(sorted(set(sources)))[:200]}")
        
        if scr_col_map.get("needs_manual_review_reason"):
            review_reasons = []
            for r in results:
                if r.get("status") in ("WARNING", "NEEDS_REVIEW"):
                    review_reasons.append(f"{r.get('id')}: {r.get('notes', '')[:80]}")
            ws_screening.cell(row=screening_row, column=scr_col_map["needs_manual_review_reason"],
                             value=" | ".join(review_reasons[:5]) if review_reasons else "")

        # Save JSON report (v5.7: organized path)
        timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        json_dir = reports_dir / "json"
        json_report_filename = f"template_screening_{sub_no}_{timestamp_str}.json"
        json_report_path = json_dir / json_report_filename

        # Build enhanced report data with v5.7 fields
        report_data = {
            "submission_no": sub_no,
            "source_file": str(file_path),
            "pdf_file": pdf_path,
            "final_status": final_status,
            "final_score": final_score,
            "structure_score": struct_score,
            "layout_score": layout_score,
            "formatting_score": fmt_score,
            "items": results,
            "screening_version": "v5.7.1-multi-source-ai-evaluator",
            "use_ai_verifier": USE_AI_VERIFIER,
            "ai_text_model": AI_TEXT_MODEL,
            "ai_vision_model": AI_VISION_MODEL,
        }
        
        with open(json_report_path, "w", encoding="utf-8") as f:
            json.dump(report_data, f, indent=2, ensure_ascii=False)

        log(f"[SCREENING] JSON report saved: {json_report_filename}")
        
        # Store JSON path in Excel
        if scr_col_map.get("json_report_path"):
            ws_screening.cell(row=screening_row, column=scr_col_map["json_report_path"], value=str(json_report_path))

        # Generate combined PDF (screening report + original article in one file)
        pdf_report_path = None
        try:
            pdf_report_path = generate_screening_pdf_report(
                report_data,
                item,
                str(reports_dir),
                log_callback=log_callback,
            )
            if pdf_report_path:
                if scr_col_map.get("template_screening_report_pdf"):
                    ws_screening.cell(row=screening_row, column=scr_col_map["template_screening_report_pdf"], value=pdf_report_path)
                log(f"[SCREENING] Combined PDF report saved: {Path(pdf_report_path).name}")
        except Exception as e:
            log(f"[SCREENING] Combined PDF report generation skipped: {e}")

        # v5.7: Copy report to status folders
        if pdf_report_path:
            pdf_report_path_obj = Path(pdf_report_path)
            if final_status == "PASS":
                target_dir = reports_dir / "pass"
            elif final_status == "NEEDS MANUAL REVIEW":
                target_dir = reports_dir / "needs_review"
            else:
                target_dir = reports_dir / "reject"
            
            target_dir.mkdir(parents=True, exist_ok=True)
            try:
                import shutil as shutil_mod
                status_copy_path = target_dir / pdf_report_path_obj.name
                shutil_mod.copy2(pdf_report_path_obj, status_copy_path)
                
                if scr_col_map.get("pdf_report_path"):
                    ws_screening.cell(row=screening_row, column=scr_col_map["pdf_report_path"], value=str(status_copy_path))
                if scr_col_map.get("combined_report_path"):
                    ws_screening.cell(row=screening_row, column=scr_col_map["combined_report_path"], value=pdf_report_path)
                if scr_col_map.get("status_folder_path"):
                    ws_screening.cell(row=screening_row, column=scr_col_map["status_folder_path"], value=str(target_dir))
                
                log(f"[SCREENING] Status copy saved to: {target_dir.name}/")
            except Exception as e:
                log(f"[SCREENING] Status copy failed: {e}")

        screened_count += 1

    # Save Excel
    wb.save(excel_path)
    wb.close()

    log(f"[SCREENING] Excel updated (screening sheet).")
    log(f"[SCREENING] Screened: {screened_count}, Errors: {error_count}.")
    log("[SCREENING] Done.")

    return {
        "status": "success",
        "message": f"Template screening completed. Screened: {screened_count}, Errors: {error_count}.",
        "screened_count": screened_count,
        "error_count": error_count,
    }