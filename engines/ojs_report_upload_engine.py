"""
OJS Report Upload Engine v5.3
Uploads Turnitin PDF reports back to OJS submission workflow.

Architecture:
  - Callable function: run_ojs_report_upload_workflow(log_callback=None)
  - No terminal input
  - UI log callback
  - Excel-driven data (reads sheet 'turnitin', writes to sheet 'upload')
  - Workflow lock (prevent concurrent runs)
  - Persistent ojs-profile
  - Returns dict with status/message
"""

from playwright.sync_api import sync_playwright
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import time
import random
import threading

# ==============================================================================
# CONFIG
# ==============================================================================

PROJECT_ROOT = Path(__file__).resolve().parent.parent

# Import stable profile paths from config
import sys
sys.path.insert(0, str(PROJECT_ROOT))
from config import STABLE_OJS_PROFILE_DIR

PROFILE_DIR  = str(STABLE_OJS_PROFILE_DIR)
MASTER_XLSX  = PROJECT_ROOT / "data" / "master update.xlsx"

OJS_HOME_URL        = "https://journals.utm.my/jurnalteknologi"
OJS_SUBMISSIONS_URL = "https://journals.utm.my/jurnalteknologi/submissions"

STUCK_SECONDS = 60
MAX_RECOVERY  = 2
HUMAN_SLOW_MO = 2000  # v5.5.1: Slower human-like typing speed (was 900)

# Sheet names
TURNITIN_SHEET = "turnitin"
UPLOAD_SHEET   = "upload"

# Column indices (0-based) in sheet 'turnitin'
COL_T_TIMESTAMP   = 0  # A
COL_T_SUB_NO      = 1  # B
COL_T_NAME        = 2  # C
COL_T_SCORE       = 3  # D
COL_T_REPORT_PATH = 4  # E
COL_T_STATUS      = 5  # F

# Column indices (0-based) in sheet 'upload'
COL_U_TIMESTAMP   = 0  # A
COL_U_SUB_NO      = 1  # B
COL_U_REPORT_PATH = 2  # C
COL_U_NAME        = 3  # D
COL_U_STATUS      = 4  # E
COL_U_NOTES       = 5  # F

STATUS_BERJAYA   = "BERJAYA UPLOAD OJS"
STATUS_TAK       = "TAK BERJAYA UPLOAD OJS"

# Workflow guard
_workflow_lock    = threading.Lock()
_workflow_running = False

# UI log callback
_ui_log_callback = None


def set_ui_log_callback(callback):
    global _ui_log_callback
    _ui_log_callback = callback


def ui_log(message):
    if _ui_log_callback:
        _ui_log_callback(message)
    else:
        print(message)


# ==============================================================================
# EXCEL HELPERS
# ==============================================================================

def _ensure_upload_sheet(wb):
    """Create sheet 'upload' with headers if it does not exist."""
    if UPLOAD_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(UPLOAD_SHEET)
    else:
        ws = wb[UPLOAD_SHEET]

    # Write headers if row 1 is empty
    if ws.cell(1, 1).value is None:
        headers = ["timestamp", "submission_no", "turnitin_report_path",
                   "turnitin_name", "status", "notes"]
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font      = Font(name="Arial", size=20, bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[1].height = 45

    return ws


def _get_already_uploaded(wb):
    """
    Return set of submission_no values that already have STATUS_BERJAYA
    in sheet 'upload'.
    """
    uploaded = set()
    if UPLOAD_SHEET not in wb.sheetnames:
        return uploaded

    ws = wb[UPLOAD_SHEET]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 5:
            continue
        sub_no = str(row[COL_U_SUB_NO] or "").strip()
        status = str(row[COL_U_STATUS] or "").strip()
        if sub_no and status == STATUS_BERJAYA:
            uploaded.add(sub_no)
    return uploaded


def find_next_pending_row():
    """
    Read sheet 'turnitin'. Find first row where:
      - Column F (Status) == 'BERJAYA'
      - Column E (Report Path) is non-empty and file exists
      - Submission No NOT already in sheet 'upload' with STATUS_BERJAYA

    Returns dict or None.
    """
    if not MASTER_XLSX.exists():
        ui_log(f"[OJS UPLOAD] ❌ Excel not found: {MASTER_XLSX}")
        return None

    try:
        wb = load_workbook(str(MASTER_XLSX), read_only=False, data_only=True)
    except Exception as e:
        ui_log(f"[OJS UPLOAD] ❌ Cannot open Excel: {e}")
        return None

    try:
        if TURNITIN_SHEET not in wb.sheetnames:
            ui_log(f"[OJS UPLOAD] ❌ Sheet '{TURNITIN_SHEET}' not found.")
            wb.close()
            return None

        # Collect already-uploaded
        already_uploaded = _get_already_uploaded(wb)
        ui_log(f"[OJS UPLOAD] Already BERJAYA UPLOAD OJS: {len(already_uploaded)} submission(s)")

        ws_t = wb[TURNITIN_SHEET]

        for row in ws_t.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 6:
                continue

            sub_no      = str(row[COL_T_SUB_NO]      or "").strip()
            turnitin_name = str(row[COL_T_NAME]       or "").strip()
            report_path = str(row[COL_T_REPORT_PATH]  or "").strip()
            status_f    = str(row[COL_T_STATUS]       or "").strip()

            if not sub_no:
                continue

            if status_f != "BERJAYA":
                continue

            if not report_path:
                ui_log(f"[OJS UPLOAD] Skip {sub_no}: report path kosong")
                continue

            if not Path(report_path).exists():
                ui_log(f"[OJS UPLOAD] Skip {sub_no}: file tidak wujud: {report_path}")
                continue

            if sub_no in already_uploaded:
                ui_log(f"[OJS UPLOAD] Skip {sub_no}: sudah BERJAYA UPLOAD OJS")
                continue

            wb.close()
            ui_log(f"[OJS UPLOAD] ✅ Akan upload: submission_no={sub_no}")
            ui_log(f"[OJS UPLOAD]   Report: {report_path}")
            ui_log(f"[OJS UPLOAD]   Name  : {turnitin_name}")

            return {
                "submission_no":    sub_no,
                "turnitin_name":    turnitin_name,
                "turnitin_report_path": report_path,
            }

        wb.close()
        ui_log("[OJS UPLOAD] Tiada row pending untuk diupload.")
        return None

    except Exception as e:
        ui_log(f"[OJS UPLOAD] ❌ Error reading Excel: {e}")
        try:
            wb.close()
        except Exception:
            pass
        return None


def log_upload_result(submission_no, turnitin_report_path, turnitin_name,
                      status, notes):
    """Write one row to sheet 'upload' in the master Excel file."""
    if not MASTER_XLSX.exists():
        ui_log(f"[OJS UPLOAD EXCEL] ❌ Excel not found: {MASTER_XLSX}")
        return

    try:
        wb = load_workbook(str(MASTER_XLSX))
        ws = _ensure_upload_sheet(wb)

        next_row = ws.max_row + 1
        now_str  = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        values = [
            now_str,
            submission_no,
            turnitin_report_path,
            turnitin_name,
            status,
            notes,
        ]

        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=next_row, column=col, value=val)
            cell.font      = Font(name="Arial", size=20)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        ws.row_dimensions[next_row].height = 50

        wb.save(str(MASTER_XLSX))
        ui_log(f"[OJS UPLOAD EXCEL] ✅ Logged row {next_row}: {status}")

    except Exception as e:
        ui_log(f"[OJS UPLOAD EXCEL] ❌ Failed to log: {e}")


# ==============================================================================
# BROWSER / PAGE HELPERS
# ==============================================================================

def random_human_wait(page, min_ms=2500, max_ms=6500):
    page.wait_for_timeout(random.randint(min_ms, max_ms))


def wait_until(page, label, check_js, timeout_seconds=STUCK_SECONDS,
               allow_refresh=True):
    """Poll until JS expression returns truthy, with optional page-refresh recovery."""
    recovery_count = 0

    while True:
        start = time.time()
        while time.time() - start < timeout_seconds:
            try:
                if page.evaluate(check_js):
                    ui_log(f"[OK] {label}")
                    return True
            except Exception:
                pass
            random_human_wait(page, 2000, 4000)

        if not allow_refresh or recovery_count >= MAX_RECOVERY:
            raise Exception(f"Timeout waiting for: {label}")

        recovery_count += 1
        ui_log(f"[RECOVERY] {label}: stuck, refreshing page...")
        page.reload(wait_until="domcontentloaded", timeout=120000)
        random_human_wait(page, 7000, 12000)


def modal(page):
    return page.locator(".pkp_modal.is_visible, .pkpModalWrapper.is_visible").last


def click_by_text_or_selector(page, label, candidates, timeout_seconds=60):
    ui_log(f"[CLICK] Waiting/clicking: {label}")
    end = time.time() + timeout_seconds

    while time.time() < end:
        for c in candidates:
            try:
                loc = page.locator(c).first
                if loc.count() > 0 and loc.is_visible():
                    loc.scroll_into_view_if_needed(timeout=10000)
                    random_human_wait(page, 2000, 4500)
                    loc.click(timeout=20000)
                    ui_log(f"[CLICKED] {label} via: {c}")
                    return True
            except Exception:
                pass
        random_human_wait(page, 2000, 4000)

    raise Exception(f"Cannot find/click: {label}")


# ==============================================================================
# NAVIGATION HELPERS
# ==============================================================================

def open_ojs_home_if_needed(page):
    ui_log("[OJS UPLOAD] Checking OJS page...")
    if "journals.utm.my" not in page.url:
        ui_log("[OJS UPLOAD] Opening OJS home...")
        page.goto(OJS_HOME_URL, wait_until="domcontentloaded", timeout=120000)
        random_human_wait(page, 7000, 12000)
    ui_log(f"[OJS UPLOAD] Current URL: {page.url}")


def go_to_dashboard_if_needed(page):
    ui_log("[OJS UPLOAD] Checking dashboard...")
    if "/submissions" in page.url or "/workflow" in page.url:
        ui_log("[OJS UPLOAD] Already inside dashboard/submissions/workflow")
        return

    try:
        click_by_text_or_selector(
            page, "User dropdown",
            ['button:has-text("adminjurnalutm")',
             'a:has-text("adminjurnalutm")',
             'text=adminjurnalutm'],
            timeout_seconds=20
        )
    except Exception:
        ui_log("[OJS UPLOAD] User dropdown not clicked / maybe already open")

    click_by_text_or_selector(
        page, "Dashboard",
        ['a:has-text("Dashboard")',
         'button:has-text("Dashboard")',
         'text=Dashboard'],
        timeout_seconds=60
    )

    page.wait_for_load_state("domcontentloaded", timeout=120000)
    random_human_wait(page, 7000, 12000)


def go_to_submissions(page):
    ui_log("[OJS UPLOAD] Going to Submissions...")
    if "/submissions" in page.url:
        ui_log("[OJS UPLOAD] Already on submissions page")
        return

    click_by_text_or_selector(
        page, "Submissions menu",
        ['a[href*="/submissions"]',
         'a:has-text("Submissions")',
         'button:has-text("Submissions")',
         'text=Submissions'],
        timeout_seconds=60
    )

    page.wait_for_load_state("domcontentloaded", timeout=120000)
    random_human_wait(page, 5000, 9000)


def go_to_active_tab(page):
    ui_log("[OJS UPLOAD] Going to Active tab...")
    click_by_text_or_selector(
        page, "Active tab",
        ['#active-button',
         'button:has-text("Active")',
         'text=Active'],
        timeout_seconds=60
    )
    random_human_wait(page, 6000, 10000)


def search_submission_id(page, submission_id):
    ui_log(f"[OJS UPLOAD] Searching submission: {submission_id}")
    ui_log(f"[OJS UPLOAD] Searching submission ID: {submission_id}")

    wait_until(
        page,
        "Visible OJS search input",
        """
        () => {
            const inputs = [...document.querySelectorAll('input.pkpSearch__input')];
            return inputs.some(i => i.offsetParent !== null);
        }
        """,
        timeout_seconds=60,
        allow_refresh=True
    )

    ok = page.evaluate(
        """
        async (submissionId) => {
            const inputs = [...document.querySelectorAll('input.pkpSearch__input')]
                .filter(i => i.offsetParent !== null);
            if (!inputs.length) return false;
            const input = inputs[inputs.length - 1];
            input.scrollIntoView({ block: 'center' });
            input.focus();
            input.value = '';
            input.dispatchEvent(new Event('input', { bubbles: true }));
            input.dispatchEvent(new Event('change', { bubbles: true }));
            await new Promise(r => setTimeout(r, 500));
            input.value = submissionId;
            input.dispatchEvent(new Event('input', { bubbles: true }));
            input.dispatchEvent(new Event('change', { bubbles: true }));
            await new Promise(r => setTimeout(r, 500));
            input.dispatchEvent(new KeyboardEvent('keydown', {
                key: 'Enter', code: 'Enter', keyCode: 13, which: 13, bubbles: true
            }));
            input.dispatchEvent(new KeyboardEvent('keyup', {
                key: 'Enter', code: 'Enter', keyCode: 13, which: 13, bubbles: true
            }));
            return true;
        }
        """,
        submission_id
    )

    if not ok:
        raise Exception("Search input not found")

    random_human_wait(page, 8000, 13000)
    ui_log(f"[OJS UPLOAD] Search entered: {submission_id}")


def open_submission_row(page, submission_id):
    ui_log(f"[OJS UPLOAD] Opening submission row: {submission_id}")

    wait_until(
        page,
        f"View link for submission {submission_id}",
        f"""
        () => {{
            return [...document.querySelectorAll('a.pkpButton, a')]
                .some(a =>
                    a.offsetParent !== null &&
                    a.href &&
                    (
                        a.href.includes('/workflow/access/{submission_id}') ||
                        a.href.includes('id={submission_id}')
                    )
                );
        }}
        """,
        timeout_seconds=90,
        allow_refresh=True
    )

    random_human_wait(page, 3000, 6000)

    ok = page.evaluate(
        """
        (submissionId) => {
            const links = [...document.querySelectorAll('a.pkpButton, a')]
                .filter(a =>
                    a.offsetParent !== null &&
                    a.href &&
                    (
                        a.href.includes('/workflow/access/' + submissionId) ||
                        a.href.includes('id=' + submissionId)
                    )
                );
            if (!links.length) return false;
            const link = links[0];
            link.scrollIntoView({ block: 'center' });
            link.click();
            return true;
        }
        """,
        submission_id
    )

    if not ok:
        raise Exception(f"Cannot click View for submission {submission_id}")

    page.wait_for_load_state("domcontentloaded", timeout=120000)
    random_human_wait(page, 8000, 13000)
    ui_log(f"[OJS UPLOAD] ✅ View clicked for submission: {submission_id}")


def navigate_to_submission(page, submission_id):
    """Full navigation: OJS → Dashboard → Submissions → Active → Search → Open."""
    open_ojs_home_if_needed(page)
    go_to_dashboard_if_needed(page)
    go_to_submissions(page)
    go_to_active_tab(page)
    search_submission_id(page, submission_id)
    open_submission_row(page, submission_id)
    ui_log("[OJS UPLOAD] Now ready for Upload File stage.")


# ==============================================================================
# UPLOAD FLOW
# ==============================================================================

def close_existing_modal_if_any(page):
    try:
        has_modal = page.locator(
            ".pkp_modal.is_visible, .pkpModalWrapper.is_visible"
        ).count() > 0
        if has_modal:
            ui_log("[OJS UPLOAD] Existing modal detected. Closing first...")
            page.keyboard.press("Escape")
            random_human_wait(page, 3000, 5000)
    except Exception:
        pass


def click_upload_file(page):
    ui_log("[OJS UPLOAD] Waiting Upload File button...")
    close_existing_modal_if_any(page)

    wait_until(
        page,
        "Upload File button visible",
        """
        () => {
            const exact = document.querySelector(
                'a[id^="component-grid-files-submission-editorsubmissiondetailsfilesgrid-addFile-button"]'
            );
            if (exact && exact.offsetParent !== null) return true;
            const links = [...document.querySelectorAll('a, button')];
            return links.some(el =>
                el.offsetParent !== null &&
                el.innerText &&
                el.innerText.toLowerCase().includes('upload file')
            );
        }
        """,
        timeout_seconds=STUCK_SECONDS,
        allow_refresh=True
    )

    selectors = [
        'a[id^="component-grid-files-submission-editorsubmissiondetailsfilesgrid-addFile-button"]',
        'a.pkp_linkaction_addFile',
        'a:has-text("Upload File")',
        'button:has-text("Upload File")',
    ]

    for sel in selectors:
        try:
            btn = page.locator(sel).first
            if btn.count() > 0 and btn.is_visible():
                random_human_wait(page, 2000, 5000)
                btn.scroll_into_view_if_needed(timeout=10000)
                random_human_wait(page, 2000, 4000)
                btn.click(timeout=20000)
                ui_log("[OJS UPLOAD] ✅ CLICKED Upload File")
                return True
        except Exception as e:
            ui_log(f"[OJS UPLOAD] Upload selector failed: {sel} — {e}")

    raise Exception("Upload File button not found")


def wait_modal_open(page):
    wait_until(
        page,
        "Upload modal opened",
        """
        () => {
            const modal = document.querySelector('.pkp_modal.is_visible, .pkpModalWrapper.is_visible');
            return modal && modal.innerText.includes('Upload Submission File');
        }
        """,
        timeout_seconds=STUCK_SECONDS,
        allow_refresh=False
    )


def select_turnitin_reports(page):
    wait_modal_open(page)
    ui_log("[OJS UPLOAD] Waiting Turnitin Reports dropdown...")

    wait_until(
        page,
        "genreId enabled",
        """
        () => {
            const el = document.querySelector(
                '.pkp_modal.is_visible #genreId, .pkpModalWrapper.is_visible #genreId'
            );
            return el && !el.disabled;
        }
        """,
        timeout_seconds=STUCK_SECONDS,
        allow_refresh=False
    )

    m = modal(page)
    random_human_wait(page, 2000, 5000)
    m.locator("#genreId").select_option(label="Turnitin Reports")
    random_human_wait(page, 4000, 7000)
    ui_log("[OJS UPLOAD] ✅ Selected 'Turnitin Reports'")


def upload_file(page, file_path):
    m = modal(page)
    file_inputs = m.locator("input[type='file']")
    ui_log(f"[OJS UPLOAD] Modal file inputs found: {file_inputs.count()}")

    if file_inputs.count() == 0:
        raise Exception("No file input found in modal")

    random_human_wait(page, 4000, 7000)
    file_inputs.last.set_input_files(file_path)
    random_human_wait(page, 7000, 12000)
    ui_log(f"[OJS UPLOAD] ✅ File set: {file_path}")


def wait_continue_enabled(page):
    wait_until(
        page,
        "Continue button enabled",
        """
        () => {
            const btn = document.querySelector(
                '.pkp_modal.is_visible #continueButton, .pkpModalWrapper.is_visible #continueButton'
            );
            return btn && !btn.disabled;
        }
        """,
        timeout_seconds=90,
        allow_refresh=False
    )


def click_continue(page):
    wait_continue_enabled(page)
    random_human_wait(page, 2000, 5000)
    btn = modal(page).locator("#continueButton").first
    btn.scroll_into_view_if_needed(timeout=10000)
    random_human_wait(page, 2000, 4000)
    btn.click(timeout=20000)
    ui_log("[OJS UPLOAD] ✅ CLICKED Continue")


def review_details_and_continue(page, file_name):
    """Fill file name on the Review Details step, then click Continue."""
    ui_log("[OJS UPLOAD] Waiting Review Details tab (Name the file)...")

    wait_until(
        page,
        "Review Details / Name the file",
        """
        () => {
            const modal = document.querySelector('.pkp_modal.is_visible, .pkpModalWrapper.is_visible');
            return modal && modal.innerText.includes('Name the file');
        }
        """,
        timeout_seconds=90,
        allow_refresh=False
    )

    m = modal(page)
    random_human_wait(page, 4000, 7000)

    name_input = m.locator(
        'input[name="name"], input[id*="name"], input[type="text"]'
    ).first
    name_input.fill(file_name)
    ui_log(f"[OJS UPLOAD] Filled file name: {file_name}")

    random_human_wait(page, 4000, 7000)
    click_continue(page)


def click_complete(page):
    ui_log("[OJS UPLOAD] Waiting Confirm tab / File Added...")

    wait_until(
        page,
        "Confirm tab / File Added",
        """
        () => {
            const modal = document.querySelector('.pkp_modal.is_visible, .pkpModalWrapper.is_visible');
            return modal && modal.innerText.includes('File Added');
        }
        """,
        timeout_seconds=90,
        allow_refresh=False
    )

    m = modal(page)
    random_human_wait(page, 4000, 7000)

    selectors = [
        'button:has-text("Complete")',
        'input[value="Complete"]',
        'text=Complete',
    ]

    for sel in selectors:
        try:
            loc = m.locator(sel).first
            count = m.locator(sel).count()
            ui_log(f"[OJS UPLOAD] TRY Complete selector: {sel} | count={count}")
            if count > 0:
                loc.scroll_into_view_if_needed(timeout=10000)
                random_human_wait(page, 2000, 5000)
                loc.click(timeout=20000)
                ui_log("[OJS UPLOAD] ✅ CLICKED Complete")
                return True
        except Exception as e:
            ui_log(f"[OJS UPLOAD] Complete selector failed: {sel} — {e}")

    raise Exception("Complete button not found")


def run_upload_flow(page, submission_id, file_path, file_name):
    """
    Run the 5-step OJS upload flow for one submission:
    Upload File → Select Turnitin Reports → Upload File → Continue →
    Fill Name → Continue → Complete
    """
    ui_log(f"[OJS UPLOAD] ▶ Starting upload flow for SUB-{submission_id}")
    ui_log(f"[OJS UPLOAD] Uploading report: {file_name}")

    click_upload_file(page)
    random_human_wait(page, 4000, 7000)

    select_turnitin_reports(page)
    random_human_wait(page, 4000, 7000)

    upload_file(page, file_path)
    random_human_wait(page, 7000, 12000)

    click_continue(page)
    random_human_wait(page, 7000, 12000)

    review_details_and_continue(page, file_name)
    random_human_wait(page, 7000, 12000)

    click_complete(page)
    random_human_wait(page, 7000, 12000)

    ui_log(f"[OJS UPLOAD] ✅ Done upload Turnitin Reports for SUB-{submission_id}")


# ==============================================================================
# MAIN CALLABLE FUNCTION
# ==============================================================================

def run_ojs_report_upload_workflow(log_callback=None):
    """
    Run OJS Report Upload workflow in the calling thread.

    Reads pending rows from sheet 'turnitin' (status=BERJAYA, valid report path,
    not yet in sheet 'upload' with BERJAYA UPLOAD OJS).

    Navigates: OJS Home → Dashboard → Submissions → Active → Search ID → Open Row
    Then runs: Upload File → Turnitin Reports → upload → Continue → name → Complete

    Logs result to sheet 'upload'.

    Returns: dict { status: 'success'|'no_submissions'|'error', message: str, ... }
    """
    global _workflow_running

    if log_callback:
        set_ui_log_callback(log_callback)

    # Prevent concurrent runs
    with _workflow_lock:
        if _workflow_running:
            ui_log("[OJS UPLOAD] ⚠️ Workflow already running. Please wait.")
            return {"status": "error", "message": "OJS Report Upload workflow already running"}
        _workflow_running = True

    _pw      = None
    _browser = None

    try:
        ui_log("[OJS UPLOAD] ═══════════════════════════════════════")
        ui_log("[OJS UPLOAD]   Upload Report to OJS — Starting")
        ui_log("[OJS UPLOAD] ═══════════════════════════════════════")

        # Find pending row from Excel
        ui_log("[OJS UPLOAD] Reading Excel for pending submissions...")
        row_data = find_next_pending_row()

        if row_data is None:
            ui_log("[OJS UPLOAD] ℹ️ No pending submissions to upload.")
            return {
                "status":  "no_submissions",
                "message": "No pending submissions found in sheet 'turnitin'."
            }

        submission_no        = row_data["submission_no"]
        turnitin_name        = row_data["turnitin_name"]
        turnitin_report_path = row_data["turnitin_report_path"]
        file_name            = Path(turnitin_report_path).name

        ui_log(f"[OJS UPLOAD] Submission No  : {submission_no}")
        ui_log(f"[OJS UPLOAD] Turnitin Name  : {turnitin_name}")
        ui_log(f"[OJS UPLOAD] Report Path    : {turnitin_report_path}")
        ui_log(f"[OJS UPLOAD] File Name      : {file_name}")

        # Remove stale lock file
        lock_file = Path(PROFILE_DIR) / "Default" / "LOCK"
        if lock_file.exists():
            ui_log("[OJS UPLOAD] Removing stale profile lock...")
            try:
                lock_file.unlink()
                ui_log("[OJS UPLOAD] Stale lock removed.")
            except Exception as le:
                ui_log(f"[OJS UPLOAD] Warning: Could not remove lock: {le}")

        # Launch browser
        ui_log("[OJS UPLOAD] Launching browser with persistent OJS profile...")
        _pw = sync_playwright().start()
        _browser = _pw.chromium.launch_persistent_context(
            user_data_dir=PROFILE_DIR,
            headless=False,
            slow_mo=HUMAN_SLOW_MO,
            accept_downloads=False,
        )

        page = _browser.pages[0] if _browser.pages else _browser.new_page()

        # Navigate to OJS
        ui_log("[OJS UPLOAD] Navigating to OJS submissions...")
        page.goto(OJS_SUBMISSIONS_URL, wait_until="domcontentloaded", timeout=120000)

        # Check login
        current_url = page.url
        if "/login" in current_url:
            ui_log("[OJS UPLOAD] Please login OJS manually. Workflow will continue automatically.")
            max_wait  = 300  # 5 minutes
            start_t   = time.time()
            logged_in = False

            while time.time() - start_t < max_wait:
                try:
                    current_url = page.url
                except Exception:
                    break
                if "/login" not in current_url:
                    ui_log("[OJS UPLOAD] ✅ Login detected. Continuing...")
                    random_human_wait(page, 3000, 5000)
                    logged_in = True
                    break
                ui_log("[OJS UPLOAD] Waiting for OJS login... (checking every 5s)")
                time.sleep(5)

            if not logged_in:
                raise Exception("Login timeout — user did not login within 5 minutes")
        else:
            ui_log("[OJS UPLOAD] ✅ Already logged in.")

        # Navigate to the submission
        ui_log(f"[OJS UPLOAD] Searching submission: {submission_no}")
        navigate_to_submission(page, submission_no)

        # Run upload flow
        run_upload_flow(page, submission_no, turnitin_report_path, file_name)

        # Log success to Excel
        log_upload_result(
            submission_no        = submission_no,
            turnitin_report_path = turnitin_report_path,
            turnitin_name        = turnitin_name,
            status               = STATUS_BERJAYA,
            notes                = f"Uploaded via ojs_report_upload_engine at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        )

        ui_log(f"[OJS UPLOAD] ✅ BERJAYA UPLOAD OJS — submission {submission_no}")
        ui_log("[OJS UPLOAD] Done upload Turnitin Reports sepenuhnya.")
        ui_log("[OJS UPLOAD] OJS Upload Completed")

        return {
            "status":       "success",
            "message":      f"BERJAYA UPLOAD OJS — submission {submission_no}",
            "submission_no": submission_no,
            "report_path":  turnitin_report_path,
        }

    except Exception as e:
        err_msg = str(e)
        ui_log(f"[OJS Upload] ❌ Upload failed: {err_msg}")

        # Log failure to Excel (best effort — row_data might not exist yet)
        try:
            if 'row_data' in dir() and row_data:
                log_upload_result(
                    submission_no        = row_data.get("submission_no", ""),
                    turnitin_report_path = row_data.get("turnitin_report_path", ""),
                    turnitin_name        = row_data.get("turnitin_name", ""),
                    status               = STATUS_TAK,
                    notes                = err_msg[:400]
                )
                ui_log(f"[OJS UPLOAD] TAK BERJAYA UPLOAD OJS logged for {row_data.get('submission_no', '')}")
        except Exception as log_err:
            ui_log(f"[OJS UPLOAD] Could not log failure: {log_err}")

        return {
            "status":  "error",
            "message": err_msg,
        }

    finally:
        # Always release workflow lock and close browser
        _workflow_running = False
        ui_log("[OJS UPLOAD] 🔓 Workflow lock released.")

        if _browser:
            try:
                _browser.close()
                ui_log("[OJS UPLOAD] Browser closed.")
            except Exception:
                pass

        if _pw:
            try:
                _pw.stop()
            except Exception:
                pass
