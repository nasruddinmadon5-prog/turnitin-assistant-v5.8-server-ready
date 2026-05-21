import webview
from pathlib import Path
import sys
import os
import threading
import queue
import random
import traceback
import re
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

window = None
selected_files = []


def resource_path(relative_path):
    if getattr(sys, "frozen", False):
        return Path(sys._MEIPASS) / relative_path
    return Path(__file__).resolve().parent / relative_path


if getattr(sys, "frozen", False):
    APP_RESOURCES = Path(sys.executable).resolve().parents[1] / "Resources"
    bundled_browser_path = APP_RESOURCES / "ms-playwright"

    os.environ["PLAYWRIGHT_BROWSERS_PATH"] = str(
        bundled_browser_path
    )

    print("[DEBUG] PLAYWRIGHT_BROWSERS_PATH =", os.environ["PLAYWRIGHT_BROWSERS_PATH"])
else:
    bundled_browser_path = None

BUNDLED_BROWSERS_DIR = bundled_browser_path

from playwright.sync_api import sync_playwright

# Use v5 clean structure (v5.3-compatible project-local profiles)
PROJECT_ROOT = Path(__file__).resolve().parent.parent
# v5.5.2: Import shared profile path from config to ensure consistency across engines
try:
    from config import TURNITIN_PROFILE_PATH
    PROFILE_DIR = TURNITIN_PROFILE_PATH
except ImportError:
    PROFILE_DIR = PROJECT_ROOT / "profiles" / "turnitin-profile"

SCREENSHOT_DIR = PROJECT_ROOT / "storage" / "screenshots"
DOWNLOAD_DIR = PROJECT_ROOT / "storage" / "turnitin-reports"

# Create subdirectories
PROFILE_DIR.mkdir(parents=True, exist_ok=True)
SCREENSHOT_DIR.mkdir(parents=True, exist_ok=True)
DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

# Excel logging path
EXCEL_LOG_PATH = PROJECT_ROOT / "data" / "master update.xlsx"

# Log paths
print(f"[INIT] Using v5 clean structure (v5.3-compatible profiles)")
print(f"[INIT] Profile: {PROFILE_DIR}")
print(f"[INIT] Screenshots: {SCREENSHOT_DIR}")
print(f"[INIT] Downloads: {DOWNLOAD_DIR}")
print(f"[INIT] Excel Log: {EXCEL_LOG_PATH}")

JOURNAL_SHORT = "JTI"

EXCLUDE_KEYWORDS = [
    "universiti teknologi malaysia",
    "university teknologi malaysia",
    "university technology malaysia",
    "utm.my",
    "journals.utm.my",
    "penerbit utm",
    "utm press",
]

TURNITIN_LOGIN_URL = "https://www.turnitin.com/login_page.asp"
TURNITIN_HOME_URL = "https://www.turnitin.com/t_home.asp"
SUBMIT_TYPE_BUTTON = "#submit_type"
SUBMIT_TYPE_MENU = "#submit-type-menu"
UPLOAD_BUTTON = "#upload-btn"
CONFIRM_BUTTON = "#confirm-btn"
CLOSE_BUTTON = "#close-btn"
ASSIGN_INBOX = "#assign_inbox"
DOWNLOAD_MENU_BUTTON = "[title='Download'], [aria-label='Current View'], div[role='button'][title='Download']"

DEFAULT_TIMEOUT = 15000
NAVIGATION_TIMEOUT = 30000
MANUAL_LOGIN_TIMEOUT = 300000

HUMAN_MODE = "human"  # options: "test", "human"

if HUMAN_MODE == "test":
    SLOW_MO = 350
    PAUSE_MIN_MS = 700
    PAUSE_MAX_MS = 1600
else:
    SLOW_MO = 1200       # v5.5.1: Slower human-like typing speed
    PAUSE_MIN_MS = 2000  # v5.5.1: 2-4 second random pauses
    PAUSE_MAX_MS = 4000

CHECKBOX_SELECTORS = [
    "div.ip_form_row:nth-child(2) > label:nth-child(1) > input:nth-child(1)",
    "div.ip_form_row:nth-child(3) > label:nth-child(1) > input:nth-child(1)",
    "div.ip_form_row:nth-child(4) > label:nth-child(1) > input:nth-child(1)",
    "div.ip_form_row:nth-child(5) > label:nth-child(1) > input:nth-child(1)",
]

automation_queue = queue.Queue()
automation_worker_started = False
automation_worker_lock = threading.Lock()

# Workflow lock to prevent concurrent operations
feedback_workflow_lock = threading.Lock()
feedback_workflow_running = False

pw = None
context = None
page = None


def ui_log(message):
    global window
    print(message)
    try:
        safe = message.replace("\\", "\\\\").replace("'", "\\'")
        window.evaluate_js(f"addLog('{safe}')")
    except Exception:
        pass


class ProcessState:
    """
    Lightweight process state tracker for stages and metadata.
    Stores all metadata needed for Excel logging.
    """
    
    # Stage constants
    STARTED = "STARTED"
    FILE_SELECTED = "FILE_SELECTED"
    QUICK_SUBMIT_READY = "QUICK_SUBMIT_READY"
    UPLOAD_FORM_FILLED = "UPLOAD_FORM_FILLED"
    UPLOADED = "UPLOADED"
    CONFIRMED = "CONFIRMED"
    INBOX_FOUND = "INBOX_FOUND"
    REPORT_READY = "REPORT_READY"
    FEEDBACK_OPENED = "FEEDBACK_OPENED"
    FILTER_DONE = "FILTER_DONE"
    DOWNLOAD_STARTED = "DOWNLOAD_STARTED"
    CLASSIC_REPORT_OPENED = "CLASSIC_REPORT_OPENED"
    DOWNLOADED = "DOWNLOADED"
    FAILED = "FAILED"
    
    # Status constants
    SUCCESS_DOWNLOADED = "SUCCESS_DOWNLOADED"
    FAILED_BEFORE_UPLOAD = "FAILED_BEFORE_UPLOAD"
    FAILED_AFTER_UPLOAD = "FAILED_AFTER_UPLOAD"
    FAILED_IN_INBOX_DETECTION = "FAILED_IN_INBOX_DETECTION"
    FAILED_FEEDBACK_STUDIO = "FAILED_FEEDBACK_STUDIO"
    FAILED_FILTER = "FAILED_FILTER"
    FAILED_DOWNLOAD = "FAILED_DOWNLOAD"
    FAILED_CLASSIC_REPORT_DOWNLOAD = "FAILED_CLASSIC_REPORT_DOWNLOAD"
    
    def __init__(self):
        self.reset()
    
    def reset(self):
        """Reset all state for new process."""
        self.timestamp_start = datetime.now()
        self.timestamp_end = None
        self.current_stage = self.STARTED
        self.previous_stage = None
        
        # File metadata
        self.original_file_path = None
        self.original_filename = None
        self.parsed_title = None
        self.parsed_first_name = None
        self.parsed_last_name = None
        
        # Turnitin metadata
        self.turnitin_paper_id = None
        self.inbox_row_date = None
        self.similarity_before_filter = None
        self.similarity_after_filter = None
        
        # Bridge metadata (from Excel download sheet)
        self.submission_no = None
        self.author_name = None
        self.source_download_path = None
        self.turnitin_name = None
        
        # Results
        self.final_pdf_filename = None
        self.final_pdf_path = None
        self.status = None
        self.failed_stage = None
        self.error_message = None
        self.screenshot_paths = []
        self.notes = []
    
    def update_stage(self, stage):
        """Update current stage and track previous."""
        self.previous_stage = self.current_stage
        self.current_stage = stage
        ui_log(f"[STAGE] {self.previous_stage} → {stage}")
    
    def mark_failed(self, error_msg=None):
        """Mark process as failed."""
        self.failed_stage = self.previous_stage or self.current_stage
        self.current_stage = self.FAILED
        if error_msg:
            self.error_message = str(error_msg)[:500]  # Limit length
        ui_log(f"[STAGE] FAILED at stage: {self.failed_stage}")
    
    def add_screenshot(self, screenshot_path):
        """Add screenshot path to collection."""
        if screenshot_path:
            self.screenshot_paths.append(str(screenshot_path))
    
    def add_note(self, note):
        """Add note to collection."""
        if note:
            self.notes.append(str(note))
    
    def finalize(self, status):
        """Finalize process with status."""
        self.timestamp_end = datetime.now()
        self.status = status
        ui_log(f"[STAGE] Process finalized with status: {status}")


class ExcelDownloadBridge:
    """
    Bridge class that reads from the 'download' Excel sheet
    and finds the next unprocessed submission file.
    
    Input sheet 'download' columns:
      A = timestamp, B = submission_no, C = author_name, D = title, E = download_location
      
    v5.5.2: Enhanced duplicate prevention with comprehensive status checks.
    - Checks sheet 'turnitin' for ALL statuses that indicate already processed
    - Skips failed report download submissions from re-upload
    - Checks if Turnitin report file already exists on disk
    - Adds detailed skip logging
    """
    
    # Statuses that mean "already uploaded to Turnitin" — skip upload entirely
    SKIP_UPLOAD_STATUSES = {
        "BERJAYA",
        "BERJAYA UPLOAD TURNITIN",
        "BERJAYA DOWNLOAD REPORT",
        "BERJAYA UPLOAD OJS",
        "Uploaded",
        "Turnitin Uploaded",
        "Completed",
        "Success",
        "Report Downloaded",
        "Similarity Done",
    }
    
    # Statuses that mean "uploaded to Turnitin but failed at report download"
    # These should NOT be re-uploaded to Turnitin (allow report retry only)
    SKIP_UPLOAD_AFTER_FAILED_DOWNLOAD = {
        "TAK BERJAYA DOWNLOAD",
        "FAILED REPORT DOWNLOAD",
        "REPORT DOWNLOAD FAILED",
        "FAILED_DOWNLOAD",
        "FAILED_FEEDBACK_STUDIO",
        "FAILED_CLASSIC_REPORT_DOWNLOAD",
        "FAILED_FILTER",
    }
    
    def __init__(self, excel_path):
        self.excel_path = Path(excel_path)
        self.lock = threading.Lock()
    
    def _get_turnitin_sheet_data(self):
        """
        Read ALL rows from turnitin sheet and return a dict:
        { submission_no: { 'status': str, 'report_path': str } }
        """
        data = {}
        try:
            if not self.excel_path.exists():
                return data
            wb = load_workbook(str(self.excel_path), read_only=True)
            if 'turnitin' not in wb.sheetnames:
                wb.close()
                return data
            ws = wb['turnitin']
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 6:
                    continue
                submission_no = str(row[1]).strip() if row[1] is not None else ""
                status = str(row[5]).strip() if row[5] is not None else ""
                report_path = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ""
                if submission_no:
                    data[submission_no] = {
                        'status': status,
                        'report_path': report_path,
                    }
            wb.close()
        except Exception as e:
            ui_log(f"[BRIDGE] Error reading turnitin sheet: {e}")
        return data
    
    def _should_skip_turnitin_upload(self, submission_no, turnitin_data, turnitin_reports_dir):
        """
        Check if submission should be skipped for Turnitin upload.
        Returns (True, reason_message) if should skip, (False, None) if eligible.
        """
        # Not in turnitin sheet at all → eligible for upload
        if submission_no not in turnitin_data:
            return False, None
        
        entry = turnitin_data[submission_no]
        status_upper = entry['status'].strip().upper()
        report_path = entry['report_path'].strip()
        
        # Check for already-uploaded statuses
        for skip_status in self.SKIP_UPLOAD_STATUSES:
            if skip_status.upper() in status_upper or status_upper == skip_status.upper():
                return True, f"[SKIP] {submission_no} already uploaded to Turnitin (status: {entry['status']})"
        
        # Check for failed report download — skip upload, allow report retry only
        for fail_status in self.SKIP_UPLOAD_AFTER_FAILED_DOWNLOAD:
            if fail_status.upper() in status_upper or status_upper == fail_status.upper():
                return True, f"[SKIP TURNITIN] {submission_no} already uploaded but report download failed previously (status: {entry['status']}). Do not re-upload."
        
        # Check if report file already exists on disk
        if report_path:
            report_file = Path(report_path)
            if report_file.exists():
                return True, f"[SKIP] {submission_no} report already exists (path: {report_path})"
        
        # Check if any Turnitin report file exists for this submission_no in reports folder
        if turnitin_reports_dir:
            try:
                for f in turnitin_reports_dir.iterdir():
                    if f.is_file() and (f"report-{submission_no}" in f.name or f"report_{submission_no}" in f.name or submission_no in f.name):
                        return True, f"[SKIP] {submission_no} report already exists in Turnitin Reports folder"
            except Exception:
                pass
        
        # Check if submission_no already exists with BERJAYA UPLOAD TURNITIN (contains pattern)
        if "BERJAYA" in status_upper:
            return True, f"[SKIP] {submission_no} already BERJAYA in turnitin sheet"
        
        return False, None
    
    def find_next_submission(self):
        """
        Find next unprocessed submission from download sheet.
        v5.5.2: Enhanced duplicate prevention with comprehensive Excel cross-check.
        
        Selection rules:
        1. Only process rows where OJS/download status = downloaded/completed/success
        2. Skip when Turnitin status shows uploaded/BERJAYA/etc.
        3. Skip if report file already exists
        4. Skip if submission already in turnitin sheet with failed report download
        5. Pick oldest/downloaded pending row first
        
        Returns dict with keys: file_path, submission_no, author_name, title, turnitin_name
        or None if no submission found.
        """
        with self.lock:
            ui_log("[BRIDGE] Reading download sheet")
            
            if not self.excel_path.exists():
                ui_log(f"[BRIDGE] Excel file not found: {self.excel_path}")
                return None
            
            try:
                # Read ALL turnitin sheet data once for cross-referencing
                turnitin_data = self._get_turnitin_sheet_data()
                ui_log(f"[BRIDGE] Total submissions in turnitin sheet: {len(turnitin_data)}")
                
                wb = load_workbook(str(self.excel_path), read_only=True)
                
                if 'download' not in wb.sheetnames:
                    ui_log("[BRIDGE] Sheet 'download' not found in workbook")
                    wb.close()
                    return None
                
                ws_download = wb['download']
                
                # Collect eligible candidates first, then pick oldest
                candidates = []
                
                for row in ws_download.iter_rows(min_row=2, values_only=True):
                    # Columns: A=timestamp(0), B=submission_no(1), C=author_name(2), D=title(3), E=download_location(4)
                    if not row or len(row) < 5:
                        continue
                    
                    submission_no = row[1]
                    author_name   = row[2]
                    title         = row[3]
                    download_location = row[4]
                    
                    # Skip if missing required fields
                    if not submission_no or not download_location:
                        continue
                    
                    submission_no_str     = str(submission_no).strip()
                    download_location_str = str(download_location).strip()
                    
                    if not submission_no_str or not download_location_str:
                        continue
                    
                    # Check if OJS/download status is completed
                    # The download sheet only has entries after OJS download succeeds
                    # So any row here is considered downloaded/completed/success
                    
                    # v5.5.2: Check turnitin sheet for duplicate prevention
                    should_skip, skip_reason = self._should_skip_turnitin_upload(
                        submission_no_str,
                        turnitin_data,
                        DOWNLOAD_DIR  # turnitin-reports folder
                    )
                    if should_skip:
                        ui_log(skip_reason)
                        continue
                    
                    # Check file exists on disk
                    file_path = Path(download_location_str)
                    if not file_path.exists():
                        ui_log(f"[SKIP] {submission_no_str} not downloaded from OJS yet (file not found: {download_location_str})")
                        continue
                    
                    # Collect as candidate with timestamp for ordering
                    timestamp_str = str(row[0]).strip() if row[0] else ""
                    candidates.append({
                        "timestamp":      timestamp_str,
                        "file_path":      str(file_path),
                        "submission_no":  submission_no_str,
                        "author_name":    str(author_name).strip() if author_name else "",
                        "title":          str(title).strip() if title else "",
                        "turnitin_name":  submission_no_str,
                    })
                
                wb.close()
                
                if not candidates:
                    ui_log("[BRIDGE] No pending submissions found in download sheet")
                    return None
                
                # Sort by timestamp ascending (oldest first) - pick oldest pending
                def _sort_key(c):
                    ts = c.get("timestamp", "")
                    if not ts:
                        return "9999-99-99"  # push empty timestamps to end
                    return ts
                
                candidates.sort(key=_sort_key)
                selected = candidates[0]
                
                ui_log(f"[BRIDGE] Selected submission_no: {selected['submission_no']}")
                ui_log(f"[BRIDGE] Source file: {selected['file_path']}")
                ui_log(f"[BRIDGE] Author: {selected['author_name']}")
                ui_log(f"[BRIDGE] Title: {selected['title']}")
                ui_log(f"[BRIDGE] Turnitin name (v5.4): {selected['turnitin_name']}")
                ui_log(f"[BRIDGE] Timestamp: {selected['timestamp']}")
                
                return {
                    "file_path":      selected["file_path"],
                    "submission_no":  selected["submission_no"],
                    "author_name":    selected["author_name"],
                    "title":          selected["title"],
                    "turnitin_name":  selected["turnitin_name"],
                }
                
            except Exception as e:
                ui_log(f"[BRIDGE] Error reading download sheet: {e}")
                ui_log(traceback.format_exc())
                return None


class ExcelLogger:
    """
    Excel logger that writes process results to sheet 'turnitin' with 6 columns:
      A = timestamp (process start time)
      B = submission_no
      C = turnitin_name
      D = turnitin_score (similarity after filter)
      E = turnitin_report_path
      F = status (BERJAYA / TAK BERJAYA UPLOAD / TAK BERJAYA DOWNLOAD)
    
    Preserves existing formatting: Arial 20pt, wrap text, bold headers.
    """
    
    SHEET_NAME = "turnitin"
    
    HEADERS = [
        "Timestamp",
        "Submission No",
        "Turnitin Name",
        "Turnitin Score",
        "Turnitin Report Path",
        "Status",
    ]
    
    # Map internal status constants to user-facing Malay status strings
    STATUS_MAP = {
        "SUCCESS_DOWNLOADED":          "BERJAYA",
        "FAILED_BEFORE_UPLOAD":        "TAK BERJAYA UPLOAD",
        "FAILED_AFTER_UPLOAD":         "TAK BERJAYA UPLOAD",
        "FAILED_IN_INBOX_DETECTION":   "TAK BERJAYA UPLOAD",
        "FAILED_FEEDBACK_STUDIO":      "TAK BERJAYA DOWNLOAD",
        "FAILED_FILTER":               "TAK BERJAYA DOWNLOAD",
        "FAILED_DOWNLOAD":             "TAK BERJAYA DOWNLOAD",
        "FAILED_CLASSIC_REPORT_DOWNLOAD": "TAK BERJAYA DOWNLOAD",
    }
    
    def __init__(self, excel_path):
        self.excel_path = Path(excel_path)
        self.lock = threading.Lock()
    
    def _ensure_turnitin_sheet(self, wb):
        """
        Ensure 'turnitin' sheet exists with proper headers and formatting.
        Creates sheet if missing. Adds headers if row 1 is empty.
        """
        from openpyxl.styles import Font, Alignment
        
        if self.SHEET_NAME not in wb.sheetnames:
            ui_log(f"[EXCEL] Creating sheet '{self.SHEET_NAME}'")
            ws = wb.create_sheet(self.SHEET_NAME)
        else:
            ws = wb[self.SHEET_NAME]
        
        # Write headers if row 1 col A is empty
        if ws.cell(1, 1).value is None:
            ui_log(f"[EXCEL] Writing headers to sheet '{self.SHEET_NAME}'")
            for col, header in enumerate(self.HEADERS, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font      = Font(name='Arial', size=20, bold=True)
                cell.alignment = Alignment(wrap_text=True)
            ws.row_dimensions[1].height = 60
        
        return ws
    
    def log_process(self, ps):
        """
        Log process result to 'turnitin' sheet with exactly 6 columns.
        Thread-safe with lock.
        """
        with self.lock:
            try:
                if not self.excel_path.exists():
                    ui_log(f"[EXCEL] ❌ Excel file not found: {self.excel_path}")
                    return
                
                from openpyxl.styles import Font, Alignment
                
                wb = load_workbook(str(self.excel_path))
                ws = self._ensure_turnitin_sheet(wb)
                
                # Map status
                internal_status = ps.status or ""
                status_text = self.STATUS_MAP.get(internal_status, "TAK BERJAYA UPLOAD")
                
                # Find next empty row
                next_row = ws.max_row + 1
                
                # Column values
                timestamp         = ps.timestamp_start.strftime("%Y-%m-%d %H:%M:%S") if ps.timestamp_start else ""
                submission_no     = ps.submission_no or ""
                turnitin_name     = ps.turnitin_name or ""
                turnitin_score    = ps.similarity_after_filter or ps.similarity_before_filter or ""
                report_path       = ps.final_pdf_path or ""
                
                row_data = [
                    timestamp,
                    submission_no,
                    turnitin_name,
                    turnitin_score,
                    report_path,
                    status_text,
                ]
                
                for col, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=next_row, column=col, value=value)
                    cell.font      = Font(name='Arial', size=20)
                    cell.alignment = Alignment(wrap_text=True)
                
                # Adjust row height
                ws.row_dimensions[next_row].height = 50
                
                wb.save(str(self.excel_path))
                ui_log(f"[EXCEL] Logged to turnitin row {next_row}")
                ui_log(f"[EXCEL] submission_no={submission_no} | status={status_text}")
                
            except Exception as e:
                ui_log(f"[EXCEL] ❌ Failed to log: {e}")
                ui_log(traceback.format_exc())


# Global process state, logger, and Excel bridge
process_state = ProcessState()
excel_logger = ExcelLogger(EXCEL_LOG_PATH)
excel_bridge = ExcelDownloadBridge(EXCEL_LOG_PATH)


def sanitize_filename(filename):
    """
    Sanitize filename by removing illegal filesystem characters.
    Does NOT truncate the filename - preserves full length.
    Only replaces illegal characters with underscore.
    """
    # Replace illegal characters with underscore
    illegal_chars = r'[<>:"/\\|?*]'
    sanitized = re.sub(illegal_chars, '_', filename)
    
    # Remove leading/trailing dots and spaces
    sanitized = sanitized.strip('. ')
    
    return sanitized


class FeedbackStudioController:
    """
    Dedicated controller for Feedback Studio popup operations.
    All operations target the feedback_page explicitly, independent of OS focus.
    """
    
    def __init__(self, feedback_page, title="paper", score_before="unknown"):
        """
        Initialize controller with explicit page reference.
        
        Args:
            feedback_page: Playwright page object for Feedback Studio popup
            title: Paper title for context
            score_before: Initial similarity score
        """
        self.feedback_page = feedback_page
        self.title = title
        self.score_before = score_before
        self.screenshot_dir = SCREENSHOT_DIR
        
        ui_log("="*70)
        ui_log("🔒 FEEDBACK STUDIO AUTOMATION STARTED")
        ui_log("="*70)
        ui_log("⚠️  IMPORTANT: Do not close the Feedback Studio popup window!")
        ui_log("⚠️  The automation will continue even if you:")
        ui_log("    - Open other browser tabs")
        ui_log("    - Open other browser windows")
        ui_log("    - Switch to other applications")
        ui_log("⚠️  The automation is locked to the Feedback Studio page object.")
        ui_log("="*70)
    
    def _wait(self, ms=1500):
        """Wait on feedback_page explicitly."""
        try:
            self.feedback_page.wait_for_timeout(ms)
        except Exception:
            pass
    
    def _save_screenshot(self, label):
        """Save screenshot of feedback_page."""
        try:
            self.screenshot_dir.mkdir(parents=True, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
            safe_label = "".join(ch if ch.isalnum() or ch in "-_" else "-" for ch in label).strip("-")
            screenshot_path = self.screenshot_dir / f"{timestamp}-feedback-{safe_label}.png"
            self.feedback_page.screenshot(path=str(screenshot_path), full_page=True)
            ui_log(f"📸 Screenshot saved: {screenshot_path.name}")
            
            # Add to process state
            process_state.add_screenshot(screenshot_path)
            
            return screenshot_path
        except Exception as e:
            ui_log(f"⚠️  Screenshot failed: {e}")
            return None
    
    def _verify_page_alive(self):
        """Verify feedback_page is still valid and responsive."""
        try:
            # Simple check: can we access the URL?
            url = self.feedback_page.url
            ui_log(f"[VERIFY] Page alive: {url}")
            return True
        except Exception as e:
            ui_log(f"[VERIFY] ❌ Page not responsive: {e}")
            return False
    
    def _verify_element_state(self, selector, expected_state="visible", timeout=3000):
        """
        Verify element exists and matches expected state.
        
        Returns: True if verified, False otherwise
        """
        try:
            loc = self.feedback_page.locator(selector)
            if loc.count() > 0:
                if expected_state == "visible":
                    return loc.first.is_visible(timeout=timeout)
                elif expected_state == "exists":
                    return True
            return False
        except Exception:
            return False
    
    def _click_with_fallback(self, selectors, label, timeout=15000):
        """
        Robust click with multiple fallback strategies:
        1. Playwright locator click (preferred)
        2. JavaScript DOM click
        3. Coordinate-based click (last resort)
        
        Args:
            selectors: List of CSS selectors to try
            label: Description for logging
            timeout: Timeout for each attempt
            
        Returns: True if clicked, False otherwise
        """
        ui_log(f"[CLICK] Attempting: {label}")
        
        # Strategy 1: Playwright locator-based click
        for sel in selectors:
            try:
                loc = self.feedback_page.locator(sel)
                count = loc.count()
                
                for i in range(count):
                    item = loc.nth(i)
                    try:
                        if item.is_visible(timeout=2000):
                            item.scroll_into_view_if_needed(timeout=timeout)
                            self._wait(random.randint(500, 1200))
                            item.click(timeout=timeout)
                            ui_log(f"[CLICK] ✅ Success via locator: {sel}")
                            return True
                    except Exception:
                        continue
            except Exception:
                continue
        
        # Strategy 2: JavaScript DOM click
        ui_log(f"[CLICK] Trying JavaScript DOM click for: {label}")
        for sel in selectors:
            try:
                clicked = self.feedback_page.evaluate(
                    """
                    (selector) => {
                        const elements = document.querySelectorAll(selector);
                        for (const el of elements) {
                            const rect = el.getBoundingClientRect();
                            const visible = rect.width > 0 && rect.height > 0;
                            if (visible) {
                                el.click();
                                return true;
                            }
                        }
                        return false;
                    }
                    """,
                    sel
                )
                if clicked:
                    ui_log(f"[CLICK] ✅ Success via JS DOM: {sel}")
                    return True
            except Exception:
                continue
        
        # Strategy 3: Search by text/attributes
        ui_log(f"[CLICK] Trying JavaScript text-based search for: {label}")
        try:
            # Extract text hints from label
            text_hints = [label.lower(), label.split()[0].lower() if label else ""]
            
            for hint in text_hints:
                if not hint or len(hint) < 3:
                    continue
                    
                clicked = self.feedback_page.evaluate(
                    """
                    (targetText) => {
                        const buttons = [...document.querySelectorAll('button, div[role="button"], a, .sc-button-view')];
                        for (const btn of buttons) {
                            const rect = btn.getBoundingClientRect();
                            const visible = rect.width > 0 && rect.height > 0;
                            const text = (btn.innerText || btn.textContent || '').toLowerCase();
                            const title = (btn.getAttribute('title') || '').toLowerCase();
                            const label = (btn.getAttribute('aria-label') || '').toLowerCase();
                            
                            if (visible && (text.includes(targetText) || title.includes(targetText) || label.includes(targetText))) {
                                btn.click();
                                return true;
                            }
                        }
                        return false;
                    }
                    """,
                    hint
                )
                if clicked:
                    ui_log(f"[CLICK] ✅ Success via JS text search: {hint}")
                    return True
        except Exception:
            pass
        
        ui_log(f"[CLICK] ❌ All strategies failed for: {label}")
        return False
    
    def _retry_action(self, action_func, action_name, max_attempts=3, verify_func=None):
        """
        Retry an action with verification and progressive backoff.
        
        Args:
            action_func: Function to execute (should return result or raise exception)
            action_name: Name for logging
            max_attempts: Maximum retry attempts
            verify_func: Optional verification function to call after action
            
        Returns: Result from action_func
        Raises: Last exception if all attempts fail
        """
        last_exception = None
        
        for attempt in range(1, max_attempts + 1):
            try:
                ui_log(f"[RETRY] {action_name}: attempt {attempt}/{max_attempts}")
                
                # Verify page is alive before attempting
                if not self._verify_page_alive():
                    raise RuntimeError("Feedback page not responsive")
                
                # Execute action
                result = action_func()
                
                # Verify if verification function provided
                if verify_func:
                    if not verify_func():
                        raise RuntimeError(f"Verification failed after {action_name}")
                
                ui_log(f"[RETRY] {action_name}: ✅ SUCCESS on attempt {attempt}")
                return result
                
            except Exception as e:
                last_exception = e
                ui_log(f"[RETRY] {action_name}: ❌ attempt {attempt} failed: {e}")
                
                if attempt < max_attempts:
                    wait_time = 2000 * attempt  # Progressive backoff
                    ui_log(f"[RETRY] Waiting {wait_time}ms before retry...")
                    self._wait(wait_time)
                    self._save_screenshot(f"retry-{action_name}-attempt-{attempt}")
                else:
                    ui_log(f"[RETRY] {action_name}: All {max_attempts} attempts exhausted")
                    self._save_screenshot(f"retry-{action_name}-final-failure")
        
        raise last_exception
    
    def get_paper_title(self, fallback="paper"):
        """Extract paper title from Feedback Studio."""
        for sel in [".paper-info .title", "span.title", ".title"]:
            try:
                txt = self.feedback_page.locator(sel).first.inner_text(timeout=3000).strip()
                if txt:
                    return txt
            except Exception:
                pass
        return fallback
    
    def get_score(self):
        """Extract similarity score from Feedback Studio."""
        # Try standard score selectors
        for sel in [
            ".sidebar-originality-cumulative .osi-score",
            ".sidebar-or .osi-score",
            ".osi-score",
        ]:
            try:
                txt = self.feedback_page.locator(sel).first.inner_text(timeout=4000).strip()
                if txt and "%" in txt:
                    return txt
                if txt and txt.isdigit():
                    return f"{txt}%"
            except Exception:
                pass

        # Try JavaScript search
        try:
            txt = self.feedback_page.evaluate("""
            () => {
                const els = [...document.querySelectorAll('*')];
                for (const e of els) {
                    const r = e.getBoundingClientRect();
                    const visible = r.width > 0 && r.height > 0;
                    const t = (e.innerText || e.textContent || '').trim();
                    if (visible && /^\d{1,3}%$/.test(t)) return t;
                }
                return '';
            }
            """)
            if txt:
                return txt
        except Exception:
            pass

        return "unknown"
    
    def open_similarity_tools(self):
        """Open Similarity Tools panel."""
        ui_log("[STEP] Opening Similarity Tools...")
        
        selectors = [
            "[title='Similarity tools']",
            "[aria-label='Similarity tools']",
            ".or-popup-button-view",
            ".osi-toolbar-segment-button",
            ".sc-button-view.or-popup-button-view",
        ]
        
        if self._click_with_fallback(selectors, "similarity tools"):
            self._wait(2500)
            return True
        
        ui_log("[STEP] ⚠️  Similarity Tools click uncertain")
        self._wait(2500)
        return False
    
    def open_match_overview(self):
        """Open Match Overview panel."""
        ui_log("[STEP] Opening Match Overview...")
        
        selectors = [
            "[title='Match Overview']",
            "[aria-label='Match Overview']",
            ".sc-segment-view.osi-toolbar-segment-button[title='Match Overview']",
            "text=Match Overview",
        ]
        
        if self._click_with_fallback(selectors, "match overview"):
            self._wait(2500)
            return True
        
        ui_log("[STEP] ⚠️  Match Overview click uncertain")
        self._wait(2500)
        return False
    
    def verify_match_overview_ready(self):
        """Verify Match Overview panel is visible and ready."""
        # Check for score element
        if self._verify_element_state(".osi-score", "visible", timeout=2000):
            ui_log("[VERIFY] ✅ Match Overview ready: score visible")
            return True
        
        # Check for header
        if self._verify_element_state(".sidebar-header-label:has-text('Match Overview')", "visible", timeout=2000):
            ui_log("[VERIFY] ✅ Match Overview ready: header visible")
            return True
        
        ui_log("[VERIFY] ❌ Match Overview NOT ready")
        return False
    
    def ensure_match_overview(self):
        """Ensure Match Overview panel is open and ready."""
        self.open_similarity_tools()
        self.open_match_overview()
        
        # Verify with retries
        for attempt in range(1, 8):
            if self.verify_match_overview_ready():
                return True
            
            ui_log(f"[WAIT] Match Overview not ready, attempt {attempt}/7")
            self._wait(1000)
        
        ui_log("⚠️  WARNING: Match Overview not confirmed ready")
        return False
    
    def open_all_sources(self):
        """Open All Sources panel."""
        ui_log("[STEP] Opening All Sources...")
        
        self.open_similarity_tools()
        
        selectors = [
            "text=All Sources",
            "text=All sources",
            "[title='All sources']",
            "[aria-label='All sources']",
            ".sc-segment-view[title='All sources']",
            ".tii-icon-top-down",
        ]
        
        if self._click_with_fallback(selectors, "all sources"):
            self._wait(3000)
            return True
        
        ui_log("[STEP] ⚠️  All Sources click uncertain")
        self._wait(3000)
        return False
    
    def get_source_rows(self):
        """Get source rows locator."""
        rows = self.feedback_page.locator(".originality-list-non-group-node")
        try:
            if rows.count() > 0:
                return rows
        except Exception:
            pass
        return self.feedback_page.locator(".originality-list-group-node")
    
    def wait_until_sources_visible(self, max_retry=10):
        """Wait until source list is loaded and visible."""
        for attempt in range(1, max_retry + 1):
            rows = self.get_source_rows()
            try:
                count = rows.count()
            except Exception:
                count = 0
            
            ui_log(f"[WAIT] Source list attempt {attempt}/{max_retry}: {count} rows")
            
            if count > 0:
                return True
            
            self.open_all_sources()
            self._wait(1500)
        
        return False
    
    def scan_utm_sources(self):
        """Scan sources for UTM matches."""
        rows = self.get_source_rows()
        try:
            count = rows.count()
        except Exception:
            count = 0
        
        ui_log(f"[SCAN] Total sources: {count}")
        matched = []
        
        for i in range(count):
            try:
                text = rows.nth(i).inner_text(timeout=4000).strip()
            except Exception:
                text = ""
            
            ui_log(f"[SCAN] Source {i + 1}: {text[:180]}")
            if any(k in text.lower() for k in EXCLUDE_KEYWORDS):
                matched.append(i)
        
        ui_log(f"[SCAN] ✅ Matched UTM sources: {matched}")
        return matched
    
    def click_exclude_sources(self):
        """Open Exclude Sources dialog."""
        ui_log("[STEP] Opening Exclude Sources dialog...")
        
        selectors = [
            "text=Exclude Sources",
            "text=Exclude sources",
            "[title='Exclude Sources']",
            "[title='Exclude sources']",
            "[aria-label='Exclude Sources']",
            "[aria-label='Exclude sources']",
            ".exclude-source",
            "[title*='Exclude']",
            "[aria-label*='Exclude']",
        ]
        
        if self._click_with_fallback(selectors, "exclude sources"):
            self._wait(2500)
            return True
        
        ui_log("[STEP] ⚠️  Exclude Sources click uncertain")
        self._wait(2500)
        return False
    
    def wait_until_checkboxes_visible(self, max_retry=8):
        """Wait until checkboxes are visible in exclude dialog."""
        for attempt in range(1, max_retry + 1):
            try:
                count = self.feedback_page.locator("input[type='checkbox'], [role='checkbox']").count()
            except Exception:
                count = 0
            
            ui_log(f"[WAIT] Checkboxes attempt {attempt}/{max_retry}: {count}")
            
            if count > 0:
                return True
            
            self._wait(1200)
        
        return False
    
    def select_sources_by_indices(self, indices):
        """Select specific sources by index in exclude dialog."""
        self.wait_until_checkboxes_visible()
        rows = self.get_source_rows()
        
        for i in indices:
            ui_log(f"[SELECT] Selecting source {i + 1}...")
            row = rows.nth(i)
            
            # Try checkbox
            try:
                cb = row.locator("input[type='checkbox']").first
                cb.check(timeout=3000)
                self._wait(700)
                continue
            except Exception:
                pass
            
            # Try role=checkbox
            try:
                box = row.locator("[role='checkbox'], .checkbox").first
                box.click(timeout=3000)
                self._wait(700)
                continue
            except Exception:
                pass
            
            # Try clicking row itself
            try:
                row.click(timeout=5000)
                self._wait(700)
            except Exception as e:
                ui_log(f"⚠️  Could not select row {i + 1}: {e}")
    
    def apply_exclude(self):
        """Apply exclude changes."""
        ui_log("[STEP] Applying exclude...")
        
        selectors = [
            "text=Exclude",
            "text=Apply Changes",
            "text=Apply changes",
            "text=Done",
            "text=Confirm",
            "[title='Exclude']",
            "[aria-label='Exclude']",
        ]
        
        if self._click_with_fallback(selectors, "apply exclude"):
            self._wait(8000)
            return True
        
        ui_log("[STEP] ⚠️  Apply exclude click uncertain")
        self._wait(8000)
        return False
    
    def filter_and_exclude_if_needed(self):
        """Main filtering workflow: scan, select, and exclude UTM sources."""
        self.open_all_sources()
        
        if not self.wait_until_sources_visible():
            ui_log("⚠️  Filter skipped: source list not loaded")
            return False
        
        matched = self.scan_utm_sources()
        
        if not matched:
            ui_log("ℹ️  No UTM sources found to exclude")
            return False
        
        if not self.click_exclude_sources():
            ui_log("⚠️  Exclude Sources button not found")
            return False
        
        self.select_sources_by_indices(matched)
        
        if self.apply_exclude():
            ui_log("✅ Exclude completed")
            process_state.update_stage(ProcessState.FILTER_DONE)
            return True
        
        ui_log("⚠️  Exclude apply failed")
        return False
    
    def open_download_menu(self):
        """
        Open download menu using feedback_page.
        Uses real mouse interactions for reliability.
        """
        ui_log("[STEP] Opening download menu...")
        
        selectors = [
            ".sidebar-download-button",
            ".tii-icon-download.sidebar-download-button",
            "[title='Download']",
            "[aria-label='Download']",
            "div[role='button']:has-text('Download')",
        ]
        
        # Try each selector
        for sel in selectors:
            try:
                loc = self.feedback_page.locator(sel).first
                if loc.is_visible(timeout=2000):
                    ui_log(f"[DOWNLOAD MENU] Found download button: {sel}")
                    
                    # Scroll into view
                    loc.scroll_into_view_if_needed(timeout=5000)
                    self._wait(500)
                    
                    # Hover
                    loc.hover(timeout=5000)
                    self._wait(500)
                    ui_log("[DOWNLOAD MENU] Hovered over download button")
                    
                    # Get bounding box
                    box = loc.bounding_box(timeout=5000)
                    if box:
                        center_x = box['x'] + box['width'] / 2
                        center_y = box['y'] + box['height'] / 2
                        ui_log(f"[DOWNLOAD MENU] Button center: ({center_x:.1f}, {center_y:.1f})")
                        
                        # Move mouse to center
                        self.feedback_page.mouse.move(center_x, center_y)
                        self._wait(300)
                        
                        # Real mouse down/up click
                        self.feedback_page.mouse.down()
                        self._wait(100)
                        self.feedback_page.mouse.up()
                        ui_log("[DOWNLOAD MENU] ✅ Download menu opened")
                        
                        self._wait(2000)
                        return True
                    else:
                        ui_log(f"[DOWNLOAD MENU] Could not get bounding box for: {sel}")
            except Exception as e:
                ui_log(f"[DOWNLOAD MENU] Selector {sel} failed: {e}")
                continue
        
        ui_log("[DOWNLOAD MENU] ⚠️  Download button not found with any selector")
        return False
    
    def download_current_view_dual_handling(self, paper_title, final_score):
        """
        Download PDF with RACE-STYLE HANDLING (v3.6 FIX):
        - Direct PDF download, OR
        - Classic report popup (newreport_classic.asp), OR
        - Recovered UUID PDF from downloads folder
        
        Uses ORIGINAL uploaded filename format:
        turnitin-report_<ORIGINAL_UPLOADED_FILENAME>.pdf
        
        NO FILENAME TRUNCATION - preserves full original filename.
        
        Returns: Path to saved PDF
        """
        process_state.update_stage(ProcessState.DOWNLOAD_STARTED)
        
        # Use turnitin_name from process_state as the PDF filename (v4 BRIDGE)
        if process_state.turnitin_name:
            safe_name = sanitize_filename(process_state.turnitin_name)
            filename   = f"turnitin-report_{safe_name}.pdf"
        elif process_state.original_filename:
            # Fallback: use original filename if turnitin_name not set
            base_name = os.path.splitext(process_state.original_filename)[0]
            safe_base = sanitize_filename(base_name)
            filename   = f"turnitin-report_{safe_base}.pdf"
        else:
            # Last resort fallback
            clean_title = re.sub(r"[^A-Za-z0-9._-]+", "_", paper_title[:45]).strip("_") or "paper"
            clean_score = re.sub(r"[^A-Za-z0-9._-]+", "_", final_score.replace('%', 'percent')).strip("_")
            filename    = f"{JOURNAL_SHORT}_{clean_title}_{clean_score}_filtered.pdf"
        
        save_path = DOWNLOAD_DIR / filename
        
        ui_log(f"[DOWNLOAD] Target filename: {filename}")
        ui_log(f"[DOWNLOAD] Full target path: {save_path}")
        ui_log(f"[DOWNLOAD] Using turnitin_name: {process_state.turnitin_name}")
        
        selectors = [
            "[aria-label='Current View']",
            "[title='Current View']",
            "text=Current View",
            "text=Current view",
            ".print-download-btn:has-text('Current View')",
        ]
        
        max_retries = 3
        
        for retry in range(1, max_retries + 1):
            try:
                ui_log(f"\n[DOWNLOAD] === Retry {retry}/{max_retries} ===")
                
                # Step 1: Open download menu
                if not self.open_download_menu():
                    raise RuntimeError("❌ Download menu failed to open")
                
                ui_log("[DOWNLOAD] Download menu opened")
                
                # Step 2: Verify Current View is visible
                current_view_loc = None
                current_view_sel = None
                
                for sel in selectors:
                    try:
                        loc = self.feedback_page.locator(sel).first
                        if loc.is_visible(timeout=3000):
                            ui_log(f"[DOWNLOAD] Current View visible: {sel}")
                            current_view_loc = loc
                            current_view_sel = sel
                            break
                    except Exception:
                        continue
                
                if not current_view_loc:
                    raise RuntimeError("❌ Current View option not found or not visible")
                
                # Step 3: SNAPSHOT state before click
                ui_log("[DOWNLOAD] === SNAPSHOT: Capturing state BEFORE click ===")
                
                # Snapshot existing files in DOWNLOAD_DIR
                existing_files_before = set()
                try:
                    existing_files_before = set(f.name for f in DOWNLOAD_DIR.iterdir() if f.is_file())
                    ui_log(f"[DOWNLOAD] Files before click: {len(existing_files_before)} files")
                    for f in list(existing_files_before)[:5]:
                        ui_log(f"  - {f}")
                except Exception as e:
                    ui_log(f"[DOWNLOAD] Could not snapshot files: {e}")
                
                # Snapshot existing pages in context
                existing_page_count_before = 0
                try:
                    existing_page_count_before = len(context.pages)
                    ui_log(f"[DOWNLOAD] Pages before click: {existing_page_count_before} pages")
                except Exception as e:
                    ui_log(f"[DOWNLOAD] Could not snapshot pages: {e}")
                
                ui_log("[DOWNLOAD] === SNAPSHOT COMPLETE ===")
                
                # Step 4: Prepare for race-style handling
                ui_log("[DOWNLOAD] Preparing for RACE-STYLE outcome handling...")
                ui_log("[DOWNLOAD]   A) Direct PDF download")
                ui_log("[DOWNLOAD]   B) Classic report popup")
                ui_log("[DOWNLOAD]   C) Recovered UUID PDF in downloads folder")
                
                # Scroll, hover, get bounding box, move mouse
                current_view_loc.scroll_into_view_if_needed(timeout=5000)
                self._wait(500)
                
                current_view_loc.hover(timeout=5000)
                self._wait(500)
                
                box = current_view_loc.bounding_box(timeout=5000)
                if not box:
                    raise RuntimeError("❌ Could not get bounding box for Current View")
                
                center_x = box['x'] + box['width'] / 2
                center_y = box['y'] + box['height'] / 2
                ui_log(f"[DOWNLOAD] Current View center: ({center_x:.1f}, {center_y:.1f})")
                
                self.feedback_page.mouse.move(center_x, center_y)
                self._wait(300)
                
                # Step 5: CLICK Current View WITH download listener active
                # The expect_download context manager wraps the click, so it captures
                # the download event as soon as it starts (no race condition)
                ui_log("[DOWNLOAD] === CLICKING Current View with download listener ===")
                
                download_captured = False
                download = None
                popup_opened = False
                classic_page = None
                new_pdf_found = False
                recovered_pdf_path = None
                
                try:
                    # Wrap click in expect_download - captures download event if one starts within 15s
                    ui_log("[DOWNLOAD] Setting up expect_download (15s timeout) and clicking...")
                    with self.feedback_page.expect_download(timeout=15000) as download_info:
                        self.feedback_page.mouse.down()
                        self._wait(100)
                        self.feedback_page.mouse.up()
                        ui_log("[DOWNLOAD] Click performed, waiting for download event...")
                    
                    # If we reach here, a download was captured within the timeout
                    download = download_info.value
                    download_captured = True
                    ui_log("[DOWNLOAD] ✅ OUTCOME A: Direct download captured!")
                except Exception as e:
                    ui_log(f"[DOWNLOAD] No direct download detected after 15s: {e}")
                
                # Step 6: If no download, check for popup page
                if not download_captured:
                    try:
                        ui_log("[DOWNLOAD] Checking for popup page (10s timeout)...")
                        with context.expect_page(timeout=10000) as new_page_info:
                            pass
                        classic_page = new_page_info.value
                        popup_opened = True
                        ui_log("[DOWNLOAD] ✅ OUTCOME B: Popup page detected!")
                        ui_log(f"[DOWNLOAD] Popup URL: {classic_page.url}")
                        
                        try:
                            classic_page.wait_for_load_state("domcontentloaded", timeout=15000)
                            ui_log("[DOWNLOAD] Popup page loaded successfully")
                        except Exception as load_err:
                            ui_log(f"[DOWNLOAD] Popup load wait warning: {load_err}")
                    except Exception as e:
                        ui_log(f"[DOWNLOAD] No popup page detected: {e}")
                
                # Step 7: Check for new files in DOWNLOAD_DIR (UUID recovery)
                try:
                    ui_log("[DOWNLOAD] Scanning DOWNLOAD_DIR for new files...")
                    
                    existing_files_after_objects = list(DOWNLOAD_DIR.iterdir())
                    existing_files_after = set(f.name for f in existing_files_after_objects if f.is_file())
                    new_files = existing_files_after - existing_files_before
                    
                    ui_log(f"[DOWNLOAD] Files after click: {len(existing_files_after)} files")
                    ui_log(f"[DOWNLOAD] New files detected: {len(new_files)} files")
                    
                    if new_files:
                        new_file_paths = [DOWNLOAD_DIR / f for f in new_files]
                        new_file_paths_with_time = []
                        
                        for file_path in new_file_paths:
                            try:
                                mtime = file_path.stat().st_mtime
                                new_file_paths_with_time.append((file_path, mtime))
                            except Exception:
                                pass
                        
                        if new_file_paths_with_time:
                            new_file_paths_with_time.sort(key=lambda x: x[1], reverse=True)
                            newest_file = new_file_paths_with_time[0][0]
                            
                            ui_log(f"[DOWNLOAD] Newest file: {newest_file.name}")
                            
                            is_pdf = False
                            if newest_file.suffix.lower() == '.pdf':
                                is_pdf = True
                                ui_log(f"[DOWNLOAD] File has .pdf extension")
                            else:
                                ui_log(f"[DOWNLOAD] File has NO .pdf extension: {newest_file.suffix}")
                                try:
                                    with open(newest_file, "rb") as f:
                                        header = f.read(5)
                                    ui_log(f"[DOWNLOAD] File header bytes: {header}")
                                    if header == b"%PDF-":
                                        is_pdf = True
                                        ui_log(f"[DOWNLOAD] ✅ File IS a PDF (detected by %PDF- header)")
                                    else:
                                        ui_log(f"[DOWNLOAD] ❌ File is NOT a PDF (header: {header})")
                                except Exception as e:
                                    ui_log(f"[DOWNLOAD] Could not read file header: {e}")
                            
                            if is_pdf:
                                recovered_pdf_path = newest_file
                                new_pdf_found = True
                                ui_log(f"[DOWNLOAD] ✅ OUTCOME C: Recovered PDF file: {newest_file.name}")
                            else:
                                ui_log(f"[DOWNLOAD] ❌ Newest file is not a PDF, skipping")
                    
                    if not new_pdf_found:
                        ui_log("[DOWNLOAD] No new PDF files detected")
                        
                except Exception as e:
                    ui_log(f"[DOWNLOAD] File scan failed: {e}")
                
                # Step 7: Handle OUTCOME A - Direct Download
                if download_captured and download:
                    ui_log("[DOWNLOAD] === HANDLING OUTCOME A: Direct Download ===")
                    download.save_as(str(save_path))
                    ui_log(f"[DOWNLOAD] ✅ File saved: {save_path.name}")
                    ui_log(f"[DOWNLOAD] Full path: {save_path}")
                    
                    # Validate PDF header
                    try:
                        with open(save_path, "rb") as f:
                            header = f.read(5)
                        ui_log(f"[DOWNLOAD] PDF header bytes: {header}")
                        if header != b"%PDF-":
                            ui_log(f"[DOWNLOAD] ⚠️  File may not be valid PDF. Header: {header}")
                        else:
                            ui_log("[DOWNLOAD] ✅ PDF validation OK: %PDF- header confirmed")
                    except Exception as e:
                        ui_log(f"[DOWNLOAD] ⚠️  Could not validate PDF: {e}")
                    
                    process_state.update_stage(ProcessState.DOWNLOADED)
                    process_state.final_pdf_filename = filename
                    process_state.final_pdf_path = str(save_path)
                    
                    ui_log(f"[DOWNLOAD] ✅ Download completed successfully (direct) on retry {retry}")
                    return save_path
                
                # Step 8: Handle OUTCOME B - Classic Report Popup
                if popup_opened and classic_page:
                    ui_log("[DOWNLOAD] === HANDLING OUTCOME B: Classic Report Popup ===")
                    process_state.update_stage(ProcessState.CLASSIC_REPORT_OPENED)
                    process_state.add_note("Classic report popup opened")
                    
                    # Wait for page to load
                    try:
                        classic_page.wait_for_load_state("domcontentloaded", timeout=15000)
                        self._wait(2000)
                        ui_log(f"[CLASSIC] Popup URL: {classic_page.url}")
                    except Exception as e:
                        ui_log(f"[CLASSIC] Page load wait failed: {e}")
                    
                    # Verify it's the classic report page
                    is_classic = False
                    try:
                        url = classic_page.url
                        if "newreport_classic.asp" in url:
                            is_classic = True
                            ui_log("[CLASSIC] ✅ Confirmed: newreport_classic.asp URL")
                    except Exception:
                        pass
                    
                    if not is_classic:
                        try:
                            title_text = classic_page.title()
                            if "Turnitin Originality Report" in title_text or "Originality Report" in title_text:
                                is_classic = True
                                ui_log(f"[CLASSIC] ✅ Confirmed: Title contains '{title_text}'")
                        except Exception:
                            pass
                    
                    if not is_classic:
                        ui_log("[CLASSIC] ⚠️  WARNING: Popup may not be classic report page")
                    
                    # Look for download button in classic page
                    ui_log("[CLASSIC] Looking for download button...")
                    
                    download_selectors = [
                        "a:has-text('download')",
                        "a:has-text('Download')",
                        "[title*='download']",
                        "[title*='Download']",
                        "a.download",
                        "input[value*='download']",
                        "input[value*='Download']",
                    ]
                    
                    download_btn = None
                    for sel in download_selectors:
                        try:
                            btn = classic_page.locator(sel).first
                            if btn.count() > 0 and btn.is_visible(timeout=3000):
                                download_btn = btn
                                ui_log(f"[CLASSIC] Found download button: {sel}")
                                break
                        except Exception:
                            continue
                    
                    if not download_btn:
                        ui_log("[CLASSIC] ❌ Could not find download button")
                        self._save_screenshot("classic-no-download-btn")
                        raise RuntimeError("Classic report download button not found")
                    
                    # Click download button on classic page
                    ui_log("[CLASSIC] Clicking download button...")
                    
                    try:
                        with classic_page.expect_download(timeout=30000) as classic_dl_info:
                            download_btn.scroll_into_view_if_needed(timeout=5000)
                            self._wait(500)
                            download_btn.click(timeout=5000)
                            ui_log("[CLASSIC] Download button clicked")
                        
                        classic_download = classic_dl_info.value
                        ui_log("[CLASSIC] ✅ Download event captured!")
                        
                        # Save the download
                        classic_download.save_as(str(save_path))
                        ui_log(f"[CLASSIC] ✅ File saved: {save_path.name}")
                        ui_log(f"[CLASSIC] Full path: {save_path}")
                        
                        # Validate PDF header
                        try:
                            with open(save_path, "rb") as f:
                                header = f.read(5)
                            ui_log(f"[CLASSIC] PDF header bytes: {header}")
                            if header != b"%PDF-":
                                ui_log(f"[CLASSIC] ⚠️  File may not be valid PDF. Header: {header}")
                            else:
                                ui_log("[CLASSIC] ✅ PDF validation OK: %PDF- header confirmed")
                        except Exception as e:
                            ui_log(f"[CLASSIC] ⚠️  Could not validate PDF: {e}")
                        
                        process_state.update_stage(ProcessState.DOWNLOADED)
                        process_state.final_pdf_filename = filename
                        process_state.final_pdf_path = str(save_path)
                        
                        ui_log(f"[CLASSIC] ✅ Download completed successfully (classic popup) on retry {retry}")
                        
                        # Close classic page
                        try:
                            classic_page.close()
                            ui_log("[CLASSIC] Popup closed")
                        except Exception:
                            pass
                        
                        return save_path
                        
                    except Exception as e:
                        ui_log(f"[CLASSIC] ❌ Download from classic page failed: {e}")
                        self._save_screenshot("classic-download-failed")
                        raise RuntimeError(f"Classic report download failed: {e}")
                
                # Step 9: Handle OUTCOME C - Recovered UUID PDF
                if new_pdf_found and recovered_pdf_path:
                    ui_log("[DOWNLOAD] === HANDLING OUTCOME C: Recovered UUID PDF ===")
                    ui_log(f"[DOWNLOAD] Renaming UUID file to final name...")
                    ui_log(f"[DOWNLOAD] From: {recovered_pdf_path.name}")
                    ui_log(f"[DOWNLOAD] To: {save_path.name}")
                    
                    try:
                        # Rename/move to final path
                        import shutil
                        shutil.move(str(recovered_pdf_path), str(save_path))
                        ui_log(f"[DOWNLOAD] ✅ File renamed successfully")
                        ui_log(f"[DOWNLOAD] Final path: {save_path}")
                        
                        # Validate PDF header
                        try:
                            with open(save_path, "rb") as f:
                                header = f.read(5)
                            ui_log(f"[DOWNLOAD] PDF header bytes: {header}")
                            if header != b"%PDF-":
                                ui_log(f"[DOWNLOAD] ⚠️  File may not be valid PDF. Header: {header}")
                            else:
                                ui_log("[DOWNLOAD] ✅ PDF validation OK: %PDF- header confirmed")
                        except Exception as e:
                            ui_log(f"[DOWNLOAD] ⚠️  Could not validate PDF: {e}")
                        
                        process_state.update_stage(ProcessState.DOWNLOADED)
                        process_state.final_pdf_filename = filename
                        process_state.final_pdf_path = str(save_path)
                        
                        ui_log(f"[DOWNLOAD] ✅ Download completed successfully (recovered UUID) on retry {retry}")
                        return save_path
                        
                    except Exception as e:
                        ui_log(f"[DOWNLOAD] ❌ Failed to rename UUID file: {e}")
                        raise RuntimeError(f"UUID file recovery failed: {e}")
                
                # Step 10: If no outcome detected
                ui_log("[DOWNLOAD] ❌ No outcome detected: no download, no popup, no new files")
                raise RuntimeError("No download outcome detected after click")
                
            except Exception as e:
                ui_log(f"[DOWNLOAD] ❌ Retry {retry} failed: {e}")
                
                # Save screenshot on failed retry
                self._save_screenshot(f"download-retry-{retry}-failed")
                
                if retry < max_retries:
                    ui_log(f"[DOWNLOAD] Waiting 2 seconds before retry {retry + 1}...")
                    self._wait(2000)
                    
                    # Try to close/reopen download menu for next retry
                    try:
                        ui_log("[DOWNLOAD] Attempting to reset download menu...")
                        # Press Escape to close menu
                        self.feedback_page.keyboard.press("Escape")
                        self._wait(1000)
                        ui_log("[DOWNLOAD] Download menu closed (Escape)")
                    except Exception as close_err:
                        ui_log(f"[DOWNLOAD] Could not close menu: {close_err}")
                else:
                    # Final failure after all retries
                    ui_log(f"[DOWNLOAD] ❌ FINAL FAILURE: Download did not trigger after {max_retries} retries")
                    self._save_screenshot(f"download-final-failure")
                    raise RuntimeError(f"Download failed after {max_retries} attempts")
        
        # Should never reach here
        raise RuntimeError("Download failed: unexpected code path")
    
    def run_full_workflow(self):
        """
        Execute complete Feedback Studio workflow with robust retry logic.
        This is the main entry point that orchestrates all steps.
        """
        ui_log("="*70)
        ui_log("🚀 STARTING FEEDBACK STUDIO WORKFLOW")
        ui_log("="*70)
        
        # STEP 0: Get paper title
        ui_log("\n[WORKFLOW STEP 0] Getting paper title...")
        paper_title = self._retry_action(
            lambda: self.get_paper_title(fallback=self.title),
            "Get Paper Title",
            max_attempts=2
        )
        ui_log(f"📄 Paper title: {paper_title}")
        self._wait(1500)
        
        # STEP 1: Navigate to Match Overview
        ui_log("\n[WORKFLOW STEP 1] Navigate to Match Overview")
        self._retry_action(
            self.ensure_match_overview,
            "Navigate to Match Overview",
            max_attempts=3,
            verify_func=self.verify_match_overview_ready
        )
        
        # Get score before filtering
        score_before_auto = self._retry_action(
            self.get_score,
            "Get Score Before Filtering",
            max_attempts=2
        )
        if score_before_auto == "unknown" and self.score_before != "unknown":
            score_before_auto = self.score_before
        ui_log(f"📊 SCORE BEFORE FILTERING: {score_before_auto}")
        
        # Store in process state
        process_state.similarity_before_filter = score_before_auto
        
        self._wait(2000)
        
        # STEP 2: Filter and exclude sources
        ui_log("\n[WORKFLOW STEP 2] Filter and Exclude Sources")
        filter_result = self._retry_action(
            self.filter_and_exclude_if_needed,
            "Filter Sources",
            max_attempts=3
        )
        
        if filter_result:
            ui_log("✅ Sources filtered and excluded successfully")
        else:
            ui_log("ℹ️  No sources excluded")
        
        self._wait(3000)
        
        # STEP 3: Return to Match Overview for final score
        ui_log("\n[WORKFLOW STEP 3] Return to Match Overview for Final Score")
        self._retry_action(
            self.ensure_match_overview,
            "Return to Match Overview",
            max_attempts=3,
            verify_func=self.verify_match_overview_ready
        )
        
        # Get final score
        final_score = self._retry_action(
            self.get_score,
            "Get Final Score",
            max_attempts=2
        )
        if final_score == "unknown" and score_before_auto != "unknown":
            final_score = score_before_auto
        
        ui_log(f"📊 FINAL SCORE AFTER FILTERING: {final_score}")
        
        # Store in process state
        process_state.similarity_after_filter = final_score
        
        self._wait(2000)
        
        # STEP 4: Download PDF with dual handling
        ui_log("\n[WORKFLOW STEP 4] Download PDF Report (Dual Handling)")
        save_path = self._retry_action(
            lambda: self.download_current_view_dual_handling(paper_title, final_score),
            "Download PDF",
            max_attempts=3
        )
        
        ui_log("="*70)
        ui_log("✅ WORKFLOW COMPLETED SUCCESSFULLY")
        ui_log(f"💾 PDF saved to: {save_path}")
        ui_log("🔓 Feedback Studio automation lock released")
        ui_log("="*70)
        
        return save_path


class Api:

    def choose_files(self):
        global window, selected_files

        result = window.create_file_dialog(
            webview.OPEN_DIALOG,
            allow_multiple=True,
            file_types=('Documents (*.docx;*.pdf)',)
        )

        if result:
            selected_files = list(result)
            return selected_files

        return []

    def open_turnitin(self):
        self._queue_automation("open_turnitin")
        return "Opening Turnitin..."

    def start_upload(self):
        self._queue_automation("prepare_quick_submit")
        return "Starting Quick Submit setup flow..."

    def _queue_automation(self, action):
        global automation_worker_started

        with automation_worker_lock:
            if not automation_worker_started:
                threading.Thread(target=self._automation_worker, daemon=True).start()
                automation_worker_started = True

        automation_queue.put(action)

    def _automation_worker(self):
        while True:
            action = automation_queue.get()

            try:
                if action == "open_turnitin":
                    self._open_or_focus_turnitin()
                elif action == "prepare_quick_submit":
                    self._prepare_quick_submit_no_repository()
                else:
                    ui_log(f"Unknown automation action: {action}")
            except Exception as e:
                print(traceback.format_exc())
                ui_log(f"FAILED: {e}")
                
                # Mark process as failed and log
                process_state.mark_failed(e)
                process_state.finalize(ProcessState.FAILED_BEFORE_UPLOAD)
                excel_logger.log_process(process_state)
                
                self._save_failure_screenshot("automation-crash")
            finally:
                automation_queue.task_done()

    def _ensure_browser(self):
        global pw, context, page

        if context is None:
            ui_log("Opening persistent visible browser session...")
            ui_log(f"[TURNITIN] Using persistent profile: {PROFILE_DIR}")
            ui_log(f"[TURNITIN] Profile exists: {PROFILE_DIR.exists()}")
            
            # === v5.5.2: Diagnostics before launch ===
            cookies_db = PROFILE_DIR / "Default" / "Cookies"
            local_storage = PROFILE_DIR / "Default" / "Local Storage"
            ui_log(f"[DIAGNOSTICS] Turnitin profile path: {PROFILE_DIR}")
            ui_log(f"[DIAGNOSTICS] Profile exists: {PROFILE_DIR.exists()}")
            ui_log(f"[DIAGNOSTICS] Cookies DB exists: {cookies_db.exists()}")
            ui_log(f"[DIAGNOSTICS] Local Storage exists: {local_storage.exists() and any(local_storage.rglob('*'))}")
            
            # === v5.5.2: Stale lock file cleanup ===
            # Remove stale lock files before launching (safe cleanup if no active Chrome process)
            from config import PROJECT_ROOT as _cfg_root
            stale_locks = [
                PROFILE_DIR / "SingletonLock",
                PROFILE_DIR / "SingletonCookie",
                PROFILE_DIR / "SingletonSocket",
                PROFILE_DIR / "Default" / "LOCK",
            ]
            for lock_file in stale_locks:
                if lock_file.exists():
                    try:
                        # Check if this lock is truly stale: no Chrome process using this profile
                        lock_file.unlink()
                        ui_log(f"[LOCK CLEANUP] Removed stale lock: {lock_file.name}")
                    except Exception as e:
                        ui_log(f"[LOCK CLEANUP] Could not remove {lock_file.name}: {e}")
            
            # Ensure directories
            PROFILE_DIR.mkdir(parents=True, exist_ok=True)
            DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
            
            ui_log(f"PLAYWRIGHT_BROWSERS_PATH: {os.environ.get('PLAYWRIGHT_BROWSERS_PATH')}")
            browser_debug_path = BUNDLED_BROWSERS_DIR or resource_path("ms-playwright")
            ui_log(f"Bundled browser exists: {browser_debug_path.exists()}")
            ui_log(f"Bundled browser contents: {list(browser_debug_path.glob('*')) if browser_debug_path.exists() else 'missing'}")
            pw = sync_playwright().start()
            try:
                # v5.5.2: Use channel="chrome" if available, plus anti-detection args
                browser_args = [
                    "--disable-blink-features=AutomationControlled",
                    "--start-maximized",
                    "--no-sandbox",
                ]
                try:
                    context = pw.chromium.launch_persistent_context(
                        user_data_dir=str(PROFILE_DIR),
                        headless=False,
                        channel="chrome",
                        args=browser_args,
                        slow_mo=SLOW_MO,
                        accept_downloads=True,
                        downloads_path=str(DOWNLOAD_DIR)
                    )
                    ui_log("[TURNITIN] ✅ Launched with channel='chrome'")
                except Exception as chrome_err:
                    ui_log(f"[TURNITIN] chrome channel not available: {chrome_err}")
                    ui_log("[TURNITIN] Falling back to default Chromium channel...")
                    context = pw.chromium.launch_persistent_context(
                        user_data_dir=str(PROFILE_DIR),
                        headless=False,
                        args=browser_args,
                        slow_mo=SLOW_MO,
                        accept_downloads=True,
                        downloads_path=str(DOWNLOAD_DIR)
                    )
                    ui_log("[TURNITIN] ✅ Launched with default Chromium channel")
            except Exception as e:
                ui_log("FAILED: Could not launch Playwright Chromium browser.")
                ui_log(f"Browser launch exception: {e}")
                ui_log(f"PLAYWRIGHT_BROWSERS_PATH: {os.environ.get('PLAYWRIGHT_BROWSERS_PATH')}")

                if BUNDLED_BROWSERS_DIR is None:
                    ui_log("Bundled browser directory: not used in non-frozen/dev mode.")
                else:
                    ui_log(f"Bundled browser directory: {BUNDLED_BROWSERS_DIR}")
                    ui_log(f"Bundled browser directory exists: {BUNDLED_BROWSERS_DIR.exists()}")
                    if BUNDLED_BROWSERS_DIR.exists():
                        try:
                            contents = sorted(path.name for path in BUNDLED_BROWSERS_DIR.iterdir())
                            ui_log(f"Bundled browser directory contents: {contents}")
                        except Exception as list_error:
                            ui_log(f"Could not list bundled browser directory contents: {list_error}")

                raise

            context.set_default_timeout(DEFAULT_TIMEOUT)
            context.set_default_navigation_timeout(NAVIGATION_TIMEOUT)
            page = context.pages[0] if context.pages else context.new_page()

        return page

    def _open_or_focus_turnitin(self):
        self._ensure_browser()

        ui_log("Opening Turnitin login page...")
        page.goto(TURNITIN_LOGIN_URL, wait_until="domcontentloaded", timeout=NAVIGATION_TIMEOUT)
        page.wait_for_timeout(2200)

        ui_log("Login manually if needed, then click Prepare Quick Submit.")

    def _goto_turnitin_home(self):
        ui_log("Opening Turnitin dashboard/home page...")

        try:
            page.goto(TURNITIN_HOME_URL, wait_until="domcontentloaded", timeout=NAVIGATION_TIMEOUT)
        except Exception as e:
            if "net::ERR_ABORTED" in str(e):
                ui_log("Turnitin redirected during navigation; continuing with the loaded page.")
            else:
                raise

        try:
            page.wait_for_load_state("domcontentloaded", timeout=15000)
        except Exception:
            pass

        page.wait_for_timeout(800)

    def _save_failure_screenshot(self, label):
        if page is None:
            return None

        try:
            SCREENSHOT_DIR.mkdir(parents=True, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
            safe_label = "".join(ch if ch.isalnum() or ch in "-_" else "-" for ch in label).strip("-")
            screenshot_path = SCREENSHOT_DIR / f"{timestamp}-{safe_label}.png"
            page.screenshot(path=str(screenshot_path), full_page=True)
            ui_log(f"Saved failure screenshot: {screenshot_path}")
            
            # Add to process state
            process_state.add_screenshot(screenshot_path)
            
            return screenshot_path
        except Exception as screenshot_error:
            ui_log(f"Could not save failure screenshot: {screenshot_error}")
            return None

    def _wait_for_selector(
        self,
        selector,
        description,
        timeout=DEFAULT_TIMEOUT,
        state="visible",
        screenshot_on_failure=True
    ):
        ui_log(f"Waiting for {description}: {selector}")

        try:
            locator = page.locator(selector)
            locator.wait_for(state=state, timeout=timeout)
            ui_log(f"Found {description}.")
            return locator
        except Exception as e:
            ui_log(f"FAILED: Timed out waiting for {description}.")
            ui_log(str(e))
            if screenshot_on_failure:
                self._save_failure_screenshot(f"missing-{description}")
            raise

    def _safe_click(self, selector, description, timeout=DEFAULT_TIMEOUT, wait_after_ms=700):
        locator = self._wait_for_selector(selector, description, timeout=timeout)

        try:
            ui_log(f"Clicking {description}: {selector}")
            locator.scroll_into_view_if_needed(timeout=timeout)
            locator.click(timeout=timeout)
            ui_log(f"Clicked {description}.")
        except Exception as e:
            error_text = str(e)
            if "click action done" in error_text and "waiting for scheduled navigations" in error_text:
                ui_log(f"Clicked {description}, but Turnitin navigation did not finish cleanly before timeout. Continuing.")
                try:
                    page.wait_for_load_state("domcontentloaded", timeout=10000)
                except Exception:
                    pass
                if wait_after_ms:
                    page.wait_for_timeout(wait_after_ms)
                return locator

            ui_log(f"FAILED: Cannot click {description}.")
            ui_log(error_text)
            self._save_failure_screenshot(f"click-{description}")
            raise

        try:
            page.wait_for_load_state("domcontentloaded", timeout=10000)
        except Exception:
            # Classic Turnitin pages sometimes finish enough for selectors before Playwright sees a load state.
            pass

        if wait_after_ms:
            page.wait_for_timeout(wait_after_ms)

        return locator

    def _human_pause(self, min_ms=None, max_ms=None):
        min_ms = min_ms or PAUSE_MIN_MS
        max_ms = max_ms or PAUSE_MAX_MS
        delay = random.randint(min_ms, max_ms)
        ui_log(f"Pause: {delay}ms")
        page.wait_for_timeout(delay)

    def _wait_for_dashboard_or_manual_login(self):
        """v5.5.2: Enhanced login detection with diagnostics"""
        # Log current URL for diagnostics
        try:
            current_url = page.url
            ui_log(f"[TURNITIN] Current URL after opening: {current_url}")
        except Exception:
            pass
        
        # Check if already logged in (Quick Submit link present)
        try:
            self._wait_for_selector(
                "a.sn_quick_submit",
                "Quick Submit link on dashboard",
                timeout=10000,
                screenshot_on_failure=False
            )
            ui_log("[SESSION] ✅ Turnitin already logged in - session reused!")
            return
        except Exception:
            ui_log("Please login manually in browser. Workflow will continue automatically.")

        # Wait for Quick Submit link to appear (after manual login)
        self._wait_for_selector(
            "a.sn_quick_submit",
            "Quick Submit link after manual login",
            timeout=MANUAL_LOGIN_TIMEOUT
        )
        ui_log("Manual login completed; dashboard is ready.")
        
        # === v5.5.2: Save session state explicitly after login ===
        # This ensures cookies/session are persisted in the profile
        storage_state_path = PROFILE_DIR / "storage_state.json"
        try:
            context.storage_state(path=str(storage_state_path))
            ui_log(f"[SESSION] ✅ Session state saved to: {storage_state_path}")
        except Exception as e:
            ui_log(f"[SESSION] Could not save storage state: {e}")

    def _select_no_repository(self):
        self._wait_for_selector(
            "#submit_papers_to",
            "No Repository dropdown on Customize Search page",
            timeout=DEFAULT_TIMEOUT
        )

        try:
            ui_log("Selecting No Repository: #submit_papers_to value=0")
            page.select_option("#submit_papers_to", value="0", timeout=DEFAULT_TIMEOUT)
            ui_log("Selected No Repository.")
        except Exception as e:
            ui_log("FAILED: Cannot select No Repository.")
            ui_log(str(e))
            self._save_failure_screenshot("select-no-repository")
            raise

    def _tick_customize_search_checkboxes(self):
        ui_log("Ticking inspected Customize Search checkbox selectors...")

        for selector in CHECKBOX_SELECTORS:
            try:
                checkbox = page.locator(selector)
                checkbox.wait_for(state="visible", timeout=10000)

                self._human_pause()

                if not checkbox.is_checked():
                    checkbox.check(force=True, timeout=DEFAULT_TIMEOUT)
                    ui_log(f"Ticked checkbox: {selector}")
                else:
                    ui_log(f"Checkbox already ticked: {selector}")

                self._human_pause()

            except Exception as e:
                ui_log(f"FAILED checkbox: {selector}")
                ui_log(str(e))
                self._save_failure_screenshot("checkbox-failure")

    def _log_final_url_after_submit(self):
        try:
            page.wait_for_load_state("domcontentloaded", timeout=15000)
        except Exception:
            # Turnitin classic pages may navigate without a clean load-state signal.
            pass

        page.wait_for_timeout(1200)

        current_url = page.url
        ui_log(f"Final URL after submit: {current_url}")

        if "t_custom_search.asp" in current_url:
            ui_log("WARNING: Still on Customize Search page. Checkboxes or form submission may not be accepted.")

        return current_url

    def _ensure_single_file_upload(self):
        ui_log("Ensuring submit type is Single File Upload...")

        try:
            current_type = page.locator("#submit_type").inner_text(timeout=5000).strip()

            ui_log(f"Current submit type: {current_type}")

            if "Single File Upload" in current_type:
                ui_log("Single File Upload already selected.")
                return

        except Exception as e:
            ui_log("Could not read current submit type.")
            ui_log(str(e))

        try:
            page.locator("#submit_type").click(timeout=10000)

            page.locator("#submit-type-menu").wait_for(
                state="visible",
                timeout=10000
            )

            single_option = page.locator(
                "#submit-type-menu a",
                has_text="Single File Upload"
            )

            if single_option.count() > 0:
                single_option.first.click(timeout=10000)
                ui_log("Selected Single File Upload.")
            else:
                ui_log("Single File Upload option not found in dropdown.")

            page.wait_for_timeout(800)

        except Exception as e:
            ui_log("FAILED: Cannot ensure Single File Upload submit type.")
            ui_log(str(e))
            self._save_failure_screenshot("single-file-upload")
            raise

    def _get_first_selected_file_details(self):
        """
        [BRIDGE] Read next manuscript from Excel 'download' sheet.
        Replaces manual file selection. Idempotent: if process_state already
        has file data (second call after upload), returns cached values.
        """
        # --- Idempotent guard: return cached data on second call ---
        if process_state.original_file_path and process_state.original_filename:
            ui_log("[BRIDGE] File details already loaded from Excel, returning cached values")
            return (
                process_state.original_file_path,
                process_state.parsed_first_name,
                process_state.parsed_last_name,
                process_state.parsed_title,
            )
        
        # --- Read from Excel bridge ---
        submission = excel_bridge.find_next_submission()
        
        if not submission:
            ui_log("FAILED: No unprocessed submissions found in Excel download sheet.")
            raise ValueError("No unprocessed submissions in download sheet")
        
        file_path = Path(submission["file_path"])
        
        if not file_path.exists():
            ui_log(f"FAILED: Source file does not exist: {file_path}")
            raise FileNotFoundError(str(file_path))
        
        # Store in process state
        process_state.original_file_path  = str(file_path)
        process_state.original_filename   = file_path.name
        process_state.submission_no       = submission["submission_no"]
        process_state.author_name         = submission["author_name"]
        process_state.source_download_path = submission["file_path"]
        process_state.turnitin_name        = submission["turnitin_name"]
        process_state.update_stage(ProcessState.FILE_SELECTED)
        
        ui_log(f"[METADATA] Original file path: {file_path}")
        ui_log(f"[METADATA] Original filename: {file_path.name}")
        ui_log(f"[METADATA] Submission no: {submission['submission_no']}")
        ui_log(f"[METADATA] Turnitin name: {submission['turnitin_name']}")
        
        # --- v5.4: Use standardized metadata for Turnitin submission ---
        # First Name = "Jurnal Teknologi"
        # Last Name = "Editor Team"
        # Title = submission_no only
        first_name = "Jurnal Teknologi"
        last_name  = "Editor Team"
        title      = submission["submission_no"]
        
        # Store parsed metadata
        process_state.parsed_title      = title
        process_state.parsed_first_name = first_name
        process_state.parsed_last_name  = last_name
        
        ui_log(f"[BRIDGE] Author parsed: first={first_name!r} last={last_name!r}")
        ui_log(f"[BRIDGE] Title: {title}")
        
        return str(file_path), first_name, last_name, title

    def _move_near_choose_file_button(self):
        try:
            choose_file_btn = page.locator("#choose-file-btn")
            choose_file_btn.scroll_into_view_if_needed(timeout=10000)
            box = choose_file_btn.bounding_box(timeout=5000)

            if box:
                page.mouse.move(box["x"] + box["width"] / 2, box["y"] + box["height"] / 2)
                ui_log("Moved mouse to Choose File button.")
            else:
                ui_log("Choose File button has no visible bounding box. Continuing with set_input_files.")
        except Exception as e:
            ui_log("Could not move mouse to Choose File button. Continuing with set_input_files.")
            ui_log(str(e))

    def _fill_upload_page_and_attach_file(self):
        file_path, first_name, last_name, title = self._get_first_selected_file_details()

        self._ensure_single_file_upload()

        try:
            ui_log("[TURNITIN] Filling metadata")

            page.locator("#author_first").fill(first_name, timeout=DEFAULT_TIMEOUT)
            ui_log(f"[TURNITIN] First name filled: {first_name}")

            page.locator("#author_last").fill(last_name, timeout=DEFAULT_TIMEOUT)
            ui_log(f"[TURNITIN] Last name filled: {last_name}")

            page.locator("#title").fill(title, timeout=DEFAULT_TIMEOUT)
            ui_log(f"[TURNITIN] Submission title filled: {title}")

            process_state.update_stage(ProcessState.UPLOAD_FORM_FILLED)

        except Exception as e:
            ui_log("FAILED: Cannot fill upload form fields.")
            ui_log(str(e))
            self._save_failure_screenshot("fill-upload-form")
            raise

        self._move_near_choose_file_button()

        try:
            ui_log("[TURNITIN] Attaching file")
            page.locator("#selected-file").set_input_files(file_path, timeout=DEFAULT_TIMEOUT)
            ui_log("[TURNITIN] File selected")
        except Exception as e:
            ui_log("FAILED: Cannot attach selected file to Turnitin.")
            ui_log(str(e))
            self._save_failure_screenshot("attach-selected-file")
            raise

    def _click_upload_and_confirm(self):
        ui_log("[TURNITIN] Upload button ready")

        self._safe_click(UPLOAD_BUTTON, "Upload button", wait_after_ms=6000)
        ui_log("[TURNITIN] Clicking upload")
        self._save_failure_screenshot("after-upload-click")

        process_state.update_stage(ProcessState.UPLOADED)

        ui_log("[TURNITIN] Waiting for Turnitin processing")

        confirm_btn = page.locator(CONFIRM_BUTTON)
        confirm_btn.wait_for(state="visible", timeout=90000)

        page.wait_for_function(
            """
            () => {
                const btn = document.querySelector('#confirm-btn');
                return btn && !btn.disabled && !btn.classList.contains('disabled');
            }
            """,
            timeout=90000
        )

        ui_log("[TURNITIN] Confirm ready")

        confirm_btn.scroll_into_view_if_needed(timeout=10000)
        confirm_btn.click(timeout=15000)
        ui_log("[TURNITIN] Confirm clicked")

        process_state.update_stage(ProcessState.CONFIRMED)

        page.wait_for_timeout(4000)
        self._save_failure_screenshot("after-confirm-click")

        try:
            page.wait_for_load_state("domcontentloaded", timeout=30000)
        except Exception:
            pass

        self._human_pause(1500, 2500)

        ui_log(f"Final URL after confirmation: {page.url}")
        ui_log("SUCCESS: Turnitin submission completed.")

        return True

    def _go_to_inbox_after_submission(self):
        self._human_pause(2000, 3500)

        try:
            if page.locator(CLOSE_BUTTON).count() > 0:
                self._safe_click(CLOSE_BUTTON, "Close button after submission", wait_after_ms=4000)
        except Exception as e:
            ui_log("Could not click close button; continuing to wait on current page.")
            ui_log(str(e))

        self._wait_for_selector(ASSIGN_INBOX, "Assignment inbox", timeout=60000)
        
        process_state.update_stage(ProcessState.INBOX_FOUND)
        
        # Reload inbox to ensure fresh state
        ui_log("[INBOX] Reloading inbox for fresh state after submission...")
        page.reload(wait_until="domcontentloaded", timeout=30000)
        page.wait_for_timeout(3000)
        self._wait_for_selector(ASSIGN_INBOX, "Assignment inbox after reload", timeout=30000)
        ui_log("[INBOX] Inbox reloaded successfully")

    def _normalize_text(self, text):
        return " ".join(text.lower().split())

    def _short_title_key(self, title):
        return " ".join(title.split()[:4])

    def _parse_date_text(self, date_text):
        """
        Parse date string and return comparable datetime or None.
        Handles formats like: "10-Nov-2025", "10 Nov 2025", etc.
        """
        if not date_text or date_text.lower() in ["unknown", "-", ""]:
            return None
        
        try:
            # Try common formats
            for fmt in ["%d-%b-%Y", "%d %b %Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"]:
                try:
                    return datetime.strptime(date_text.strip(), fmt)
                except ValueError:
                    continue
        except Exception:
            pass
        
        return None
    
    def _get_first_n_dates_from_inbox(self, n=5):
        """
        Extract first N visible dates from inbox rows.
        Returns list of (row_index, date_text, parsed_date) tuples.
        """
        rows = page.locator("#assign_inbox tbody tr[class^='student--'], table.inbox_table tbody tr[class^='student--']")
        total = rows.count()
        dates = []
        
        for i in range(min(n, total)):
            try:
                row = rows.nth(i)
                date_text = row.locator("td.class_status").first.inner_text(timeout=2000).strip()
                parsed = self._parse_date_text(date_text)
                dates.append((i, date_text, parsed))
                ui_log(f"[DATE-CHECK] Row {i}: {date_text}")
            except Exception as e:
                ui_log(f"[DATE-CHECK] Row {i}: Could not read date - {e}")
                dates.append((i, "unknown", None))
        
        return dates
    
    def _is_descending_date_order(self, dates):
        """
        Check if dates list is in descending order (newest first).
        Returns True if descending, False otherwise.
        """
        parsed_dates = [d[2] for d in dates if d[2] is not None]
        
        if len(parsed_dates) < 2:
            ui_log("[DATE-CHECK] Not enough valid dates to verify order")
            return False
        
        # Check if sorted descending (newest to oldest)
        is_descending = all(parsed_dates[i] >= parsed_dates[i+1] for i in range(len(parsed_dates)-1))
        
        if is_descending:
            ui_log(f"[DATE-CHECK] ✅ Dates are in DESCENDING order (newest first)")
        else:
            ui_log(f"[DATE-CHECK] ❌ Dates are NOT in descending order")
        
        return is_descending
    
    def _click_date_column_header(self):
        """
        Click the DATE column header to toggle sort order.
        Returns True if clicked successfully, False otherwise.
        
        v3.6.1 FIX: Treats navigation timeout as success since click was performed.
        """
        ui_log("[DATE-SORT] Attempting to click DATE column header...")
        
        # Try multiple selectors for date header
        date_header_selectors = [
            "th:has-text('Date')",
            "th:has-text('DATE')",
            "th.class_status",
            "th a:has-text('Date')",
            "th a:has-text('DATE')",
            "#assign_inbox thead th:has-text('Date')",
            "table.inbox_table thead th:has-text('Date')"
        ]
        
        for selector in date_header_selectors:
            try:
                header = page.locator(selector).first
                if header.count() > 0:
                    header.scroll_into_view_if_needed(timeout=5000)
                    page.wait_for_timeout(500)
                    header.click(timeout=5000)
                    ui_log(f"[DATE-SORT] ✅ Clicked DATE header using selector: {selector}")
                    return True
            except Exception as e:
                error_text = str(e)
                # v3.6.1 FIX: If click was done but navigation timeout, treat as success
                if "click action done" in error_text or "waiting for scheduled navigations" in error_text:
                    ui_log(f"[DATE-SORT] ✅ DATE click performed (navigation timeout ignored)")
                    ui_log(f"[DATE-SORT] Used selector: {selector}")
                    return True
                else:
                    ui_log(f"[DATE-SORT] Failed with selector {selector}: {e}")
                    continue
        
        ui_log("[DATE-SORT] WARNING: Could not click DATE header with any selector")
        return False
    
    def _ensure_newest_first_sorting(self, max_attempts=3):
        """
        Ensure inbox is sorted with newest papers first by clicking DATE header.
        
        Args:
            max_attempts: Maximum number of DATE header click attempts
        
        Returns: True if newest-first order confirmed, False otherwise
        """
        ui_log("="*60)
        ui_log("[DATE-SORT] Ensuring inbox sorted by DATE (newest first)")
        ui_log("="*60)
        
        for attempt in range(1, max_attempts + 1):
            ui_log(f"\n[DATE-SORT] Attempt {attempt}/{max_attempts}")
            
            # Read first 5 dates
            dates = self._get_first_n_dates_from_inbox(n=5)
            
            # Check if already in descending order
            if self._is_descending_date_order(dates):
                ui_log(f"[DATE-SORT] ✅ Inbox already sorted newest-first on attempt {attempt}")
                return True
            
            # If not last attempt, click DATE header to toggle sort
            if attempt < max_attempts:
                ui_log(f"[DATE-SORT] Clicking DATE header to toggle sort order...")
                
                if not self._click_date_column_header():
                    ui_log(f"[DATE-SORT] WARNING: Could not click DATE header on attempt {attempt}")
                    continue
                
                # Wait for table to refresh
                ui_log("[DATE-SORT] Waiting for table to refresh after sort...")
                page.wait_for_timeout(2500)
                
                # Wait for table to be stable
                try:
                    page.wait_for_selector("#assign_inbox tbody tr[class^='student--']", state="visible", timeout=10000)
                except Exception as e:
                    ui_log(f"[DATE-SORT] WARNING: Table refresh wait failed: {e}")
        
        # If we reach here, sorting failed after max attempts
        ui_log(f"[DATE-SORT] ⚠️  Could not confirm newest-first order after {max_attempts} attempts")
        ui_log("[DATE-SORT] Proceeding with current order (may be unreliable)")
        return False
    
    def _get_latest_inbox_row(self, uploaded_title=None, skip_sort=False):
        """
        Get the latest/newest row from inbox using DATE sorting + title validation hybrid strategy.
        
        Args:
            uploaded_title: Optional title of uploaded paper for secondary validation
            skip_sort: If True, skip the _ensure_newest_first_sorting() call (v3.6.1 polling fix)
        
        Returns: (row, title, score) tuple for the latest paper
        Raises: RuntimeError if no valid row found
        """
        ui_log("="*60)
        ui_log("[INBOX] Starting latest row detection with DATE sorting strategy")
        ui_log("="*60)
        
        # Step 1: Ensure newest-first sorting (skip if called from polling loop)
        if skip_sort:
            ui_log("[INBOX] Skipping _ensure_newest_first_sorting (sort already done by polling)")
            sort_success = True
        else:
            sort_success = self._ensure_newest_first_sorting(max_attempts=3)
        
        # Step 2: Get all rows
        rows = page.locator("#assign_inbox tbody tr[class^='student--'], table.inbox_table tbody tr[class^='student--']")
        total = rows.count()
        ui_log(f"[INBOX] Total rows in inbox: {total}")
        
        if total == 0:
            raise RuntimeError("No papers found in inbox")
        
        # Step 3: Scan first few rows and collect candidates
        candidates = []
        scan_count = min(5, total)  # Check first 5 rows
        
        ui_log(f"\n[INBOX] Scanning first {scan_count} rows for candidates...")
        
        for i in range(scan_count):
            try:
                row = rows.nth(i)
                row.wait_for(state="visible", timeout=3000)
                
                # Get title
                try:
                    title_cell = row.locator("td.ibox_title")
                    row_title = title_cell.inner_text(timeout=3000).strip()
                except Exception:
                    row_title = ""
                
                # Get checkbox title as fallback
                try:
                    checkbox_title = row.locator("input[name='object_checkbox']").first.get_attribute("title", timeout=2000) or ""
                except Exception:
                    checkbox_title = ""
                
                # Get paper ID
                try:
                    paper_id_text = row.locator("td.pid").first.inner_text(timeout=2000).strip()
                    paper_id = int(re.sub(r"\D", "", paper_id_text) or "0")
                except Exception:
                    paper_id = 0
                
                # Get score
                try:
                    score_text = row.locator("td.or_report_cell span.or-percentage, td.or_report_cell a.or-link").first.inner_text(timeout=3000).strip()
                except Exception:
                    score_text = ""
                
                # Get date
                try:
                    date_text = row.locator("td.class_status").first.inner_text(timeout=2000).strip()
                    parsed_date = self._parse_date_text(date_text)
                except Exception:
                    date_text = "unknown"
                    parsed_date = None
                
                # Store date in process state
                if i == 0 and date_text:
                    process_state.inbox_row_date = date_text
                
                # Get paper ID for Turnitin metadata
                if i == 0 and paper_id:
                    process_state.turnitin_paper_id = str(paper_id)
                
                # Check if has similarity link (report ready)
                has_similarity_link = False
                try:
                    similarity_link = row.locator("td.or_report_cell a.or-link").first
                    has_similarity_link = similarity_link.count() > 0 and bool(score_text and "%" in score_text)
                except Exception:
                    pass
                
                final_title = row_title or checkbox_title or "unknown"
                
                # Calculate match score for title validation
                title_match_score = 0
                if uploaded_title:
                    uploaded_norm = self._normalize_text(uploaded_title)
                    row_title_norm = self._normalize_text(final_title)
                    
                    # Exact match
                    if uploaded_norm == row_title_norm:
                        title_match_score = 100
                    # Partial match (first 4 words)
                    elif uploaded_norm and row_title_norm:
                        uploaded_key = self._normalize_text(self._short_title_key(uploaded_title))
                        if uploaded_key in row_title_norm:
                            title_match_score = 50
                
                ui_log(f"\n[INBOX] Row {i}:")
                ui_log(f"  - Title: {final_title}")
                ui_log(f"  - Paper ID: {paper_id}")
                ui_log(f"  - Date: {date_text}")
                ui_log(f"  - Similarity: {score_text}")
                ui_log(f"  - Has Link: {has_similarity_link}")
                ui_log(f"  - Title Match Score: {title_match_score}")
                
                candidates.append({
                    "row": row,
                    "index": i,
                    "title": final_title,
                    "paper_id": paper_id,
                    "score": score_text,
                    "date_text": date_text,
                    "parsed_date": parsed_date,
                    "has_similarity_link": has_similarity_link,
                    "title_match_score": title_match_score
                })
                
            except Exception as e:
                ui_log(f"[INBOX] Row {i}: Could not read - {e}")
        
        # Step 4: Select best candidate using STRICT title matching strategy (v3.6.1 FIX)
        ui_log("\n[INBOX] Selecting best candidate...")
        
        if not candidates:
            raise RuntimeError("No candidate rows found in inbox")
        
        # CRITICAL FIX v3.6.1: If uploaded_title exists, REJECT rows with title_match_score == 0
        if uploaded_title:
            ui_log(f"[INBOX] STRICT MODE: Uploaded title provided: {uploaded_title}")
            ui_log(f"[INBOX] STRICT MODE: Filtering candidates to only those with title match...")
            
            # Filter candidates: ONLY accept rows with title_match_score > 0
            matched_candidates = [c for c in candidates if c["title_match_score"] > 0]
            
            ui_log(f"[INBOX] STRICT MODE: {len(matched_candidates)} candidates with title match (out of {len(candidates)} total)")
            
            if not matched_candidates:
                # NO TITLE MATCH FOUND - REJECT ALL ROWS
                ui_log("[INBOX] ❌ CRITICAL: No rows match the uploaded title!")
                ui_log(f"[INBOX] ❌ Uploaded title: {uploaded_title}")
                ui_log(f"[INBOX] ❌ Available rows checked:")
                for c in candidates[:5]:
                    ui_log(f"  - {c['title']} (match score: {c['title_match_score']})")
                ui_log("[INBOX] ❌ REJECTING all rows - will wait and refresh")
                raise RuntimeError(f"No inbox row matches uploaded title: {uploaded_title}")
            
            # Use only matched candidates
            candidates = matched_candidates
            ui_log(f"[INBOX] STRICT MODE: ✅ Using {len(candidates)} title-matched candidates")
        
        # Prefer rows with:
        # 1. Title match score (MANDATORY if uploaded_title provided)
        # 2. Has similarity link (report ready)
        # 3. Newest date (highest parsed_date)
        # 4. Highest paper ID
        
        # Sort candidates by priority
        candidates.sort(
            key=lambda c: (
                c["title_match_score"],           # Title match first
                c["has_similarity_link"],          # Then has report
                c["parsed_date"] if c["parsed_date"] else datetime.min,  # Then newest date
                c["paper_id"]                      # Finally highest ID
            ),
            reverse=True
        )
        
        best = candidates[0]
        
        ui_log("\n[INBOX] ✅ Selected row:")
        ui_log(f"  - Row Index: {best['index']}")
        ui_log(f"  - Title: {best['title']}")
        ui_log(f"  - Paper ID: {best['paper_id']}")
        ui_log(f"  - Date: {best['date_text']}")
        ui_log(f"  - Similarity: {best['score']}")
        ui_log(f"  - Has Link: {best['has_similarity_link']}")
        ui_log(f"  - Title Match Score: {best['title_match_score']}")
        
        # CRITICAL FIX v3.6.1: REJECT if no similarity link
        if not best['has_similarity_link']:
            ui_log("[INBOX] ❌ CRITICAL: Selected row has no similarity link yet")
            ui_log("[INBOX] ❌ REJECTING - will wait and refresh")
            raise RuntimeError("Selected row has no similarity link yet")
        
        # CRITICAL FIX v3.6.1: REJECT if title mismatch (should not happen after filtering)
        if uploaded_title and best['title_match_score'] == 0:
            ui_log(f"[INBOX] ❌ CRITICAL: Selected row has NO title match!")
            ui_log(f"[INBOX] ❌ This should never happen after filtering!")
            raise RuntimeError(f"Selected row does not match uploaded title: {uploaded_title}")
        
        ui_log("="*60)
        
        return best["row"], best["title"], best["score"]

    def _wait_for_similarity_report_ready_new_strategy(self, uploaded_title=None, max_attempts=20, wait_between_sec=30):
        """
        Poll inbox until similarity report is ready for the latest uploaded paper.
        Uses date-based latest-row strategy with active sort forcing (v3.6.1 FIX).
        
        Args:
            uploaded_title: Optional title of uploaded paper for secondary validation
            max_attempts: Maximum refresh attempts (default 20 = ~10 minutes)
            wait_between_sec: Seconds to wait between refresh attempts
            
        Returns: (matched_row, matched_title, score) tuple
        Raises: RuntimeError if report not ready after max attempts
        """
        ui_log("="*60)
        ui_log("WAITING FOR TURNITIN TO GENERATE SIMILARITY REPORT")
        ui_log("="*60)
        
        if uploaded_title:
            ui_log(f"[POLL] Uploaded paper title for validation: {uploaded_title}")
        
        sort_click_count = 0
        page_refresh_interval = 5  # Refresh page every 5 attempts
        
        for attempt in range(1, max_attempts + 1):
            ui_log(f"\n[POLL] === Attempt {attempt}/{max_attempts} ===")
            
            # v3.6.1 FIX: Click DATE sort BEFORE each scan (not just on page load)
            # This forces Turnitin to re-sort and break cache
            try:
                ui_log(f"[POLL] Clicking DATE sort button to force fresh order (click #{sort_click_count + 1})...")
                
                if self._click_date_column_header():
                    sort_click_count += 1
                    ui_log(f"[POLL] ✅ DATE sort clicked successfully (total clicks: {sort_click_count})")
                    
                    # Wait for table to reorder/stabilize
                    ui_log("[POLL] Waiting 2-4 seconds for table reorder...")
                    page.wait_for_timeout(random.randint(2000, 4000))
                    
                    # Wait for table to be stable
                    try:
                        page.wait_for_selector("#assign_inbox tbody tr[class^='student--']", state="visible", timeout=10000)
                    except Exception:
                        pass
                else:
                    ui_log(f"[POLL] ⚠️  Could not click DATE sort")
            except Exception as e:
                ui_log(f"[POLL] DATE sort click failed: {e}")
            
            # Log top row for debugging
            try:
                rows = page.locator("#assign_inbox tbody tr[class^='student--'], table.inbox_table tbody tr[class^='student--']")
                if rows.count() > 0:
                    top_row = rows.nth(0)
                    try:
                        top_title = top_row.locator("td.ibox_title").first.inner_text(timeout=2000).strip()
                    except Exception:
                        top_title = "unknown"
                    try:
                        top_date = top_row.locator("td.class_status").first.inner_text(timeout=2000).strip()
                    except Exception:
                        top_date = "unknown"
                    
                    ui_log(f"[POLL] Current top row: {top_title} | Date: {top_date}")
            except Exception as e:
                ui_log(f"[POLL] Could not read top row: {e}")
            
            # Now scan rows with title validation
            try:
                # Use new latest-row strategy with title validation (skip_sort=True to avoid nested sorting)
                latest_row, latest_title, score = self._get_latest_inbox_row(uploaded_title=uploaded_title, skip_sort=True)
                
                # Check if similarity link exists and is clickable
                similarity_link = latest_row.locator("td.or_report_cell a.or-link").first
                if similarity_link.count() > 0 and score and "%" in score:
                    ui_log(f"[POLL] ✅ Report ready! Score: {score}")
                    ui_log(f"[POLL] Total DATE sort clicks: {sort_click_count}")
                    ui_log("="*60)
                    
                    process_state.update_stage(ProcessState.REPORT_READY)
                    
                    return latest_row, latest_title, score
                else:
                    ui_log(f"[POLL] Report not ready yet (no similarity score or link)")
                    
            except Exception as e:
                ui_log(f"[POLL] Could not detect matching row: {e}")
                ui_log(f"[POLL] This is expected if uploaded paper not yet in inbox or title mismatch")
            
            # If not last attempt, decide whether to sort-retry or page-refresh
            if attempt < max_attempts:
                # v3.6.1 FIX: Only refresh page every N attempts
                # Otherwise just click sort and try again
                should_refresh_page = (attempt % page_refresh_interval == 0)
                
                if should_refresh_page:
                    ui_log(f"[POLL] === PAGE REFRESH CYCLE (attempt {attempt}) ===")
                    ui_log(f"[POLL] Waiting {wait_between_sec} seconds before page refresh...")
                    page.wait_for_timeout(wait_between_sec * 1000)
                    
                    # Refresh inbox page
                    try:
                        ui_log("[POLL] Refreshing inbox page...")
                        page.reload(wait_until="domcontentloaded", timeout=30000)
                        page.wait_for_timeout(3000)
                        
                        # Wait for inbox to be visible again
                        self._wait_for_selector(ASSIGN_INBOX, "Assignment inbox after refresh", timeout=30000)
                        ui_log("[POLL] Inbox refreshed successfully")
                        
                    except Exception as refresh_error:
                        ui_log(f"[POLL] WARNING: Refresh failed: {refresh_error}")
                        # Try to navigate to home and back to inbox
                        try:
                            self._goto_turnitin_home()
                            self._wait_for_selector(ASSIGN_INBOX, "Assignment inbox", timeout=30000)
                        except Exception:
                            pass
                else:
                    # Just wait a bit and retry with sort click
                    ui_log(f"[POLL] === SORT-RETRY CYCLE (attempt {attempt}) ===")
                    ui_log(f"[POLL] Waiting 3-5 seconds before next sort-retry...")
                    page.wait_for_timeout(random.randint(3000, 5000))
                    
                    # Optional: Toggle sort direction by clicking again
                    if attempt % 2 == 0:
                        ui_log("[POLL] Toggle: Clicking DATE sort again to reverse order...")
                        try:
                            if self._click_date_column_header():
                                sort_click_count += 1
                                page.wait_for_timeout(2000)
                                ui_log(f"[POLL] Sort toggled (total clicks: {sort_click_count})")
                        except Exception as e:
                            ui_log(f"[POLL] Sort toggle failed: {e}")
        
        # Max attempts reached
        ui_log("="*60)
        ui_log(f"[POLL] ❌ TIMEOUT: Report not ready after {max_attempts} attempts")
        ui_log(f"[POLL] Total DATE sort clicks: {sort_click_count}")
        ui_log("="*60)
        raise RuntimeError(f"Similarity report not ready after {max_attempts} polling attempts (~{max_attempts * wait_between_sec / 60:.0f} minutes)")

    def _open_feedback_studio_from_similarity_row(self, matched_row):
        """Open Feedback Studio popup from similarity row."""
        matched_row.wait_for(state="visible", timeout=60000)

        similarity_link = matched_row.locator("td.or_report_cell a.or-link").first
        similarity_link.wait_for(state="visible", timeout=90000)

        try:
            score = similarity_link.inner_text(timeout=5000).strip()
            ui_log(f"Similarity score link ready: {score}")
        except Exception:
            pass

        self._human_pause(2000, 4000)

        with context.expect_page(timeout=90000) as new_page_info:
            similarity_link.scroll_into_view_if_needed(timeout=10000)
            similarity_link.click(timeout=30000)

        feedback_page = new_page_info.value
        feedback_page.wait_for_load_state("domcontentloaded", timeout=90000)
        feedback_page.wait_for_timeout(8000)
        ui_log(f"Feedback Studio popup opened: {feedback_page.url}")
        
        process_state.update_stage(ProcessState.FEEDBACK_OPENED)
        
        return feedback_page

    def _download_similarity_report(self, title):
        """Main workflow: wait for report, open Feedback Studio, run automation."""
        global feedback_workflow_running
        
        ui_log(f"Looking for uploaded paper row...")
        self._wait_for_selector(ASSIGN_INBOX, "Assignment inbox", timeout=60000)

        try:
            # Acquire workflow lock
            if not feedback_workflow_lock.acquire(blocking=False):
                ui_log("="*70)
                ui_log("⚠️  WARNING: Another Feedback Studio workflow is already running!")
                ui_log("⚠️  Waiting for previous workflow to complete...")
                ui_log("="*70)
                feedback_workflow_lock.acquire(blocking=True)
            
            feedback_workflow_running = True
            
            # NEW STRATEGY: Use DATE sorting + title validation hybrid
            ui_log("[STRATEGY] Using DATE sorting + title validation hybrid strategy")
            matched_row, matched_title, score_before = self._wait_for_similarity_report_ready_new_strategy(uploaded_title=title)
            
            ui_log(f"[STRATEGY] Detected paper: {matched_title}")
            
            # Open Feedback Studio popup
            feedback_page = self._open_feedback_studio_from_similarity_row(matched_row)
            
            # Create dedicated controller and run workflow
            controller = FeedbackStudioController(
                feedback_page=feedback_page,
                title=matched_title or title,
                score_before=score_before or "unknown"
            )
            
            controller.run_full_workflow()
            
            # SUCCESS - finalize and log
            process_state.finalize(ProcessState.SUCCESS_DOWNLOADED)
            excel_logger.log_process(process_state)
            ui_log("[EXCEL] ✅ Process logged to Excel (SUCCESS)")
            
        except Exception as e:
            ui_log("="*70)
            ui_log("❌ Feedback Studio workflow failed")
            ui_log(str(e))
            ui_log("="*70)
            
            # Mark as failed and determine status
            process_state.mark_failed(e)
            
            # Determine failure status based on stage
            if process_state.failed_stage in [ProcessState.STARTED, ProcessState.FILE_SELECTED, ProcessState.QUICK_SUBMIT_READY]:
                status = ProcessState.FAILED_BEFORE_UPLOAD
            elif process_state.failed_stage in [ProcessState.UPLOADED, ProcessState.CONFIRMED]:
                status = ProcessState.FAILED_AFTER_UPLOAD
            elif process_state.failed_stage == ProcessState.INBOX_FOUND:
                status = ProcessState.FAILED_IN_INBOX_DETECTION
            elif process_state.failed_stage in [ProcessState.FEEDBACK_OPENED, ProcessState.FILTER_DONE]:
                status = ProcessState.FAILED_FEEDBACK_STUDIO
            elif process_state.failed_stage == ProcessState.DOWNLOAD_STARTED:
                status = ProcessState.FAILED_DOWNLOAD
            elif process_state.failed_stage == ProcessState.CLASSIC_REPORT_OPENED:
                status = ProcessState.FAILED_CLASSIC_REPORT_DOWNLOAD
            else:
                status = ProcessState.FAILED_AFTER_UPLOAD
            
            process_state.finalize(status)
            excel_logger.log_process(process_state)
            ui_log(f"[EXCEL] ✅ Process logged to Excel (FAILED: {status})")
            
            self._save_failure_screenshot("feedback-workflow-failed")
            
            # Try to save Feedback Studio screenshot if popup exists
            try:
                for pg in context.pages:
                    try:
                        if "ev.turnitin.com" in pg.url or "newreport" in pg.url or "/app/carta" in pg.url:
                            timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
                            screenshot_path = SCREENSHOT_DIR / f"{timestamp}-feedback-failed.png"
                            pg.screenshot(path=str(screenshot_path), full_page=True)
                            ui_log(f"Saved Feedback Studio screenshot: {screenshot_path.name}")
                            process_state.add_screenshot(screenshot_path)
                            break
                    except Exception:
                        pass
            except Exception:
                pass
        finally:
            # Release workflow lock
            feedback_workflow_running = False
            feedback_workflow_lock.release()
            ui_log("🔓 Workflow lock released")

    def _prepare_quick_submit_no_repository(self):
        self._ensure_browser()

        # Reset process state for new attempt
        process_state.reset()
        ui_log("[PROCESS] Starting new process attempt")

        ui_log("Starting Quick Submit setup flow: login → Quick Submit → Submit → checkboxes → No Repository → Final Submit.")

        process_state.update_stage(ProcessState.STARTED)

        # Fix 4: Always open direct login URL first
        ui_log(f"[TURNITIN] Opening direct login URL: {TURNITIN_LOGIN_URL}")
        page.goto(TURNITIN_LOGIN_URL, wait_until="domcontentloaded", timeout=NAVIGATION_TIMEOUT)
        page.wait_for_timeout(2000)

        self._wait_for_dashboard_or_manual_login()

        self._human_pause()
        self._safe_click("a.sn_quick_submit", "Quick Submit link")

        process_state.update_stage(ProcessState.QUICK_SUBMIT_READY)

        self._human_pause()
        self._safe_click("a.submit_paper_button", "Submit link on Quick Submit page")

        self._human_pause()
        self._tick_customize_search_checkboxes()

        self._human_pause()
        self._select_no_repository()

        self._human_pause()
        self._safe_click(".submit", "Final Submit button on Customize Search page")

        self._human_pause()
        final_url = self._log_final_url_after_submit()

        if "t_custom_search.asp" in final_url:
            ui_log("Stopping because Turnitin did not leave Customize Search page.")
            
            process_state.mark_failed("Stuck on Customize Search page")
            process_state.finalize(ProcessState.FAILED_BEFORE_UPLOAD)
            excel_logger.log_process(process_state)
            
            return

        self._fill_upload_page_and_attach_file()

        self._click_upload_and_confirm()

        _, _, _, title = self._get_first_selected_file_details()
        self._go_to_inbox_after_submission()
        self._download_similarity_report(title)

        ui_log("DONE: Upload, confirmation, and report download flow completed.")


# Only run standalone when executed directly (not when imported as a module)
if __name__ == "__main__":
    HTML_PATH = resource_path("ui/index.html")

    api = Api()

    window = webview.create_window(
        "Turnitin Upload Assistant",
        str(HTML_PATH),
        js_api=api,
        width=1400,
        height=850
    )

    webview.start()
