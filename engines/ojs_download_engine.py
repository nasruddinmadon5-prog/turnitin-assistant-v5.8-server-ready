from playwright.sync_api import sync_playwright
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
import time
import random
import re
import threading

# =========================
# CONFIG
# =========================

OJS_SUBMISSIONS_URL = "https://journals.utm.my/jurnalteknologi/submissions"

# Get absolute paths relative to v5.3.1 project root with stable profiles
PROJECT_ROOT = Path(__file__).resolve().parent.parent

# Import stable profile paths from config
import sys
sys.path.insert(0, str(PROJECT_ROOT))
from config import STABLE_OJS_PROFILE_DIR, OJS_DOWNLOADS_DIR as CONFIG_OJS_DOWNLOADS_DIR

PROFILE_DIR = str(STABLE_OJS_PROFILE_DIR)
DOWNLOAD_DIR = CONFIG_OJS_DOWNLOADS_DIR
DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

MASTER_XLSX = PROJECT_ROOT / "data" / "master update.xlsx"

MAX_SCAN_PAGES = 50
MIN_REVIEW_AFTER_SUBMISSION = 2
MAX_DOWNLOAD = 1
ITEMS_PER_PAGE = 30

STUCK_SECONDS = 60
HUMAN_SLOW_MO = 2000  # v5.5.1: Slower human-like typing speed (was 900)

# Workflow guard - prevent concurrent OJS runs
_workflow_lock = threading.Lock()
_workflow_running = False

# UI Logging callback (will be set by main app)
_ui_log_callback = None

def set_ui_log_callback(callback):
    """Set the UI logging callback function"""
    global _ui_log_callback
    _ui_log_callback = callback

def ui_log(message):
    """Log message to UI if callback is set, otherwise print"""
    if _ui_log_callback:
        _ui_log_callback(message)
    else:
        print(message)


# =========================
# HELPERS
# =========================

def human_wait(page, a=2000, b=5000):
    page.wait_for_timeout(random.randint(a, b))


def clean_filename(text):
    text = re.sub(r'[\\/:*?"<>|]+', "_", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text[:160]


def wait_until(page, label, js, timeout=STUCK_SECONDS):
    start = time.time()

    while time.time() - start < timeout:
        try:
            if page.evaluate(js):
                ui_log(f"[OK] {label}")
                return True
        except:
            pass

        human_wait(page, 1500, 3000)

    raise Exception(f"Timeout: {label}")


# =========================
# EXCEL
# =========================

def get_or_create_master():
    """
    Guna workbook:
    /Users/nasruddinmadon/turnitin-assistant-v2/master update.xlsx

    Guna sheet khusus:
    download

    Susunan kolum:
    A = timestamp
    B = submission_no
    C = author_name
    D = title
    E = download path
    """

    if MASTER_XLSX.exists():
        wb = load_workbook(MASTER_XLSX)

        if "download" in wb.sheetnames:
            ws = wb["download"]
        else:
            ws = wb.create_sheet("download")
    else:
        wb = Workbook()

        # Buang default sheet supaya workbook bersih
        default_ws = wb.active
        wb.remove(default_ws)

        ws = wb.create_sheet("download")

    # Pastikan header ikut format yang diminta
    if ws.max_row == 1 and ws["A1"].value is None:
        ws["A1"] = "timestamp"
        ws["B1"] = "submission_no"
        ws["C1"] = "author_name"
        ws["D1"] = "title"
        ws["E1"] = "download path"

    wb.save(MASTER_XLSX)

    return wb, ws


def estimate_wrapped_lines(value, width):
    """Anggar bilangan baris visual untuk wrap text dalam Excel.
    Excel tidak sentiasa auto-fit row height bila fail dijana melalui openpyxl,
    jadi kita kira tinggi row berdasarkan panjang teks + lebar kolum.
    """
    text = str(value or "")
    if not text:
        return 1

    # Pecah ikut newline dulu, kemudian anggar wrap ikut column width.
    lines = 0
    usable_width = max(8, int(width * 0.85))

    for part in text.split("\n"):
        part_len = len(part.strip())
        lines += max(1, (part_len // usable_width) + 1)

    return max(1, lines)


def apply_excel_format(ws, row_number=None):
    """
    Format Excel:
    - Font Arial
    - Size 20pt
    - Wrap text
    - Top vertical alignment
    - Row height dikira supaya teks tidak crop
    """

    target_rows = range(1, ws.max_row + 1) if row_number is None else [row_number]

    column_widths = {
        "A": 22,
        "B": 18,
        "C": 28,
        "D": 58,
        "E": 95,
    }

    # Lebarkan kolum supaya wrap lebih kemas.
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    for row in target_rows:
        max_lines = 1

        for col_idx, col_letter in enumerate(["A", "B", "C", "D", "E"], start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.font = Font(name="Arial", size=20, bold=(row == 1))
            cell.alignment = Alignment(
                wrap_text=True,
                vertical="top",
                horizontal="center" if col_letter in ["A", "B", "C"] else "left"
            )

            max_lines = max(
                max_lines,
                estimate_wrapped_lines(cell.value, column_widths[col_letter])
            )

        # 20pt font perlukan lebih kurang 25pt setiap wrapped line.
        # Had maksimum Excel row height sekitar 409pt.
        if row == 1:
            ws.row_dimensions[row].height = 38
        else:
            ws.row_dimensions[row].height = min(409, max(45, max_lines * 26))


def log_excel(data):
    """
    Log hanya ke sheet download.

    Kolum:
    B = submission_no
    C = author_name
    D = title
    E = download path

    Format:
    - Arial
    - 20pt
    - Wrap text
    """

    wb, ws = get_or_create_master()

    next_row = ws.max_row + 1

    ws[f"A{next_row}"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws[f"B{next_row}"] = data.get("submission_id", "")
    ws[f"C{next_row}"] = data.get("author", "")
    ws[f"D{next_row}"] = data.get("title", "")
    ws[f"E{next_row}"] = data.get("downloaded_path", "")

    # Apply format pada header + row baru
    apply_excel_format(ws, 1)
    apply_excel_format(ws, next_row)

    wb.save(MASTER_XLSX)


def get_downloaded_submission_ids():
    """Baca sheet download dan pulangkan submission_no yang sudah berjaya ada download path."""
    downloaded = set()

    if not MASTER_XLSX.exists():
        return downloaded

    try:
        wb = load_workbook(MASTER_XLSX, read_only=True, data_only=True)
        if "download" not in wb.sheetnames:
            return downloaded

        ws = wb["download"]

        for row in ws.iter_rows(min_row=2, values_only=True):
            # A timestamp, B submission_no, C author, D title, E download path
            submission_no = str(row[1] or "").strip() if len(row) > 1 else ""
            download_path = str(row[4] or "").strip() if len(row) > 4 else ""

            # Hanya anggap selesai kalau ada submission_no dan ada download path.
            if submission_no and download_path:
                downloaded.add(submission_no)

    except Exception as e:
        ui_log("[WARN] Gagal baca Excel downloaded list:" + str(e))

    return downloaded


# =========================
# LOGIN
# =========================

def wait_manual_login_no_input(page):
    """Wait for manual login without terminal input prompt"""
    ui_log("\n======================================")
    ui_log("LOGIN MANUAL DULU")
    ui_log("1. Login OJS")
    ui_log("2. Pergi page Submissions")
    ui_log("3. Pastikan nampak Active / Archived")
    ui_log("4. Workflow akan continue automatically")
    ui_log("======================================\n")

    # Wait for page to be ready (not on login page)
    max_wait = 300  # 5 minutes
    start_time = time.time()
    
    while time.time() - start_time < max_wait:
        current_url = page.url
        ui_log(f"Current URL: {current_url}")
        
        if "/login" not in current_url:
            ui_log("[OK] Login detected, proceeding...")
            human_wait(page, 3000, 5000)
            return
        
        ui_log("[WAIT] Waiting for login... (checking again in 5 seconds)")
        time.sleep(5)
    
    ui_log("[STOP] Login timeout after 5 minutes")
    raise Exception("Login timeout - user did not login within 5 minutes")


# =========================
# SUBMISSION LIST
# =========================

def go_active_tab(page):
    ui_log("Pergi tab Active...")

    selectors = [
        '#active-button',
        'button:has-text("Active")',
        'a:has-text("Active")',
        'text=Active'
    ]

    for sel in selectors:
        try:
            loc = page.locator(sel).first

            if loc.count() > 0 and loc.is_visible():
                loc.scroll_into_view_if_needed()
                human_wait(page, 1000, 2500)
                loc.click(timeout=20000)
                human_wait(page, 5000, 8000)

                ui_log("[OK] Active tab")
                return
        except:
            pass

    ui_log("[WARN] Active tab mungkin sudah aktif.")


def click_next_page(page):
    selectors = [
        'button:has-text("Next")',
        'a:has-text("Next")',
        '[aria-label="Next page"]',
        '.pkpPagination__next',
        'button.next',
        'a.next'
    ]

    for sel in selectors:
        try:
            loc = page.locator(sel).first

            if loc.count() > 0 and loc.is_visible():
                loc.scroll_into_view_if_needed()

                human_wait(page, 1000, 2500)

                loc.click(timeout=20000)

                human_wait(page, 5000, 9000)

                return True

        except:
            pass

    return False


def extract_rows(page):
    """Extract visible submission rows from the current OJS list page."""
    return page.evaluate("""
    () => {
        const clean = (s) => (s || '').replace(/\s+/g, ' ').trim();
        const results = [];
        const items = [...document.querySelectorAll('.pkpTab--isActive li.listPanel__item')];

        for (const item of items) {
            const txt = item.innerText || '';
            const idText = clean(item.querySelector('.listPanel__item--submission__id')?.innerText || '');
            const idMatch = idText.match(/\b(\d{4,8})\b/) || txt.match(/\b(\d{4,8})\b/);
            if (!idMatch) continue;
            const submissionId = idMatch[1];
            const author = clean(item.querySelector('.listPanel__itemTitle')?.innerText || '');
            const title = clean(item.querySelector('.listPanel__itemSubtitle')?.innerText || '');
            const stageText = clean(item.querySelector('.listPanel__item--submission__stage')?.innerText || '');
            const incomplete = /\bIncomplete\b/i.test(stageText) || /Currently in the Incomplete stage/i.test(txt);
            let status = '';
            if (/\bIncomplete\b/i.test(stageText)) status = 'Incomplete';
            else if (/\bSubmission\b/i.test(stageText)) status = 'Submission';
            else if (/\bReview\b/i.test(stageText)) status = 'Review';
            else if (/\bCopyediting\b/i.test(stageText)) status = 'Copyediting';
            else if (/\bProduction\b/i.test(stageText)) status = 'Production';
            else if (/\bSubmission\b/i.test(txt)) status = 'Submission';
            else if (/\bReview\b/i.test(txt)) status = 'Review';
            const links = [...item.querySelectorAll('a[href]')];
            const workflow = links.find(a => /\/workflow\/access\//.test(a.href || ''));
            const view = links.find(a => clean(a.innerText) === 'View' || /View/i.test(a.getAttribute('aria-label') || ''));
            const any = links.find(a => /submission\?id=|\/workflow\/access\//.test(a.href || ''));
            const href = (workflow || view || any || {}).href || '';
            results.push({submission_id: submissionId, text: txt.trim(), href, status, incomplete, author, title, stage_text: stageText});
        }
        if (results.length) return results;

        const links = [...document.querySelectorAll('a[href*="/workflow/access/"]')];
        const unique = [];
        const seen = new Set();
        for (const a of links) {
            const row = a.closest('li.listPanel__item') || a.closest('tr') || a.closest('.listPanel__item--submission');
            if (!row) continue;
            const txt = row.innerText || '';
            const href = a.href || '';
            const idMatch = href.match(/workflow\/access\/(\d+)/) || txt.match(/\b(\d{5})\b/);
            if (!idMatch) continue;
            const submissionId = idMatch[1];
            if (seen.has(submissionId)) continue;
            seen.add(submissionId);
            const status = /\bIncomplete\b/i.test(txt) ? 'Incomplete' : /\bSubmission\b/i.test(txt) ? 'Submission' : /\bReview\b/i.test(txt) ? 'Review' : /\bCopyediting\b/i.test(txt) ? 'Copyediting' : /\bProduction\b/i.test(txt) ? 'Production' : '';
            unique.push({submission_id: submissionId, text: txt.trim(), href, status, incomplete: /Incomplete/i.test(txt), author: '', title: '', stage_text: status});
        }
        return unique;
    }
    """)

def parse_row_metadata(row_or_text):

    if isinstance(row_or_text, dict):
        author = (row_or_text.get("author") or "").strip()
        title = (row_or_text.get("title") or "").strip()
        if author or title:
            return author, title
        row_text = row_or_text.get("text") or ""
    else:
        row_text = row_or_text or ""

    lines = [x.strip() for x in row_text.split("\n") if x.strip()]
    author = ""
    title = ""

    for line in lines:
        if " et al." in line:
            author = line
            break

    skip_words = ["view", "submission", "review", "incomplete", "copyediting", "production", "currently in the", "show more details"]
    candidates = [x for x in lines if len(x) > 20 and not any(w in x.lower() for w in skip_words)]
    if candidates:
        title = max(candidates, key=len)
    return author, title


def find_first_valid_submission_row(rows, downloaded_ids=None, scan_bottom_to_top=False):
    """
    Cari row Submission yang valid.

    scan_bottom_to_top=True:
    - mula dari row paling bawah
    - naik ke atas

    downloaded_ids:
    - submission_no yang sudah ada dalam Excel
    - kalau sudah ada, skip supaya tidak download berulang
    """

    downloaded_ids = downloaded_ids or set()

    if scan_bottom_to_top:
        indexed_rows = list(enumerate(rows, start=1))
        indexed_rows = list(reversed(indexed_rows))
    else:
        indexed_rows = list(enumerate(rows, start=1))

    for i, row in indexed_rows:
        sid = str(row.get("submission_id") or "").strip()

        if not sid:
            ui_log(f"[SKIP] row {i} submission_id kosong")
            continue

        if sid in downloaded_ids:
            ui_log(f"[SKIP] {sid} sudah ada dalam Excel")
            continue

        if row.get("incomplete"):
            ui_log(f"[SKIP] {sid} incomplete")
            continue

        if row.get("status") != "Submission":
            ui_log(f"[SKIP] {sid} status={row.get('status')}")
            continue

        if not row.get("href"):
            ui_log(f"[SKIP] {sid} href kosong")
            continue

        return i, row

    return None, None


def get_last_visible_page_number(page):
    try:
        return int(page.evaluate("""
        () => {
            const nums = [...document.querySelectorAll('.pkpTab--isActive nav.pkpPagination button')]
                .map(b => (b.innerText || '').trim())
                .filter(t => /^\d+$/.test(t))
                .map(t => parseInt(t, 10));
            return nums.length ? Math.max(...nums) : 1;
        }
        """))
    except Exception:
        return 1


def click_page_number(page, page_number):
    clicked = page.evaluate("""
    (pageNumber) => {
        const buttons = [...document.querySelectorAll('.pkpTab--isActive nav.pkpPagination button')];
        const btn = buttons.find(b => (b.innerText || '').trim() === String(pageNumber));
        if (!btn) return false;
        btn.scrollIntoView({block: 'center', inline: 'center'});
        btn.click();
        return true;
    }
    """, page_number)
    if clicked:
        human_wait(page, 5000, 9000)
        return True
    return False


def go_to_last_visible_page(page):
    last_seen = 1
    for _ in range(10):
        last_page = get_last_visible_page_number(page)
        ui_log(f"[PAGINATION] Last visible page: {last_page}")
        if last_page <= last_seen:
            break
        last_seen = last_page
        if not click_page_number(page, last_page):
            break
    return last_seen


def apply_stage_submission_filter(page):
    ui_log("Apply filter: Stage = Submission...")
    filter_buttons = [
        '.pkpTab--isActive .pkpHeader__actions button:has-text("Filters")',
        '.pkpTab--isActive button:has-text("Filters")',
        'button:has-text("Filters")'
    ]
    opened = False
    for sel in filter_buttons:
        try:
            loc = page.locator(sel).first
            if loc.count() > 0 and loc.is_visible():
                loc.scroll_into_view_if_needed()
                human_wait(page, 800, 1800)
                loc.click(timeout=20000)
                human_wait(page, 1500, 3500)
                opened = True
                break
        except Exception:
            pass
    if not opened:
        ui_log("[WARN] Filter button tak jumpa / mungkin filter panel sudah terbuka.")

    clicked = False
    candidates = ['button:has-text("Submission")', 'label:has-text("Submission")', 'input[value="1"]', '[data-param="stageIds"][data-value="1"]', 'text=Submission']
    for sel in candidates:
        try:
            loc = page.locator(sel).last
            if loc.count() > 0 and loc.is_visible():
                loc.scroll_into_view_if_needed()
                human_wait(page, 800, 1800)
                loc.click(timeout=20000)
                human_wait(page, 5000, 9000)
                clicked = True
                ui_log("[OK] Submission filter clicked")
                break
        except Exception:
            pass
    if not clicked:
        ui_log("[WARN] Tak dapat klik filter UI. Akan cuba guna API/pagination DOM semasa.")
    return clicked


def fetch_filtered_rows_by_api(page, target_page=1):
    return page.evaluate("""
    async ({targetPage, count}) => {
        const base = window.pkp?.app?.baseUrl || 'https://journals.utm.my';
        const context = window.pkp?.app?.contextPath || 'jurnalteknologi';
        const url = new URL(`${base}/${context}/api/v1/_submissions`);
        url.searchParams.set('status', '1');
        url.searchParams.append('stageIds[]', '1');
        url.searchParams.set('count', String(count));
        url.searchParams.set('page', String(targetPage));
        const res = await fetch(url.toString(), {credentials: 'same-origin', headers: {'Accept': 'application/json'}});
        if (!res.ok) return {ok: false, status: res.status, error: await res.text()};
        const data = await res.json();
        const items = Array.isArray(data.items) ? data.items : (Array.isArray(data) ? data : []);
        const total = data.itemsMax || data.count || data.total || items.length || 0;
        const lastPage = Math.max(1, Math.ceil(total / count));
        const rows = items.map(item => {
            const id = item.id || item.submissionId || '';
            const author = item.authorsStringShort || item.authorString || item.author || item.authorsString || '';
            let title = '';
            if (typeof item.title === 'string') title = item.title;
            else if (item.title && typeof item.title === 'object') title = item.title.en || Object.values(item.title)[0] || '';
            else if (typeof item.fullTitle === 'string') title = item.fullTitle;
            const stageId = item.stageId || item.currentStageId || item.stage_id || '';
            const isIncomplete = !!(item.isIncomplete || item.incomplete || item.submissionProgress);
            const stageLabel = item.stageLabel || item.currentStageLabel || (stageId == 1 ? 'Submission' : '');
            let href = item.urlWorkflow || item.urlEditorialWorkflow || item.workflowUrl || item.url || '';
            if (!href && id) href = `${base}/${context}/workflow/access/${id}`;
            return {submission_id: String(id), author: String(author || '').trim(), title: String(title || '').replace(/<[^>]*>/g, '').trim(), href, status: isIncomplete ? 'Incomplete' : 'Submission', incomplete: isIncomplete, text: '', stage_text: stageLabel};
        }).filter(r => r.submission_id && r.href);
        return {ok: true, total, lastPage, rows, url: url.toString()};
    }
    """, {"targetPage": target_page, "count": ITEMS_PER_PAGE})


def get_target_row_submission_filter(page):
    """
    Rule baru:
    1. Ambil filtered Submission list.
    2. Kira last_page sebenar dari API/OJS.
    3. Guna konsep human-readable:
       BACK PAGE 1 = real OJS last_page
       BACK PAGE 2 = real OJS last_page - 1
       BACK PAGE 3 = real OJS last_page - 2
    4. Dalam setiap page, scan row dari bawah ke atas.
    5. Skip kalau incomplete / bukan Submission / sudah ada dalam Excel.
    """

    downloaded_ids = get_downloaded_submission_ids()
    ui_log(f"[EXCEL] Sudah downloaded dalam Excel: {len(downloaded_ids)}")

    # Cara utama: guna API OJS sebab lebih stabil daripada klik pagination UI.
    try:
        first = fetch_filtered_rows_by_api(page, 1)
        if first and first.get("ok"):
            total = int(first.get("total") or 0)
            last_page = int(first.get("lastPage") or 1)
            ui_log(f"[API] Filtered Submission total: {total} | real last page: {last_page}")
            ui_log("[RULE] BACK PAGE 1 = real OJS last page")

            max_back_pages = min(MAX_SCAN_PAGES, last_page)

            for back_page in range(1, max_back_pages + 1):
                real_page = last_page - back_page + 1
                ui_log(f"[SCAN] BACK PAGE {back_page} / REAL OJS PAGE {real_page}")

                data = fetch_filtered_rows_by_api(page, real_page)
                if not data or not data.get("ok"):
                    ui_log(f"[WARN] API gagal fetch real page {real_page}")
                    continue

                rows = data.get("rows") or []
                ui_log(f"[API] Page {real_page}: {len(rows)} rows. Scan bottom -> top")

                idx, row = find_first_valid_submission_row(
                    rows,
                    downloaded_ids=downloaded_ids,
                    scan_bottom_to_top=True
                )

                if row:
                    row["api_mode"] = True
                    row["back_page"] = back_page
                    row["real_ojs_page"] = real_page
                    return real_page, idx, row

            ui_log("[API] Tiada row Submission valid yang belum download dalam semua BACK PAGE.")

    except Exception as e:
        ui_log("[WARN] API filter gagal:" + str(e))

    # Fallback: guna UI filter. Ini kurang ideal tapi masih boleh jalan kalau API berubah.
    apply_stage_submission_filter(page)
    last_page = go_to_last_visible_page(page)
    ui_log(f"[UI] real last page: {last_page}")

    for back_page in range(1, min(MAX_SCAN_PAGES, last_page) + 1):
        real_page = last_page - back_page + 1
        ui_log(f"[UI SCAN] BACK PAGE {back_page} / REAL OJS PAGE {real_page}")

        if real_page != get_last_visible_page_number(page):
            click_page_number(page, real_page)

        rows = extract_rows(page)
        ui_log(f"[UI] Page {real_page}: {len(rows)} rows. Scan bottom -> top")

        idx, row = find_first_valid_submission_row(
            rows,
            downloaded_ids=downloaded_ids,
            scan_bottom_to_top=True
        )

        if row:
            row["back_page"] = back_page
            row["real_ojs_page"] = real_page
            return real_page, idx, row

    return None, None, None

def find_start_row_by_pattern(rows):

    for i, row in enumerate(rows):

        if row.get("incomplete"):
            continue

        if row.get("status") != "Submission":
            continue

        after_rows = rows[i+1:i+5]

        review_count = sum(
            1 for r in after_rows
            if r.get("status") == "Review"
        )

        if review_count >= MIN_REVIEW_AFTER_SUBMISSION:
            return i + 1

    return None


# =========================
# WORKFLOW
# =========================

def wait_workflow_page(page):

    wait_until(
        page,
        "Workflow page",
        """
        () =>
            document.body.innerText.includes("Submission Files")
            ||
            document.querySelector("#submissionFilesGridDiv")
        """,
        timeout=90
    )


def click_submission_tab(page):

    selectors = [
        'a[href*="stageId=1"]:has-text("Submission")',
        '#ui-id-1',
        'text=Submission'
    ]

    for sel in selectors:
        try:
            loc = page.locator(sel).first

            if loc.count() > 0 and loc.is_visible():

                loc.scroll_into_view_if_needed()

                human_wait(page, 1000, 2500)

                loc.click(timeout=20000)

                human_wait(page, 3000, 6000)

                return

        except:
            pass



def extract_workflow_metadata(page, meta):
    """Ambil metadata sebenar dari workflow breadcrumb/top header.
    Contoh OJS: 26840 / Elamin / Impact of Antenna...
    """
    current = dict(meta)
    try:
        data = page.evaluate("""
        (submissionId) => {
            const clean = (s) => (s || '').replace(/\s+/g, ' ').trim();
            const selectors = [
                '.pkpWorkflow__identification',
                '.pkpWorkflow__title',
                '.pkpPageHeader',
                'header',
                'main',
                'body'
            ];
            const blocks = [];
            for (const sel of selectors) {
                for (const el of document.querySelectorAll(sel)) {
                    const t = clean(el.innerText || '');
                    if (t && t.includes(String(submissionId))) blocks.push(t);
                }
            }
            const bodyText = clean(document.body.innerText || '');
            return {blocks, bodyText};
        }
        """, str(current.get("submission_id", "")))

        sid = re.escape(str(current.get("submission_id", "")))
        blocks = data.get("blocks") or []
        if data.get("bodyText"):
            blocks.append(data.get("bodyText"))

        for block in blocks:
            # Cari pattern: 26840 / Author / Title
            m = re.search(rf"\b{sid}\b\s*/\s*([^/\n]+?)\s*/\s*(.+?)(?:\s+Activity Log|\s+Library|\s+Workflow|\s+Publication|$)", block)
            if m:
                author = m.group(1).strip()
                title = m.group(2).strip()
                if author and author.lower() not in ["submission", "view"]:
                    current["author"] = author
                if title and len(title) > 5:
                    current["title"] = title
                break
    except Exception as e:
        ui_log("[WARN] Gagal extract workflow metadata:" + str(e))

    return current

def find_article_text_file(page):

    return page.evaluate("""
    () => {

        const rows = [...document.querySelectorAll('tr.gridRow')];

        for (const row of rows) {

            const txt = row.innerText || '';

            if (!/Article Text/i.test(txt))
                continue;

            const a = row.querySelector('a[href*="download-file"]');

            if (!a)
                continue;

            const href = a.href;

            const fileIdMatch =
                href.match(/submissionFileId=(\\d+)/);

            const submissionIdMatch =
                href.match(/submissionId=(\\d+)/);

            return {
                file_id: fileIdMatch ? fileIdMatch[1] : "",
                submission_id: submissionIdMatch ? submissionIdMatch[1] : "",
                filename: (a.innerText || "").trim(),
                component: "Article Text",
                href
            };
        }

        return null;
    }
    """)


def download_article_text(page, meta):

    article = find_article_text_file(page)

    if not article:
        raise Exception("Article Text tak jumpa")

    # Pastikan metadata diambil dari workflow page dahulu.
    meta = extract_workflow_metadata(page, meta)

    submission_id = clean_filename(str(meta["submission_id"]))

    # Folder hanya guna submission number sahaja. Jangan campur author/title/status.
    folder = DOWNLOAD_DIR / submission_id
    folder.mkdir(parents=True, exist_ok=True)

    ui_log("[DOWNLOAD]" + article["filename"])

    link = page.locator(
        f'a[href*="submissionFileId={article["file_id"]}"][href*="download-file"]'
    ).first

    with page.expect_download(timeout=120000) as download_info:
        link.click(timeout=30000)

    download = download_info.value

    # Nama fail ikut nama asal dari OJS/download. Jangan rename gabung author/title.
    original_name = (article.get("filename") or "").strip() or download.suggested_filename or "article.docx"
    original_name = clean_filename(original_name)

    target_path = folder / original_name

    # Kalau nama dari UI tiada extension, guna suggested filename extension.
    if not target_path.suffix and download.suggested_filename:
        target_path = target_path.with_suffix(Path(download.suggested_filename).suffix)

    download.save_as(str(target_path))

    # Untuk Excel: kalau title dari row kosong, guna nama fail tanpa extension.
    if not (meta.get("title") or "").strip():
        meta["title"] = Path(original_name).stem

    return {
        **meta,
        "file_id": article["file_id"],
        "filename": article["filename"],
        "component": "Article Text",
        "downloaded_path": str(target_path),
        "status": "SUCCESS",
        "fail_reason": "",
        "workflow_url": page.url
    }


# =========================
# MAIN CALLABLE FUNCTION
# =========================

def run_ojs_download_workflow(log_callback=None):
    """
    Run OJS download workflow in the calling thread.

    Launches its own persistent browser session with ojs-profile.
    Login check is smart:
      - IF already logged in  → continue download automation immediately
      - IF not logged in      → log message and poll until login detected (same thread, no input())

    All Playwright operations (browser creation, navigation, download) happen
    in the SAME thread. No cross-thread object sharing.

    Returns: dict with status and message
    """
    global _workflow_running

    if log_callback:
        set_ui_log_callback(log_callback)

    # Prevent concurrent runs
    with _workflow_lock:
        if _workflow_running:
            ui_log("[OJS] ⚠️ Workflow already running. Please wait for it to complete.")
            return {"status": "error", "message": "OJS workflow already running"}
        _workflow_running = True

    _pw = None
    _browser = None

    try:
        ui_log("[OJS] Starting download workflow...")

        # --- Profile lock check ---
        # Remove stale Chromium lock file before launch to prevent ProcessSingleton errors.
        lock_file = Path(PROFILE_DIR) / "Default" / "LOCK"
        if lock_file.exists():
            ui_log("[OJS] Detected stale profile lock. Removing before launch...")
            try:
                lock_file.unlink()
                ui_log("[OJS] Stale lock removed.")
            except Exception as le:
                ui_log(f"[OJS] Warning: Could not remove lock file: {le}")

        # --- Launch browser in THIS thread ---
        # All Playwright objects created here are owned by this thread only.
        ui_log("[OJS] Launching browser with persistent OJS profile...")
        _pw = sync_playwright().start()
        _browser = _pw.chromium.launch_persistent_context(
            user_data_dir=PROFILE_DIR,
            headless=False,
            slow_mo=HUMAN_SLOW_MO,
            accept_downloads=True,
            downloads_path=str(DOWNLOAD_DIR)
        )

        page = _browser.pages[0] if _browser.pages else _browser.new_page()

        # Navigate to OJS submissions page
        ui_log("[OJS] Navigating to OJS submissions page...")
        page.goto(OJS_SUBMISSIONS_URL, wait_until="domcontentloaded", timeout=120000)

        # --- Login state check (same thread polling) ---
        current_url = page.url
        if "/login" in current_url:
            # Not logged in: log message and poll until login detected
            ui_log("[OJS] Please login manually in browser. Workflow will continue automatically.")

            max_wait = 300  # 5 minutes
            start_time = time.time()
            logged_in = False

            while time.time() - start_time < max_wait:
                try:
                    current_url = page.url
                except Exception:
                    break

                if "/login" not in current_url:
                    ui_log("[OJS] ✅ Login detected. Continuing workflow...")
                    human_wait(page, 3000, 5000)
                    logged_in = True
                    break

                time.sleep(5)

            if not logged_in:
                ui_log("[OJS] ❌ Login timeout after 5 minutes.")
                return {"status": "error", "message": "Login timeout - user did not login within 5 minutes"}
        else:
            # Already logged in: continue immediately
            ui_log("[OJS] ✅ Already logged in. Continuing workflow...")

        # ---- Existing download logic (preserved exactly) ----
        go_active_tab(page)

        downloaded_count = 0
        start_found = False

        scan_page, row_index, row = get_target_row_submission_filter(page)

        if not row:
            ui_log("[STOP] Tiada row Submission valid yang belum download dalam semua BACK PAGE filter Submission.")
            return {"status": "no_submissions", "message": "No valid submissions found to download"}
        else:
            start_found = True
            submission_id = row["submission_id"]
            author, title = parse_row_metadata(row)

            meta = {
                "submission_id": submission_id,
                "author": author,
                "title": title,
                "ojs_page": scan_page,
                "back_page": row.get("back_page", ""),
                "real_ojs_page": row.get("real_ojs_page", scan_page),
                "row_index": row_index,
            }

            ui_log("\n==============================")
            ui_log("PROCESS: " + submission_id)
            ui_log("RULE   : BACK PAGE 1 = real OJS last page, scan bottom -> top, skip Excel downloaded")
            ui_log("FILTER : Active > Stage Submission")
            ui_log("BACK PG: " + str(row.get("back_page", "")))
            ui_log("REAL PG: " + str(row.get("real_ojs_page", scan_page)))
            ui_log("ROW    : " + str(row_index))
            ui_log("STATUS : " + str(row.get("status")))
            ui_log("AUTHOR : " + author)
            ui_log("TITLE  : " + title[:100])
            ui_log("==============================")

            try:
                page.goto(row["href"], wait_until="domcontentloaded", timeout=120000)
                human_wait(page, 6000, 10000)
                wait_workflow_page(page)
                click_submission_tab(page)
                human_wait(page, 3000, 6000)
                result = download_article_text(page, meta)
                log_excel(result)
                ui_log("[SUCCESS] " + result["downloaded_path"])
                downloaded_count = 1

                page.goto(OJS_SUBMISSIONS_URL, wait_until="domcontentloaded", timeout=120000)
                human_wait(page, 4000, 7000)
                go_active_tab(page)

            except Exception as e:
                fail = {
                    **meta,
                    "file_id": "",
                    "filename": "",
                    "component": "Article Text",
                    "downloaded_path": "",
                    "status": "FAILED",
                    "fail_reason": str(e),
                    "workflow_url": page.url
                }
                log_excel(fail)
                ui_log("[FAILED] " + submission_id + " " + str(e))
                return {"status": "error", "message": f"Download failed: {str(e)}"}

        ui_log("\nDONE.")
        ui_log("Start found: " + str(start_found))
        ui_log("Downloaded: " + str(downloaded_count))
        ui_log("Folder: " + str(DOWNLOAD_DIR))
        ui_log("Excel: " + str(MASTER_XLSX))

        return {
            "status": "success",
            "message": f"Downloaded {downloaded_count} submission(s)",
            "downloaded_count": downloaded_count
        }

    except Exception as e:
        ui_log(f"[ERROR] OJS workflow failed: {str(e)}")
        return {"status": "error", "message": str(e)}
    finally:
        _workflow_running = False
        # Close browser to release the profile lock for the next run.
        # Persistent profile (cookies/session) is preserved in profiles/ojs-profile/.
        if _browser:
            try:
                _browser.close()
                ui_log("[OJS] Browser closed. Profile session preserved.")
            except Exception:
                pass
        if _pw:
            try:
                _pw.stop()
            except Exception:
                pass
