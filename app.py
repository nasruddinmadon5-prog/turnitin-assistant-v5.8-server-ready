"""
Turnitin Assistant v5.7 - Main Application
DOCX-first structure screening with AI as third checker.

© 2026 Nasruddin Madon.

Turnitin Assistant v5.7 is a proprietary workflow automation system
developed specially for Unit Jurnal, Penerbit UTM Press.

All rights reserved. Internal use only.
"""
import webview
import threading
import queue
import subprocess
import atexit
from pathlib import Path
import config
import openpyxl
from datetime import datetime
import json

# Import engine APIs
from engines import ojs_download_engine
from engines import turnitin_engine_api
from engines import ojs_report_upload_engine
from engines import template_screening_engine


# Register cleanup handler for browser sessions
def cleanup_on_exit():
    """Clean up browser sessions on app exit"""
    try:
        print("[CLEANUP] Closing Turnitin browser session...")
        turnitin_engine_api.close_turnitin_browser()
        print("[CLEANUP] ✅ Cleanup complete")
    except Exception as e:
        print(f"[CLEANUP] Error during cleanup: {e}")


atexit.register(cleanup_on_exit)

# Log queue for UI communication
log_queue = queue.Queue()


class AcademicBridgeAPI:
    """Backend API for the Turnitin Assistant dashboard"""
    
    def __init__(self):
        self.window = None
        
    def set_window(self, window):
        """Set the webview window reference for logging"""
        self.window = window
        
    def send_log_to_ui(self, message):
        """Send log message to UI terminal"""
        if self.window:
            try:
                # Escape message for JavaScript
                safe_message = message.replace("\\", "\\\\").replace("'", "\\'").replace("\n", "\\n")
                self.window.evaluate_js(f"addLog('{safe_message}', 'INFO')")
            except Exception as e:
                print(f"[LOG ERROR] Could not send to UI: {e}")
                print(f"[LOG] {message}")
        else:
            print(f"[LOG] {message}")

    def _open_existing_path(self, path, label, expected_type):
        """Open an existing file/folder in the default system application.
        Cross-platform: macOS (open), Windows (os.startfile), Linux (xdg-open).
        If no opener is available, returns the path string instead of crashing.
        """
        import platform
        target = Path(path)

        if expected_type == "file" and not target.is_file():
            return f"❌ {label} not found: {target}"
        if expected_type == "folder" and not target.is_dir():
            return f"❌ {label} folder not found: {target}"

        system = platform.system()
        try:
            if system == "Darwin":
                subprocess.run(["open", str(target)], check=True)
            elif system == "Windows":
                os.startfile(str(target))
            elif system == "Linux":
                subprocess.run(["xdg-open", str(target)], check=True)
            else:
                return f"❌ Unsupported platform: {system}. Path: {target}"
            return f"✅ Opened {label}: {target}"
        except FileNotFoundError:
            # Command not found on this system (e.g., xdg-open not installed)
            return f"ℹ️ {label}: {target}"
        except Exception as e:
            return f"❌ Failed to open {label}: {str(e)}"

    def open_excel_file(self):
        """Open the master Excel file using the default macOS app."""
        return self._open_existing_path(config.MASTER_EXCEL, "Excel file", "file")

    def open_ojs_downloads_folder(self):
        """Open the OJS downloads folder in Finder."""
        return self._open_existing_path(config.OJS_DOWNLOADS_DIR, "OJS downloads", "folder")

    def open_turnitin_reports_folder(self):
        """Open the Turnitin reports folder in Finder."""
        return self._open_existing_path(config.TURNITIN_REPORTS_DIR, "Turnitin reports", "folder")

    def open_screening_reports_folder(self):
        """Open the Template Screening reports folder in Finder."""
        screening_dir = Path(config.TEMPLATE_SCREENING_CONFIG["reports_dir"])
        return self._open_existing_path(screening_dir, "Screening reports", "folder")

    def start_ojs_download(self):
        """
        Start OJS download workflow.
        
        A single daemon thread handles the FULL lifecycle:
          - launch_persistent_context with ojs-profile
          - check login state
          - if not logged in: log message and poll until login detected
          - then run existing OJS download automation
        
        No separate login thread. No passing Playwright objects across threads.
        """
        try:
            self.send_log_to_ui("[API] Starting OJS download workflow...")
            
            def run_in_thread():
                try:
                    result = ojs_download_engine.run_ojs_download_workflow(
                        log_callback=self.send_log_to_ui
                    )
                    
                    if result["status"] == "success":
                        self.send_log_to_ui(f"[OJS] ✅ {result['message']}")
                        self.send_log_to_ui(f"[OJS] Downloaded: {result.get('downloaded_count', 0)} submission(s)")
                    elif result["status"] == "no_submissions":
                        self.send_log_to_ui(f"[OJS] ℹ️ {result['message']}")
                    else:
                        self.send_log_to_ui(f"[OJS] ❌ {result['message']}")
                        
                except Exception as e:
                    self.send_log_to_ui(f"[OJS] ❌ Error: {str(e)}")
            
            threading.Thread(target=run_in_thread, daemon=True).start()
            return "OJS download workflow started..."
            
        except Exception as e:
            return f"Error: {str(e)}"

    def start_turnitin_upload(self):
        """
        Start Turnitin upload workflow.
        
        A single daemon thread handles the FULL lifecycle:
          - launch_persistent_context with turnitin-profile
          - check login state
          - if not logged in: log message and wait for manual login
          - then run existing Turnitin upload automation
        
        No separate login thread. No passing Playwright objects across threads.
        """
        try:
            self.send_log_to_ui("[API] Starting Turnitin upload workflow...")
            
            def run_in_thread():
                try:
                    result = turnitin_engine_api.run_turnitin_upload_workflow(
                        log_callback=self.send_log_to_ui
                    )
                    
                    if result["status"] == "success":
                        self.send_log_to_ui(f"[TURNITIN] ✅ {result['message']}")
                        if result.get("submission_no"):
                            self.send_log_to_ui(f"[TURNITIN] Submission: {result['submission_no']}")
                        if result.get("report_path"):
                            self.send_log_to_ui(f"[TURNITIN] Report: {result['report_path']}")
                    else:
                        self.send_log_to_ui(f"[TURNITIN] ❌ {result['message']}")
                        
                except Exception as e:
                    self.send_log_to_ui(f"[TURNITIN] ❌ Error: {str(e)}")
            
            threading.Thread(target=run_in_thread, daemon=True).start()
            return "Turnitin upload workflow started..."
            
        except Exception as e:
            return f"Error: {str(e)}"

    def start_ojs_report_upload(self):
        """
        Start OJS Report Upload workflow (third engine).

        A single daemon thread handles the FULL lifecycle:
          - launch_persistent_context with ojs-profile
          - check login state
          - if not logged in: log message and poll until login detected
          - read pending row from sheet 'turnitin' in Excel
          - navigate to submission in OJS
          - upload Turnitin PDF report
          - log result to sheet 'upload'

        No separate login thread. No passing Playwright objects across threads.
        """
        try:
            self.send_log_to_ui("[API] Starting OJS Report Upload workflow...")

            def run_in_thread():
                try:
                    result = ojs_report_upload_engine.run_ojs_report_upload_workflow(
                        log_callback=self.send_log_to_ui
                    )

                    if result["status"] == "success":
                        self.send_log_to_ui(f"[OJS UPLOAD] ✅ {result['message']}")
                        if result.get("submission_no"):
                            self.send_log_to_ui(f"[OJS UPLOAD] Submission: {result['submission_no']}")
                        if result.get("report_path"):
                            self.send_log_to_ui(f"[OJS UPLOAD] Report: {result['report_path']}")
                    elif result["status"] == "no_submissions":
                        self.send_log_to_ui(f"[OJS UPLOAD] ℹ️ {result['message']}")
                    else:
                        self.send_log_to_ui(f"[OJS UPLOAD] ❌ {result['message']}")

                except Exception as e:
                    self.send_log_to_ui(f"[OJS UPLOAD] ❌ Error: {str(e)}")

            threading.Thread(target=run_in_thread, daemon=True).start()
            return "OJS Report Upload workflow started..."

        except Exception as e:
            return f"Error: {str(e)}"

    def start_template_screening(self):
        """
        Start Template Screening workflow.
        Checks downloaded manuscript files against required journal template.
        """
        try:
            self.send_log_to_ui("[API] Starting Template Screening workflow...")

            def run_in_thread():
                try:
                    result = template_screening_engine.run_template_screening_workflow(
                        log_callback=self.send_log_to_ui
                    )

                    if result["status"] == "success":
                        self.send_log_to_ui(f"[SCREENING] ✅ {result['message']}")
                        if result.get("screened_count", 0) > 0:
                            self.send_log_to_ui(f"[SCREENING] Screened: {result['screened_count']} file(s)")
                        if result.get("error_count", 0) > 0:
                            self.send_log_to_ui(f"[SCREENING] Errors: {result['error_count']} file(s)")
                    else:
                        self.send_log_to_ui(f"[SCREENING] ❌ {result['message']}")

                except Exception as e:
                    self.send_log_to_ui(f"[SCREENING] ❌ Error: {str(e)}")

            threading.Thread(target=run_in_thread, daemon=True).start()
            return "Template Screening workflow started..."

        except Exception as e:
            return f"Error: {str(e)}"

    def get_paper_progress(self):
        """
        Read all 3 Excel sheets (download, turnitin, upload) and return
        a JSON list of submission progress cards for the live dashboard.
        """
        try:
            excel_path = Path(config.MASTER_EXCEL)
            if not excel_path.exists():
                return json.dumps({"error": "Excel file not found"})
            
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            
            # Collect data from all sheets
            downloads = {}
            if 'download' in wb.sheetnames:
                ws = wb['download']
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[1]:
                        sub_no = str(int(row[1])) if isinstance(row[1], (int, float)) else str(row[1])
                        downloads[sub_no] = {
                            'author': row[2] if row[2] else '',
                            'title': row[3] if row[3] else '',
                            'timestamp': str(row[0]) if row[0] else '',
                        }
            
            turnitin = {}
            if 'turnitin' in wb.sheetnames:
                ws = wb['turnitin']
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[1]:
                        sub_no = str(int(row[1])) if isinstance(row[1], (int, float)) else str(row[1])
                        turnitin[sub_no] = {
                            'score': str(row[3]) if row[3] else '',
                            'status': str(row[5]) if row[5] else '',
                        }
            
            uploads = {}
            if 'upload' in wb.sheetnames:
                ws = wb['upload']
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[1]:
                        sub_no = str(int(row[1])) if isinstance(row[1], (int, float)) else str(row[1])
                        uploads[sub_no] = {
                            'status': str(row[5]) if row[5] else '',
                        }
            
            wb.close()
            
            # Build cards sorted by timestamp (latest first)
            all_subs = sorted(
                downloads.keys(),
                key=lambda k: downloads[k].get('timestamp', ''),
                reverse=True
            )
            cards = []
            
            for sn in all_subs:
                dl = downloads[sn]
                ti = turnitin.get(sn, {})
                up = uploads.get(sn, {})
                
                # Determine status for each stage
                dl_status = 'completed'
                
                ti_berjaya = ti.get('status') == 'BERJAYA'
                ti_fail = ti.get('status') == 'TAK BERJAYA DOWNLOAD'
                ti_pending = not ti
                
                up_berjaya = 'BERJAYA' in up.get('status', '')
                up_fail = 'TAK' in up.get('status', '') or 'FAIL' in up.get('status', '').upper()
                up_pending = not up
                
                # Current phase description
                if up_berjaya:
                    current_phase = f"Uploaded to OJS successfully"
                elif ti_fail:
                    current_phase = f"Turnitin Download Failed (Score: {ti.get('score', 'N/A')})" if ti.get('score') else "Turnitin Download Failed"
                elif ti_berjaya:
                    current_phase = f"Turnitin Completed (Score: {ti.get('score', 'N/A')}) - Pending Upload"
                elif dl_status == 'completed':
                    current_phase = f"Downloaded - Pending Turnitin"
                else:
                    current_phase = "Processing"
                
                card = {
                    'sub_no': f"SUB-{sn}",
                    'author': dl.get('author', ''),
                    'title': dl.get('title', ''),
                    'timestamp': dl.get('timestamp', ''),
                    'current': current_phase,
                    'download': dl_status,
                    'turnitin': 'completed' if ti_berjaya else ('error' if ti_fail else 'pending'),
                    'upload': 'completed' if up_berjaya else ('error' if up_fail else 'pending'),
                }
                cards.append(card)
            
            if not cards:
                cards.append({
                    'sub_no': '',
                    'author': '',
                    'title': '',
                    'current': 'No submissions found in Excel',
                    'download': 'pending',
                    'turnitin': 'pending',
                    'upload': 'pending'
                })
            
            return json.dumps({"cards": cards})
            
        except Exception as e:
            return json.dumps({"error": str(e)})

    def start_full_workflow(self):
        """Placeholder for full workflow - KIV for now"""
        try:
            self.send_log_to_ui("[API] Full workflow requested (KIV)...")
            self.send_log_to_ui("[FULL WORKFLOW] This feature is kept in view for future implementation.")
            self.send_log_to_ui("[FULL WORKFLOW] Please run OJS Download first, then Turnitin Upload manually.")
            
            return {
                "ojs_message": "[FULL WORKFLOW] KIV. Run OJS Download first, then Turnitin Upload manually.",
                "turnitin_message": "[FULL WORKFLOW] This feature is kept in view for future implementation."
            }
        except Exception as e:
            return {
                "ojs_message": f"Error: {str(e)}",
                "turnitin_message": f"Error: {str(e)}"
            }


def main():
    """Main application entry point"""
    # Ensure all directories exist
    config.ensure_directories()

    # Fail early with a clear message instead of letting pywebview show a blank/404 page.
    if not config.UI_INDEX.exists():
        raise FileNotFoundError(
            "Dashboard UI file not found. Expected index.html at: "
            f"{config.UI_INDEX}"
        )
    
    print("="*60)
    print(f"  Turnitin Assistant v5.7")
    print(f"  DOCX-First + AI as 3rd Checker")
    print("="*60)
    print(f"\n[INIT] Project root: {config.PROJECT_ROOT}")
    print(f"[INIT] UI path: {config.UI_INDEX}")
    print(f"[INIT] Excel path: {config.MASTER_EXCEL}")
    print(f"[INIT] OJS downloads: {config.OJS_DOWNLOADS_DIR}")
    print(f"[INIT] Turnitin reports: {config.TURNITIN_REPORTS_DIR}")
    print(f"\n[PROFILE] Turnitin: {config.TURNITIN_PROFILE_PATH}")
    print(f"[PROFILE] OJS: {config.OJS_PROFILE_PATH}")
    print("\n[INFO] v5.7: DOCX-first structure screening with AI as 3rd checker.")
    print("[INFO] AI participates early: runs when DOC/PDF conflict, not only when both FAILED.")
    print("[INFO] New status: PASS / NEEDS MANUAL REVIEW / REJECT.")
    print("[INFO] Login state persists across runs - no repeated manual login needed.")
    print("[INFO] Dashboard dynamically updates based on backend workflow state.")
    print("\n[READY] Starting dashboard...\n")
    
    # Create API instance
    api = AcademicBridgeAPI()
    
    # Create and start window
    window = webview.create_window(
        title="Turnitin Assistant v5.7",
        url=config.UI_INDEX.as_uri(),
        js_api=api,
        width=config.APP_CONFIG['window_width'],
        height=config.APP_CONFIG['window_height'],
        resizable=True
    )
    
    # Set window reference for logging
    api.set_window(window)
    
    webview.start(debug=False)


if __name__ == "__main__":
    main()